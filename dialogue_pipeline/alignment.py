from __future__ import annotations

import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

from .project import load_source_data
from .segmentation import materialize_derived_segment
from .util import (
    is_nonverbal_script,
    is_vocalization_script,
    normalize_text,
    read_json,
    resolve_model_cache_root,
    resolve_project_path,
    stable_hash,
    word_count,
    write_json,
)
from .workbook_io import lines_for_session


@dataclass
class TraceNode:
    previous: "TraceNode | None"
    action: dict[str, Any] | None


@dataclass
class StateRecord:
    score: float
    trace: TraceNode


def text_similarity(expected: str, observed: str) -> float:
    expected_normalized = normalize_text(expected)
    observed_normalized = normalize_text(observed)
    if not expected_normalized or not observed_normalized:
        return 0.0
    expected_tokens = expected_normalized.split()
    observed_tokens = observed_normalized.split()
    expected_counts = Counter(expected_tokens)
    observed_counts = Counter(observed_tokens)
    covered = sum(
        min(count, observed_counts.get(token, 0))
        for token, count in expected_counts.items()
    )
    coverage = 100.0 * covered / max(1, len(expected_tokens))
    ratio = fuzz.ratio(expected_normalized, observed_normalized)
    weighted = fuzz.WRatio(expected_normalized, observed_normalized)
    partial = fuzz.partial_ratio(expected_normalized, observed_normalized)
    token_sort = fuzz.token_sort_ratio(expected_normalized, observed_normalized)
    token_set = fuzz.token_set_ratio(expected_normalized, observed_normalized)
    compact_ratio = fuzz.ratio(
        expected_normalized.replace(" ", "").replace("'", ""),
        observed_normalized.replace(" ", "").replace("'", ""),
    )

    if len(expected_tokens) <= 3:
        score = 0.55 * ratio + 0.25 * weighted + 0.20 * coverage
    else:
        score = 0.40 * ratio + 0.30 * weighted + 0.15 * partial + 0.15 * coverage

    length_ratio = len(observed_tokens) / max(1, len(expected_tokens))
    if length_ratio < 0.45:
        score *= 0.72
    elif length_ratio < 0.70:
        score *= 0.88
    if length_ratio > 2.2:
        score *= 0.85

    order_insensitive = max(
        0.50 * token_sort + 0.30 * token_set + 0.20 * compact_ratio,
        0.60 * token_set + 0.40 * compact_ratio,
    )
    if compact_ratio < 98.0 and length_ratio < 0.45:
        order_insensitive *= 0.80
    elif compact_ratio < 98.0 and length_ratio < 0.70:
        order_insensitive *= 0.97
    if length_ratio > 2.2:
        order_insensitive *= 0.92
    score = max(score, order_insensitive)
    return max(0.0, min(100.0, score))


def transcript_fidelity(
    expected: str,
    observed: str,
    *,
    token_min_similarity: float = 78.0,
) -> dict[str, float | int]:
    """Measure ordered agreement separately from tolerant candidate retrieval."""

    expected_normalized = normalize_text(expected)
    observed_normalized = normalize_text(observed)
    expected_tokens = expected_normalized.split()
    observed_tokens = observed_normalized.split()
    if not expected_tokens or not observed_tokens:
        return {
            "ordered_similarity": 0.0,
            "token_coverage": 0.0,
            "token_precision": 0.0,
            "extra_word_count": len(observed_tokens),
        }

    possible_matches = sorted(
        (
            (
                fuzz.ratio(expected_token, observed_token),
                expected_index,
                observed_index,
            )
            for expected_index, expected_token in enumerate(expected_tokens)
            for observed_index, observed_token in enumerate(observed_tokens)
        ),
        reverse=True,
    )
    matched_expected: set[int] = set()
    matched_observed: set[int] = set()
    for score, expected_index, observed_index in possible_matches:
        if score < token_min_similarity:
            break
        if (
            expected_index in matched_expected
            or observed_index in matched_observed
        ):
            continue
        matched_expected.add(expected_index)
        matched_observed.add(observed_index)

    matched_count = len(matched_expected)
    return {
        "ordered_similarity": fuzz.ratio(
            expected_normalized,
            observed_normalized,
        ),
        "token_coverage": matched_count / len(expected_tokens),
        "token_precision": matched_count / len(observed_tokens),
        "extra_word_count": len(observed_tokens) - matched_count,
    }


def script_clauses(value: str) -> list[str]:
    return [
        clause
        for clause in (part.strip() for part in re.split(r"[.!?]+", value or ""))
        if normalize_text(clause)
    ]


def sentence_fidelity(
    expected: str,
    observed: str,
    *,
    token_min_similarity: float = 78.0,
    minimum_clause_score: float = 55.0,
) -> dict[str, Any]:
    clauses = script_clauses(expected)
    observed_normalized = normalize_text(observed)
    clause_scores = []
    for clause in clauses:
        clause_fidelity = transcript_fidelity(
            clause,
            observed,
            token_min_similarity=token_min_similarity,
        )
        partial_score = (
            fuzz.partial_ratio(normalize_text(clause), observed_normalized)
            if observed_normalized
            else 0.0
        )
        clause_scores.append(
            0.50 * partial_score
            + 50.0 * float(clause_fidelity["token_coverage"])
        )
    minimum_score = min(clause_scores) if clause_scores else 0.0
    return {
        "clause_count": len(clauses),
        "clause_scores": clause_scores,
        "minimum_clause_score": minimum_score,
        "missing_clause_count": sum(
            score < minimum_clause_score for score in clause_scores
        ),
    }


def _technical_score(segment: dict[str, Any]) -> float:
    metrics = segment.get("metrics") or {}
    score = 100.0
    if int(metrics.get("clipping_samples") or 0) > 0:
        score -= 35.0
    peak = metrics.get("peak_dbfs")
    if peak is None or not math.isfinite(float(peak)):
        score -= 40.0
    elif float(peak) > -0.05:
        score -= 10.0
    rms = metrics.get("rms_dbfs")
    if rms is None or not math.isfinite(float(rms)):
        score -= 30.0
    elif float(rms) < -45.0:
        score -= 20.0
    probability = segment.get("asr_probability")
    if probability is not None:
        score += max(-20.0, min(0.0, (float(probability) - 0.75) * 40.0))
    return max(0.0, min(100.0, score))


def _valid_span(
    segments: list[dict[str, Any]],
    start_index: int,
    count: int,
    *,
    max_merge_gap_seconds: float,
    max_span_seconds: float,
) -> bool:
    selected = segments[start_index : start_index + count]
    if len(selected) != count:
        return False
    if selected[-1]["end_seconds"] - selected[0]["start_seconds"] > max_span_seconds:
        return False
    for left, right in zip(selected, selected[1:]):
        gap = max(0.0, right["start_seconds"] - left["end_seconds"])
        if gap > max_merge_gap_seconds:
            return False
    return True


def _span_preview(
    segments: list[dict[str, Any]], start_index: int, count: int
) -> dict[str, Any]:
    selected = segments[start_index : start_index + count]
    probabilities = [
        segment["asr_probability"]
        for segment in selected
        if segment.get("asr_probability") is not None
    ]
    return {
        "start_index": start_index,
        "count": count,
        "transcript": " ".join(
            segment.get("transcript", "").strip()
            for segment in selected
            if segment.get("transcript", "").strip()
        ),
        "start_seconds": selected[0]["start_seconds"],
        "end_seconds": selected[-1]["end_seconds"],
        "asr_probability": (
            sum(probabilities) / len(probabilities) if probabilities else None
        ),
    }


def _update_state(
    states: list[dict[int, StateRecord]],
    segment_index: int,
    line_index: int,
    candidate: StateRecord,
) -> None:
    existing = states[segment_index].get(line_index)
    if existing is None or candidate.score > existing.score:
        states[segment_index][line_index] = candidate


def _nonverbal_policy(settings: dict[str, Any]) -> str:
    policy = str(settings.get("nonverbal_policy", "review")).strip().lower()
    if policy not in {"review", "skip", "weak_order"}:
        raise ValueError(
            "alignment.nonverbal_policy must be 'review', 'skip', or "
            f"'weak_order', got {policy!r}"
        )
    return policy


def _text_matchable_lines(
    lines: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[int]]:
    source_indexes = [
        index
        for index, line in enumerate(lines)
        if not is_vocalization_script(line["line"])
    ]
    return [lines[index] for index in source_indexes], source_indexes


def _restore_source_line_indexes(
    actions: list[dict[str, Any]],
    source_indexes: list[int],
) -> list[dict[str, Any]]:
    for action in actions:
        action["line_index"] = source_indexes[int(action["line_index"])]
        for match in action.get("top_matches") or []:
            match["line_index"] = source_indexes[int(match["line_index"])]
    return actions


def sequence_align(
    segments: list[dict[str, Any]],
    lines: list[dict[str, Any]],
    settings: dict[str, Any],
) -> list[dict[str, Any]]:
    lines, source_indexes = _text_matchable_lines(lines)
    segment_count = len(segments)
    line_count = len(lines)
    if not segments or not lines:
        return []
    states: list[dict[int, StateRecord]] = [
        {} for _ in range(segment_count + 1)
    ]
    root = TraceNode(previous=None, action=None)
    states[0][0] = StateRecord(score=0.0, trace=root)

    max_merge = int(settings.get("max_merge_segments", 3))
    max_gap = float(settings.get("max_merge_gap_seconds", 2.5))
    max_span = float(settings.get("max_span_seconds", 35.0))
    lookahead = int(settings.get("lookahead_lines", 8))
    minimum_path_score = float(settings.get("path_min_score", 35.0))
    noise_penalty = float(settings.get("noise_penalty", 2.2))
    skip_penalty = float(settings.get("skip_line_penalty", 1.8))
    repeat_penalty = float(settings.get("repeat_take_penalty", 0.8))

    for segment_index in range(segment_count):
        for next_line, state in list(states[segment_index].items()):
            noise_node = TraceNode(
                previous=state.trace,
                action={
                    "type": "unmatched",
                    "start_index": segment_index,
                    "count": 1,
                },
            )
            _update_state(
                states,
                segment_index + 1,
                next_line,
                StateRecord(state.score - noise_penalty, noise_node),
            )
            if next_line >= line_count:
                continue

            for count in range(1, max_merge + 1):
                if not _valid_span(
                    segments,
                    segment_index,
                    count,
                    max_merge_gap_seconds=max_gap,
                    max_span_seconds=max_span,
                ):
                    break
                span = _span_preview(segments, segment_index, count)
                last_target = min(line_count, next_line + lookahead + 1)
                for target_line in range(next_line, last_target):
                    match_score = text_similarity(
                        lines[target_line]["line"], span["transcript"]
                    )
                    if match_score < minimum_path_score:
                        continue
                    contribution = (match_score - 50.0) / 5.0
                    base_score = (
                        state.score
                        + contribution
                        - (target_line - next_line) * skip_penalty
                    )
                    action = {
                        "type": "assigned",
                        "start_index": segment_index,
                        "count": count,
                        "line_index": target_line,
                        "match_score": match_score,
                        "transcript": span["transcript"],
                    }
                    advance_node = TraceNode(previous=state.trace, action=action)
                    _update_state(
                        states,
                        segment_index + count,
                        target_line + 1,
                        StateRecord(base_score, advance_node),
                    )
                    repeat_action = dict(action)
                    repeat_action["repeat_transition"] = True
                    repeat_node = TraceNode(
                        previous=state.trace, action=repeat_action
                    )
                    _update_state(
                        states,
                        segment_index + count,
                        target_line,
                        StateRecord(base_score - repeat_penalty, repeat_node),
                    )

    if not states[segment_count]:
        return []
    _, best_state = max(
        states[segment_count].items(),
        key=lambda item: item[1].score - (line_count - item[0]) * skip_penalty,
    )
    actions = []
    node: TraceNode | None = best_state.trace
    while node and node.action is not None:
        actions.append(node.action)
        node = node.previous
    actions.reverse()
    return _restore_source_line_indexes(
        [action for action in actions if action["type"] == "assigned"],
        source_indexes,
    )


def _duration_plausibility(
    line: dict[str, Any],
    span: dict[str, Any],
) -> float:
    expected_words = max(1, word_count(line["line"]))
    expected_seconds = max(0.35, expected_words / 2.7)
    observed_seconds = max(
        0.05,
        float(span["end_seconds"]) - float(span["start_seconds"]),
    )
    duration_ratio = observed_seconds / expected_seconds
    distance = abs(math.log(max(0.05, duration_ratio)))
    return 100.0 * math.exp(-distance / 1.25)


def _order_hint(
    *,
    segment_index: int,
    segment_count: int,
    line_index: int,
    line_count: int,
) -> float:
    segment_position = segment_index / max(1, segment_count - 1)
    line_position = line_index / max(1, line_count - 1)
    return max(0.0, 1.0 - abs(segment_position - line_position))


def order_independent_align(
    segments: list[dict[str, Any]],
    lines: list[dict[str, Any]],
    settings: dict[str, Any],
) -> list[dict[str, Any]]:
    """Match non-overlapping audio spans without requiring script order.

    Each selected span has one primary line plus a small set of alternative
    line matches. Multiple separate spans may select the same line, which is
    how repeated takes are represented.
    """

    lines, source_indexes = _text_matchable_lines(lines)
    if not segments or not lines:
        return []

    segment_count = len(segments)
    line_count = len(lines)
    max_merge = int(settings.get("max_merge_segments", 3))
    max_gap = float(settings.get("max_merge_gap_seconds", 2.5))
    max_span = float(settings.get("max_span_seconds", 35.0))
    minimum_score = max(
        float(settings.get("path_min_score", 35.0)),
        float(settings.get("candidate_min_score", 45.0)),
    )
    top_k = max(1, int(settings.get("candidate_top_k", 8)))
    noise_penalty = float(settings.get("noise_penalty", 2.2))
    merge_penalty = float(settings.get("merge_span_penalty", 0.2))
    duration_weight = float(settings.get("duration_hint_weight", 1.0))
    order_weight = float(settings.get("order_hint_weight", 0.0))

    proposals_by_start: list[list[dict[str, Any]]] = [
        [] for _ in range(segment_count)
    ]
    for segment_index in range(segment_count):
        for count in range(1, max_merge + 1):
            if not _valid_span(
                segments,
                segment_index,
                count,
                max_merge_gap_seconds=max_gap,
                max_span_seconds=max_span,
            ):
                break
            span = _span_preview(segments, segment_index, count)
            line_matches = []
            for line_index, line in enumerate(lines):
                match_score = text_similarity(line["line"], span["transcript"])
                duration_score = _duration_plausibility(line, span)
                order_score = _order_hint(
                    segment_index=segment_index,
                    segment_count=segment_count,
                    line_index=line_index,
                    line_count=line_count,
                )
                ranking_score = (
                    match_score
                    + duration_weight * ((duration_score - 50.0) / 50.0)
                    + order_weight * order_score
                )
                line_matches.append(
                    {
                        "line_index": line_index,
                        "match_score": match_score,
                        "ranking_score": ranking_score,
                        "duration_plausibility": duration_score,
                        "order_hint": order_score,
                    }
                )
            line_matches.sort(
                key=lambda match: (
                    match["ranking_score"],
                    match["match_score"],
                    -match["line_index"],
                ),
                reverse=True,
            )
            pure_scores = sorted(
                (
                    match["match_score"],
                    int(match["line_index"]),
                )
                for match in line_matches
            )
            best_pure = pure_scores[-1] if pure_scores else (0.0, -1)
            second_pure = (
                pure_scores[-2] if len(pure_scores) > 1 else (0.0, -1)
            )
            for match in line_matches:
                other_score = (
                    second_pure[0]
                    if int(match["line_index"]) == best_pure[1]
                    else best_pure[0]
                )
                match["confidence_margin"] = (
                    float(match["match_score"]) - float(other_score)
                )
            primary = line_matches[0]
            if primary["match_score"] < minimum_score:
                continue
            proposal = {
                "type": "assigned",
                "start_index": segment_index,
                "count": count,
                "line_index": primary["line_index"],
                "match_score": primary["match_score"],
                "transcript": span["transcript"],
                "duration_plausibility": primary["duration_plausibility"],
                "order_hint": primary["order_hint"],
                "top_matches": [
                    match
                    for match in line_matches[:top_k]
                    if match["match_score"] >= minimum_score
                ],
            }
            proposal["path_utility"] = (
                (primary["ranking_score"] - 50.0) / 5.0
                - merge_penalty * (count - 1)
            )
            proposals_by_start[segment_index].append(proposal)

    best_scores = [0.0] * (segment_count + 1)
    choices: list[dict[str, Any] | None] = [None] * segment_count
    for segment_index in range(segment_count - 1, -1, -1):
        best_score = best_scores[segment_index + 1] - noise_penalty
        best_choice = None
        for proposal in proposals_by_start[segment_index]:
            next_index = segment_index + int(proposal["count"])
            proposal_score = float(proposal["path_utility"]) + best_scores[next_index]
            if proposal_score > best_score:
                best_score = proposal_score
                best_choice = proposal
        best_scores[segment_index] = best_score
        choices[segment_index] = best_choice

    selected = []
    segment_index = 0
    while segment_index < segment_count:
        choice = choices[segment_index]
        if choice is None:
            segment_index += 1
            continue
        selected.append(choice)
        segment_index += int(choice["count"])
    return _restore_source_line_indexes(selected, source_indexes)


def _apply_duplicate_line_policy(
    actions: list[dict[str, Any]],
    *,
    lines: list[dict[str, Any]],
    base_segments: list[dict[str, Any]],
    settings: dict[str, Any],
) -> None:
    policy = str(settings.get("duplicate_line_policy", "review")).lower()
    if policy not in {"review", "weak_order", "reuse"}:
        raise ValueError(
            "alignment.duplicate_line_policy must be 'review', "
            f"'weak_order', or 'reuse', got {policy!r}"
        )
    if policy != "weak_order":
        return

    duplicate_indexes: dict[str, list[int]] = defaultdict(list)
    for line_index, line in enumerate(lines):
        duplicate_indexes[normalize_text(line["line"])].append(line_index)
    duplicate_indexes = {
        text: indexes
        for text, indexes in duplicate_indexes.items()
        if text and len(indexes) > 1
    }
    if not duplicate_indexes:
        return

    maximum_gap = float(settings.get("take_group_gap_seconds", 12.0))
    for normalized, indexes in duplicate_indexes.items():
        matching_actions = []
        for action in actions:
            primary_index = int(action["line_index"])
            if normalize_text(lines[primary_index]["line"]) != normalized:
                continue
            matching_actions.append(action)
        matching_actions.sort(key=lambda action: int(action["start_index"]))
        if not matching_actions:
            continue

        clusters: list[list[dict[str, Any]]] = []
        previous_end = None
        for action in matching_actions:
            start_index = int(action["start_index"])
            count = int(action["count"])
            start_seconds = float(base_segments[start_index]["start_seconds"])
            end_seconds = float(
                base_segments[start_index + count - 1]["end_seconds"]
            )
            if (
                not clusters
                or previous_end is None
                or start_seconds - previous_end > maximum_gap
            ):
                clusters.append([action])
            else:
                clusters[-1].append(action)
            previous_end = end_seconds
        if len(clusters) < len(indexes):
            continue

        for cluster_index, cluster in enumerate(clusters):
            if len(clusters) == 1:
                duplicate_index = indexes[0]
            else:
                duplicate_position = round(
                    cluster_index
                    * (len(indexes) - 1)
                    / (len(clusters) - 1)
                )
                duplicate_index = indexes[duplicate_position]
            for action in cluster:
                action["line_index"] = duplicate_index
                action["duplicate_resolution"] = "weak_order"
                action["duplicate_resolved"] = True
                matches = list(action.get("top_matches") or [])
                selected = next(
                    (
                        match
                        for match in matches
                        if int(match["line_index"]) == duplicate_index
                    ),
                    None,
                )
                if selected is None:
                    selected = {
                        "line_index": duplicate_index,
                        "match_score": action["match_score"],
                        "ranking_score": action["match_score"],
                        "duration_plausibility": action.get(
                            "duration_plausibility", 0.0
                        ),
                        "order_hint": action.get("order_hint", 0.0),
                        "confidence_margin": 0.0,
                    }
                action["top_matches"] = [
                    selected,
                    *[
                        match
                        for match in matches
                        if int(match["line_index"]) != duplicate_index
                    ],
                ]


def _multisentence_fragment_join_actions(
    actions: list[dict[str, Any]],
    *,
    lines: list[dict[str, Any]],
    base_segments: list[dict[str, Any]],
    settings: dict[str, Any],
) -> list[dict[str, Any]]:
    if not bool(settings.get("fragment_join_enabled", True)):
        return []
    ordered_actions = sorted(
        actions,
        key=lambda action: (
            int(action["start_index"]),
            int(action["count"]),
        ),
    )
    maximum_actions = max(
        2,
        int(
            settings.get(
                "fragment_join_max_actions",
                settings.get("max_merge_segments", 3),
            )
        ),
    )
    maximum_segments = int(settings.get("max_merge_segments", 3))
    maximum_gap = float(settings.get("max_merge_gap_seconds", 2.5))
    maximum_span = float(settings.get("max_span_seconds", 35.0))
    minimum_match_gain = float(
        settings.get("fragment_join_min_match_gain", 5.0)
    )
    minimum_clause_gain = float(
        settings.get("fragment_join_min_clause_gain", 20.0)
    )
    minimum_ordered_score = float(
        settings.get("fragment_join_min_ordered_score", 70.0)
    )
    token_min_similarity = float(
        settings.get("fidelity_token_min_similarity", 78.0)
    )
    minimum_clause_score = float(
        settings.get("reliable_min_clause_score", 55.0)
    )
    existing = {
        (
            int(action["line_index"]),
            int(action["start_index"]),
            int(action["count"]),
        )
        for action in actions
    }
    joined = []

    for first_position, first_action in enumerate(ordered_actions):
        line_index = int(first_action["line_index"])
        line = lines[line_index]
        clauses = script_clauses(line["line"])
        if len(clauses) < 2:
            continue
        first_sentence = sentence_fidelity(
            line["line"],
            str(first_action.get("transcript") or ""),
            token_min_similarity=token_min_similarity,
            minimum_clause_score=minimum_clause_score,
        )
        first_clause_scores = list(first_sentence["clause_scores"])
        if (
            not first_clause_scores
            or first_clause_scores.index(max(first_clause_scores)) != 0
        ):
            continue

        window = [first_action]
        for next_position in range(
            first_position + 1,
            min(len(ordered_actions), first_position + maximum_actions),
        ):
            next_action = ordered_actions[next_position]
            if int(next_action["line_index"]) != line_index:
                break
            previous_action = window[-1]
            previous_end_index = (
                int(previous_action["start_index"])
                + int(previous_action["count"])
                - 1
            )
            next_start_index = int(next_action["start_index"])
            if next_start_index <= previous_end_index:
                break
            gap = max(
                0.0,
                float(base_segments[next_start_index]["start_seconds"])
                - float(base_segments[previous_end_index]["end_seconds"]),
            )
            if gap > maximum_gap:
                break
            window.append(next_action)

            start_index = int(first_action["start_index"])
            last_action = window[-1]
            end_index = (
                int(last_action["start_index"])
                + int(last_action["count"])
                - 1
            )
            count = end_index - start_index + 1
            if count > maximum_segments:
                break
            if not _valid_span(
                base_segments,
                start_index,
                count,
                max_merge_gap_seconds=maximum_gap,
                max_span_seconds=maximum_span,
            ):
                continue
            key = (line_index, start_index, count)
            if key in existing:
                continue

            span = _span_preview(base_segments, start_index, count)
            combined_match = text_similarity(
                line["line"],
                span["transcript"],
            )
            combined_fidelity = transcript_fidelity(
                line["line"],
                span["transcript"],
                token_min_similarity=token_min_similarity,
            )
            combined_sentence = sentence_fidelity(
                line["line"],
                span["transcript"],
                token_min_similarity=token_min_similarity,
                minimum_clause_score=minimum_clause_score,
            )
            fragment_match = max(
                float(action["match_score"]) for action in window
            )
            fragment_clause_score = max(
                float(
                    sentence_fidelity(
                        line["line"],
                        str(action.get("transcript") or ""),
                        token_min_similarity=token_min_similarity,
                        minimum_clause_score=minimum_clause_score,
                    )["minimum_clause_score"]
                )
                for action in window
            )
            if combined_match - fragment_match < minimum_match_gain:
                continue
            if (
                float(combined_sentence["minimum_clause_score"])
                - fragment_clause_score
                < minimum_clause_gain
            ):
                continue
            if combined_sentence["missing_clause_count"] > 0:
                continue
            if (
                float(combined_fidelity["ordered_similarity"])
                < minimum_ordered_score
            ):
                continue

            other_score = max(
                (
                    text_similarity(other_line["line"], span["transcript"])
                    for other_index, other_line in enumerate(lines)
                    if other_index != line_index
                    and not is_vocalization_script(other_line["line"])
                ),
                default=0.0,
            )
            duration_score = _duration_plausibility(line, span)
            order_score = sum(
                float(action.get("order_hint", 0.0))
                for action in window
            ) / len(window)
            joined.append(
                {
                    "type": "assigned",
                    "start_index": start_index,
                    "count": count,
                    "line_index": line_index,
                    "match_score": combined_match,
                    "transcript": span["transcript"],
                    "duration_plausibility": duration_score,
                    "order_hint": order_score,
                    "top_matches": [
                        {
                            "line_index": line_index,
                            "match_score": combined_match,
                            "ranking_score": combined_match,
                            "duration_plausibility": duration_score,
                            "order_hint": order_score,
                            "confidence_margin": combined_match - other_score,
                        }
                    ],
                    "fragment_join": True,
                    "fragment_source_count": len(window),
                }
            )
            existing.add(key)
            break
    return joined


def _expand_alignment_actions(
    actions: list[dict[str, Any]],
    *,
    base_segments: list[dict[str, Any]],
    session_id: str,
    settings: dict[str, Any],
) -> list[dict[str, Any]]:
    take_group_gap = float(settings.get("take_group_gap_seconds", 12.0))
    group_number = 0
    previous_primary: dict[str, Any] | None = None
    expanded = []

    for action in actions:
        primary_line_index = int(action["line_index"])
        start_index = int(action["start_index"])
        count = int(action["count"])
        start_seconds = float(base_segments[start_index]["start_seconds"])
        end_seconds = float(
            base_segments[start_index + count - 1]["end_seconds"]
        )
        if (
            previous_primary is not None
            and int(previous_primary["line_index"]) == primary_line_index
            and start_seconds - float(previous_primary["end_seconds"])
            <= take_group_gap
        ):
            take_group_id = str(previous_primary["take_group_id"])
        else:
            group_number += 1
            take_group_id = f"{session_id}__g{group_number:05d}"
        previous_primary = {
            "line_index": primary_line_index,
            "end_seconds": end_seconds,
            "take_group_id": take_group_id,
        }

        matches = action.get("top_matches") or [
            {
                "line_index": primary_line_index,
                "match_score": action["match_score"],
                "duration_plausibility": action.get(
                    "duration_plausibility", 0.0
                ),
                "order_hint": action.get("order_hint", 0.0),
            }
        ]
        for match_rank, match in enumerate(matches, start=1):
            expanded.append(
                {
                    "type": "assigned",
                    "start_index": start_index,
                    "count": count,
                    "line_index": int(match["line_index"]),
                    "primary_line_index": primary_line_index,
                    "match_score": float(match["match_score"]),
                    "transcript": action["transcript"],
                    "duration_plausibility": float(
                        match.get("duration_plausibility", 0.0)
                    ),
                    "order_hint": float(match.get("order_hint", 0.0)),
                    "confidence_margin": float(
                        match.get("confidence_margin", 0.0)
                    ),
                    "segment_match_rank": match_rank,
                    "is_primary_match": match_rank == 1,
                    "take_group_id": (
                        take_group_id if match_rank == 1 else None
                    ),
                    "duplicate_resolution": action.get(
                        "duplicate_resolution"
                    ),
                    "duplicate_resolved": bool(
                        action.get("duplicate_resolved", False)
                        and match_rank == 1
                    ),
                    "fragment_join": bool(
                        action.get("fragment_join", False)
                        and match_rank == 1
                    ),
                    "fragment_source_count": (
                        int(action.get("fragment_source_count", 0))
                        if match_rank == 1
                        else 0
                    ),
                }
            )
    return expanded


def _vocalization_key(value: str) -> str:
    normalized = re.sub(r"[^a-z]+", "", normalize_text(value))
    normalized = re.sub(r"(.)\1+", r"\1", normalized)
    return normalized.replace("gh", "h")


def _vocalization_review_actions(
    *,
    base_segments: list[dict[str, Any]],
    lines: list[dict[str, Any]],
    session_id: str,
    settings: dict[str, Any],
    existing_actions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if _nonverbal_policy(settings) != "weak_order":
        return []
    vocalization = dict(settings.get("vocalization_alignment") or {})
    if not bool(vocalization.get("enabled", True)):
        return []

    line_indexes = [
        index
        for index, line in enumerate(lines)
        if is_vocalization_script(line["line"])
    ]
    if not line_indexes:
        return []
    maximum_duration = float(vocalization.get("max_segment_seconds", 4.0))
    maximum_words = int(vocalization.get("max_transcript_words", 3))
    candidates_per_line = max(
        1,
        int(vocalization.get("candidates_per_line", 3)),
    )
    sequence_weight = float(vocalization.get("weak_order_weight", 4.0))
    eligible_segments = [
        (index, segment)
        for index, segment in enumerate(base_segments)
        if (
            float(segment.get("end_seconds", 0.0))
            - float(segment.get("start_seconds", 0.0))
            <= maximum_duration
            and int(
                segment.get("word_count")
                or word_count(str(segment.get("transcript") or ""))
            )
            <= maximum_words
        )
    ]
    if not eligible_segments:
        return []

    existing = {
        (
            int(action["line_index"]),
            int(action["start_index"]),
            int(action["count"]),
        )
        for action in existing_actions
    }
    actions = []
    for line_index in line_indexes:
        line = lines[line_index]
        expected_key = _vocalization_key(line["line"])
        matches = []
        for segment_index, segment in eligible_segments:
            transcript = str(segment.get("transcript") or "")
            observed_key = _vocalization_key(transcript)
            phonetic_score = (
                max(
                    fuzz.ratio(expected_key, observed_key),
                    fuzz.partial_ratio(expected_key, observed_key),
                )
                if expected_key and observed_key
                else 0.0
            )
            text_score = text_similarity(line["line"], transcript)
            order_score = _order_hint(
                segment_index=segment_index,
                segment_count=len(base_segments),
                line_index=line_index,
                line_count=len(lines),
            )
            score = min(
                100.0,
                max(text_score, 0.85 * phonetic_score)
                + sequence_weight * order_score,
            )
            matches.append((score, segment_index, transcript, order_score))
        matches.sort(reverse=True)
        for match_rank, (
            score,
            segment_index,
            transcript,
            order_score,
        ) in enumerate(matches[:candidates_per_line], start=1):
            key = (line_index, segment_index, 1)
            if key in existing:
                continue
            actions.append(
                {
                    "type": "assigned",
                    "start_index": segment_index,
                    "count": 1,
                    "line_index": line_index,
                    "primary_line_index": line_index,
                    "match_score": score,
                    "transcript": transcript,
                    "duration_plausibility": 0.0,
                    "order_hint": order_score,
                    "confidence_margin": 0.0,
                    "segment_match_rank": match_rank,
                    "is_primary_match": False,
                    "take_group_id": (
                        f"{session_id}__v{line_index + 1:05d}"
                    ),
                    "force_candidate": True,
                    "forced_review_reason": "VOCALIZATION_REVIEW",
                    "duplicate_resolution": None,
                    "duplicate_resolved": False,
                }
            )
    return actions


def _candidate_reliability(
    *,
    line: dict[str, Any],
    match_score: float,
    margin: float,
    settings: dict[str, Any],
    observed: str | None = None,
    duplicate_text: bool = False,
    unsafe_untranscribed_merge: bool = False,
    duration_plausibility: float | None = None,
) -> tuple[bool, str]:
    if is_nonverbal_script(line["line"]):
        return False, "NONVERBAL_SCRIPT"
    if unsafe_untranscribed_merge:
        return False, "MERGED_UNTRANSCRIBED_AUDIO"
    if (
        duration_plausibility is not None
        and duration_plausibility
        < float(settings.get("reliable_min_duration_plausibility", 25.0))
    ):
        return False, "POSSIBLE_REPEATED_TAKES"
    if (
        observed
        and not duplicate_text
        and normalize_text(line["line"]) == normalize_text(observed)
    ):
        return True, ""
    fidelity = transcript_fidelity(
        line["line"],
        observed or "",
        token_min_similarity=float(
            settings.get("fidelity_token_min_similarity", 78.0)
        ),
    )
    sentence = sentence_fidelity(
        line["line"],
        observed or "",
        token_min_similarity=float(
            settings.get("fidelity_token_min_similarity", 78.0)
        ),
        minimum_clause_score=float(
            settings.get("reliable_min_clause_score", 55.0)
        ),
    )
    is_short_line = word_count(line["line"]) <= 3
    if is_short_line:
        minimum_score = float(settings.get("short_line_min_score", 88.0))
        minimum_margin = float(settings.get("short_line_min_margin", 15.0))
        if match_score < minimum_score:
            return False, "SHORT_LINE_LOW_SCORE"
        if margin < minimum_margin:
            return False, "SHORT_LINE_AMBIGUOUS"
        if (
            sentence["clause_count"] >= 2
            and sentence["missing_clause_count"] > 0
        ):
            return False, "MISSING_SENTENCE"
        if fidelity["token_coverage"] < float(
            settings.get("short_line_min_token_coverage", 1.0)
        ):
            return False, "SHORT_LINE_INCOMPLETE_TRANSCRIPT"
        if fidelity["token_precision"] < float(
            settings.get("short_line_min_token_precision", 1.0)
        ):
            return False, "SHORT_LINE_EXTRA_WORDS"
        if fidelity["ordered_similarity"] < float(
            settings.get("short_line_min_ordered_score", 70.0)
        ):
            return False, "SHORT_LINE_ORDER_MISMATCH"
        return True, ""
    if match_score < float(settings.get("reliable_min_score", 72.0)):
        return False, "LOW_MATCH_SCORE"
    if margin < float(settings.get("reliable_min_margin", 8.0)):
        return False, "AMBIGUOUS_MATCH"
    if (
        sentence["clause_count"] >= 2
        and sentence["missing_clause_count"] > 0
    ):
        return False, "MISSING_SENTENCE"
    if fidelity["token_coverage"] < float(
        settings.get("reliable_min_token_coverage", 0.60)
    ):
        return False, "INCOMPLETE_TRANSCRIPT"
    if fidelity["token_precision"] < float(
        settings.get("reliable_min_token_precision", 0.70)
    ):
        return False, "EXCESS_TRANSCRIPT_WORDS"
    if fidelity["ordered_similarity"] < float(
        settings.get("reliable_min_ordered_score", 55.0)
    ):
        return False, "LOW_ORDERED_SIMILARITY"
    return True, ""


def _has_unsafe_untranscribed_merge(
    *,
    action: dict[str, Any],
    base_segments: list[dict[str, Any]],
    settings: dict[str, Any],
) -> bool:
    if not bool(settings.get("auto_reject_untranscribed_merge", True)):
        return False
    start_index = int(action["start_index"])
    count = int(action["count"])
    if count <= 1:
        return False
    minimum_seconds = float(
        settings.get("untranscribed_merge_min_seconds", 0.5)
    )
    minimum_rms = float(
        settings.get("untranscribed_merge_min_rms_dbfs", -45.0)
    )
    for segment in base_segments[start_index : start_index + count]:
        if str(segment.get("transcript") or "").strip():
            continue
        duration = float(
            (segment.get("metrics") or {}).get("duration_seconds")
            or (
                float(segment.get("end_seconds", 0.0))
                - float(segment.get("start_seconds", 0.0))
            )
        )
        rms = (segment.get("metrics") or {}).get("rms_dbfs")
        if (
            duration >= minimum_seconds
            and (rms is None or float(rms) >= minimum_rms)
        ):
            return True
    return False


def _apply_local_asr_rescue(
    *,
    project_dir: Path,
    project: dict[str, Any],
    session: dict[str, Any],
    session_lines: list[dict[str, Any]],
    base_segments: list[dict[str, Any]],
    settings: dict[str, Any],
    runtime: dict[str, Any],
) -> int:
    rescue = dict(settings.get("local_asr_rescue") or {})
    if not bool(rescue.get("enabled", False)):
        return 0

    eligible_lines = [
        line
        for line in session_lines
        if not is_vocalization_script(line["line"])
        and word_count(line["line"])
        <= int(rescue.get("max_candidate_words", 8))
    ]
    if not eligible_lines:
        return 0

    maximum_segments = max(0, int(rescue.get("max_segments_per_session", 80)))
    maximum_duration = float(rescue.get("max_segment_seconds", 6.0))
    maximum_words = int(rescue.get("max_observed_words", 6))
    trigger_score = float(rescue.get("trigger_score", 88.0))
    trigger_probability = float(rescue.get("trigger_probability", 0.65))
    minimum_probability = float(rescue.get("minimum_probability", 0.55))
    minimum_gain = float(rescue.get("minimum_score_gain", 6.0))
    minimum_score = float(rescue.get("minimum_score", 72.0))
    prompt_top_k = max(1, int(rescue.get("prompt_top_k", 8)))
    prompt_characters = max(100, int(rescue.get("prompt_characters", 800)))

    transcription_settings = dict(project.get("transcription") or {})
    model_name = str(
        rescue.get("model")
        or transcription_settings.get("model")
        or "large-v3"
    )
    device = str(
        rescue.get("device")
        or transcription_settings.get("device")
        or "auto"
    )
    compute_type = str(transcription_settings.get("compute_type", "auto"))
    model_root = resolve_model_cache_root(project_dir, transcription_settings)
    model_root.mkdir(parents=True, exist_ok=True)

    attempted = 0
    accepted = 0
    for segment in base_segments:
        if attempted >= maximum_segments:
            break
        transcript = str(segment.get("transcript") or "").strip()
        if not transcript:
            continue
        duration = float(segment.get("end_seconds", 0.0)) - float(
            segment.get("start_seconds", 0.0)
        )
        if duration > maximum_duration:
            continue
        observed_words = int(segment.get("word_count") or word_count(transcript))
        probability = segment.get("asr_probability")
        scored_lines = sorted(
            [
                (
                    text_similarity(line["line"], transcript),
                    line,
                )
                for line in eligible_lines
            ],
            key=lambda item: item[0],
        )
        old_score = scored_lines[-1][0] if scored_lines else 0.0
        if (
            observed_words > maximum_words
            and old_score >= trigger_score
            and probability is not None
            and float(probability) >= trigger_probability
        ):
            continue
        if (
            old_score >= trigger_score
            and probability is not None
            and float(probability) >= trigger_probability
        ):
            continue

        candidates = [
            line for _, line in reversed(scored_lines[-prompt_top_k:])
        ]
        cache_key = stable_hash(
            {
                "segment_id": segment["segment_id"],
                "source_sha256": segment.get("source_sha256"),
                "transcript": transcript,
                "model": model_name,
                "device": device,
                "compute_type": compute_type,
                "settings": rescue,
                "candidate_line_ids": [line["line_id"] for line in candidates],
            }
        )
        previous = segment.get("local_asr_rescue") or {}
        if previous.get("cache_key") == cache_key:
            continue

        if runtime.get("model") is None:
            from .transcription import _make_model

            print(
                f"[local ASR rescue] Loading {model_name} from {model_root}",
                flush=True,
            )
            model, resolved_device, resolved_compute = _make_model(
                model_name,
                device=device,
                compute_type=compute_type,
                download_root=model_root,
            )
            runtime.update(
                {
                    "model": model,
                    "device": resolved_device,
                    "compute_type": resolved_compute,
                    "model_name": model_name,
                }
            )

        prompt = " | ".join(line["line"] for line in candidates)[
            :prompt_characters
        ]
        segment_path = resolve_project_path(project_dir, segment["file"])
        attempted += 1
        try:
            segments_iter, _ = runtime["model"].transcribe(
                str(segment_path),
                language=project.get("language", "en"),
                beam_size=int(rescue.get("beam_size", 5)),
                word_timestamps=True,
                condition_on_previous_text=False,
                vad_filter=False,
                initial_prompt=prompt,
                hotwords=prompt,
            )
            rescue_parts = []
            rescue_probabilities = []
            for rescue_segment in segments_iter:
                text = str(rescue_segment.text or "").strip()
                if text:
                    rescue_parts.append(text)
                for rescue_word in rescue_segment.words or []:
                    if rescue_word.probability is not None:
                        rescue_probabilities.append(
                            float(rescue_word.probability)
                        )
            rescued_transcript = " ".join(rescue_parts).strip()
            rescued_probability = (
                sum(rescue_probabilities) / len(rescue_probabilities)
                if rescue_probabilities
                else None
            )
            new_scores = sorted(
                (
                    text_similarity(line["line"], rescued_transcript),
                    line["line_id"],
                )
                for line in eligible_lines
            )
            new_score = new_scores[-1][0] if new_scores else 0.0
            accepted_rescue = bool(
                rescued_transcript
                and rescued_probability is not None
                and rescued_probability >= minimum_probability
                and new_score >= minimum_score
                and new_score >= old_score + minimum_gain
            )
            segment["local_asr_rescue"] = {
                "cache_key": cache_key,
                "accepted": accepted_rescue,
                "transcript": rescued_transcript,
                "asr_probability": rescued_probability,
                "old_best_score": old_score,
                "new_best_score": new_score,
                "candidate_line_ids": [
                    line["line_id"] for line in candidates
                ],
                "model": model_name,
                "device": runtime.get("device"),
                "compute_type": runtime.get("compute_type"),
            }
            if accepted_rescue:
                segment.setdefault("original_transcript", transcript)
                segment.setdefault(
                    "original_asr_probability",
                    segment.get("asr_probability"),
                )
                segment["transcript"] = rescued_transcript
                segment["word_count"] = word_count(rescued_transcript)
                segment["asr_probability"] = rescued_probability
                segment["transcript_source"] = "local_asr_rescue"
                accepted += 1
        except Exception as error:
            segment["local_asr_rescue"] = {
                "cache_key": cache_key,
                "accepted": False,
                "error": str(error),
                "candidate_line_ids": [
                    line["line_id"] for line in candidates
                ],
                "model": model_name,
            }
            print(
                f"[local ASR rescue] {session['id']} "
                f"{segment['segment_id']} failed: {error}",
                flush=True,
            )
    if attempted:
        print(
            f"[local ASR rescue] {session['id']}: "
            f"{accepted}/{attempted} improved",
            flush=True,
        )
    return accepted


def align_project(
    *,
    project_dir: Path,
    project: dict[str, Any],
    session_filter: set[str] | None = None,
) -> dict[str, Path]:
    manifest_path = project_dir / "segments_manifest.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(
            f"Missing segment manifest: {manifest_path}. Run segment first."
        )
    manifest = read_json(manifest_path)
    source_data = load_source_data(project_dir, project)
    line_by_id = {line["line_id"]: line for line in source_data["lines"]}
    settings = dict(project.get("alignment") or {})
    nonverbal_policy = _nonverbal_policy(settings)
    alignment_mode = str(settings.get("mode", "unordered")).strip().lower()
    if alignment_mode not in {"unordered", "sequence"}:
        raise ValueError(
            "alignment.mode must be either 'unordered' or 'sequence', "
            f"got {alignment_mode!r}"
        )
    manifest_session_by_id = {
        entry["session_id"]: entry for entry in manifest["sessions"]
    }

    candidates: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []
    alignment_sessions = []
    rescue_runtime: dict[str, Any] = {}

    selected_sessions = [
        item
        for item in project["sessions"]
        if item.get("enabled", True)
        and (not session_filter or item["id"] in session_filter)
    ]
    if not selected_sessions:
        raise ValueError("No enabled sessions matched the requested filter.")
    for session_index, session in enumerate(selected_sessions, start=1):
        session_entry = manifest_session_by_id.get(session["id"])
        if not session_entry:
            raise KeyError(f"Segments missing for session: {session['id']}")
        base_segments = session_entry.get("segments", [])
        session_lines = lines_for_session(source_data, session)
        if not session_lines:
            continue
        text_matchable_session_lines = [
            line
            for line in session_lines
            if not is_vocalization_script(line["line"])
        ]
        normalized_line_counts = Counter(
            normalize_text(line["line"]) for line in session_lines
        )
        _apply_local_asr_rescue(
            project_dir=project_dir,
            project=project,
            session=session,
            session_lines=session_lines,
            base_segments=base_segments,
            settings=settings,
            runtime=rescue_runtime,
        )
        print(
            f"[align {session_index}] {session['id']}: "
            f"{len(base_segments)} segments → {len(session_lines)} lines "
            f"({alignment_mode})",
            flush=True,
        )
        if alignment_mode == "sequence":
            primary_actions = sequence_align(
                base_segments,
                session_lines,
                settings,
            )
        else:
            primary_actions = order_independent_align(
                base_segments,
                session_lines,
                settings,
            )
        _apply_duplicate_line_policy(
            primary_actions,
            lines=session_lines,
            base_segments=base_segments,
            settings=settings,
        )
        fragment_join_actions = _multisentence_fragment_join_actions(
            primary_actions,
            lines=session_lines,
            base_segments=base_segments,
            settings=settings,
        )
        candidate_actions = sorted(
            [*primary_actions, *fragment_join_actions],
            key=lambda action: (
                int(action["start_index"]),
                int(action["count"]),
            ),
        )
        actions = _expand_alignment_actions(
            candidate_actions,
            base_segments=base_segments,
            session_id=session["id"],
            settings=settings,
        )
        actions.extend(
            _vocalization_review_actions(
                base_segments=base_segments,
                lines=session_lines,
                session_id=session["id"],
                settings=settings,
                existing_actions=actions,
            )
        )
        reliable_coverage: set[int] = set()
        assignment_by_base: dict[int, list[dict[str, Any]]] = defaultdict(list)
        serialized_actions = []

        for action in actions:
            line = session_lines[action["line_index"]]
            segment = materialize_derived_segment(
                project_dir=project_dir,
                project=project,
                session_entry=session_entry,
                base_segments=base_segments,
                start_index=action["start_index"],
                count=action["count"],
            )
            if "confidence_margin" in action:
                margin = float(action["confidence_margin"])
            else:
                all_scores = sorted(
                    (
                        text_similarity(
                            other_line["line"],
                            segment["transcript"],
                        ),
                        other_line["line_id"],
                    )
                    for other_line in session_lines
                    if other_line["line_id"] != line["line_id"]
                )
                second_score = all_scores[-1][0] if all_scores else 0.0
                margin = action["match_score"] - second_score
            duplicate_policy = str(
                settings.get("duplicate_line_policy", "review")
            ).lower()
            primary_line = session_lines[
                int(action.get("primary_line_index", action["line_index"]))
            ]
            reusable_duplicate = bool(
                duplicate_policy == "reuse"
                and normalize_text(primary_line["line"])
                == normalize_text(line["line"])
            )
            fidelity = transcript_fidelity(
                line["line"],
                segment.get("transcript", ""),
                token_min_similarity=float(
                    settings.get("fidelity_token_min_similarity", 78.0)
                ),
            )
            sentence = sentence_fidelity(
                line["line"],
                segment.get("transcript", ""),
                token_min_similarity=float(
                    settings.get("fidelity_token_min_similarity", 78.0)
                ),
                minimum_clause_score=float(
                    settings.get("reliable_min_clause_score", 55.0)
                ),
            )
            unsafe_untranscribed_merge = _has_unsafe_untranscribed_merge(
                action=action,
                base_segments=base_segments,
                settings=settings,
            )
            reliable, reason = _candidate_reliability(
                line=line,
                match_score=action["match_score"],
                margin=margin,
                settings=settings,
                observed=segment.get("transcript", ""),
                duplicate_text=(
                    normalized_line_counts[normalize_text(line["line"])] > 1
                    and not action.get("duplicate_resolved", False)
                    and not reusable_duplicate
                ),
                unsafe_untranscribed_merge=unsafe_untranscribed_merge,
                duration_plausibility=(
                    float(action["duration_plausibility"])
                    if action.get("duration_plausibility") is not None
                    else None
                ),
            )
            if (
                not action.get("is_primary_match", True)
                and not reusable_duplicate
            ):
                reliable = False
                reason = "SEGMENT_BETTER_MATCH_ELSEWHERE"
            if action.get("forced_review_reason"):
                reliable = False
                reason = str(action["forced_review_reason"])
            technical_score = _technical_score(segment)
            selection_score = (
                action["match_score"]
                + 0.10 * technical_score
                + 5.0 * float(segment.get("asr_probability") or 0.0)
                + float(
                    settings.get("clause_completeness_weight", 5.0)
                )
                * (
                    float(sentence["minimum_clause_score"]) / 100.0
                    if sentence["clause_count"] >= 2
                    else 1.0
                )
                + (
                    float(settings.get("primary_match_bonus", 2.0))
                    if action.get("is_primary_match", True)
                    else 0.0
                )
            )
            candidate = {
                "line_id": line["line_id"],
                "session_id": session["id"],
                "segment_id": segment["segment_id"],
                "segment_file": segment["file"],
                "transcript": segment.get("transcript", ""),
                "match_score": action["match_score"],
                "ordered_similarity": fidelity["ordered_similarity"],
                "token_coverage": fidelity["token_coverage"],
                "token_precision": fidelity["token_precision"],
                "extra_word_count": fidelity["extra_word_count"],
                "clause_count": sentence["clause_count"],
                "minimum_clause_score": sentence["minimum_clause_score"],
                "missing_clause_count": sentence["missing_clause_count"],
                "confidence_margin": margin,
                "technical_score": technical_score,
                "selection_score": selection_score,
                "reliable": reliable,
                "reliability_reason": reason,
                "source_audio": segment["source_audio"],
                "start_seconds": segment["start_seconds"],
                "end_seconds": segment["end_seconds"],
                "duration_seconds": segment["metrics"]["duration_seconds"],
                "asr_probability": segment.get("asr_probability"),
                "base_indices": segment["base_indices"],
                "alignment_mode": alignment_mode,
                "is_primary_match": bool(
                    action.get("is_primary_match", True)
                ),
                "segment_match_rank": int(
                    action.get("segment_match_rank", 1)
                ),
                "take_group_id": action.get("take_group_id"),
                "duration_plausibility": float(
                    action.get("duration_plausibility", 0.0)
                ),
                "order_hint": float(action.get("order_hint", 0.0)),
                "duplicate_resolution": action.get("duplicate_resolution"),
                "transcript_source": segment.get(
                    "transcript_source", "session_asr"
                ),
                "local_asr_rescued": bool(
                    (segment.get("local_asr_rescue") or {}).get("accepted")
                ),
                "unsafe_untranscribed_merge": unsafe_untranscribed_merge,
                "fragment_join": bool(action.get("fragment_join", False)),
                "fragment_source_count": int(
                    action.get("fragment_source_count", 0)
                ),
            }
            if (
                action["match_score"]
                >= float(settings.get("candidate_min_score", 45.0))
                or action.get("force_candidate", False)
            ):
                candidates.append(candidate)
            for base_index in segment["base_indices"]:
                assignment_by_base[base_index].append(candidate)
                if reliable:
                    reliable_coverage.add(base_index)
            serialized_actions.append(candidate)

        for base_index, segment in enumerate(base_segments):
            if base_index in reliable_coverage:
                continue
            transcript = segment.get("transcript", "")
            suggestions = sorted(
                (
                    text_similarity(line["line"], transcript),
                    line["line_id"],
                    line["line"],
                )
                for line in text_matchable_session_lines
            )
            best = suggestions[-1] if suggestions else (0.0, "", "")
            second = suggestions[-2] if len(suggestions) > 1 else (0.0, "", "")
            related = assignment_by_base.get(base_index, [])
            if not transcript.strip():
                reason = "TRANSCRIPTION_FAILED_OR_NONVERBAL"
            elif related:
                reason = "AMBIGUOUS_OR_LOW_CONFIDENCE"
            else:
                reason = "NO_RELIABLE_MATCH"
            unmatched_rows.append(
                {
                    "segment_id": segment["segment_id"],
                    "segment_file": segment["file"],
                    "source_wav": segment["source_audio"],
                    "start_seconds": segment["start_seconds"],
                    "end_seconds": segment["end_seconds"],
                    "duration_seconds": segment["metrics"]["duration_seconds"],
                    "transcript": transcript,
                    "asr_confidence": segment.get("asr_probability"),
                    "reason": reason,
                    "suggested_line_1": best[1],
                    "suggested_line_1_text": best[2],
                    "suggested_line_1_score": best[0],
                    "suggested_line_2": second[1],
                    "suggested_line_2_text": second[2],
                    "suggested_line_2_score": second[0],
                    "technical_flags": _technical_flags(segment),
                }
            )

        alignment_sessions.append(
            {
                "session_id": session["id"],
                "alignment_mode": alignment_mode,
                "nonverbal_policy": nonverbal_policy,
                "script_line_count": len(session_lines),
                "base_segment_count": len(base_segments),
                "assignments": serialized_actions,
                "reliable_base_segment_count": len(reliable_coverage),
            }
        )

    by_line: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for candidate in candidates:
        by_line[candidate["line_id"]].append(candidate)
    for line_candidates in by_line.values():
        line_candidates.sort(
            key=lambda candidate: candidate["selection_score"], reverse=True
        )
        for rank, candidate in enumerate(line_candidates, start=1):
            candidate["rank"] = rank

    alignment_payload = {
        "schema_version": 1,
        "nonverbal_policy": nonverbal_policy,
        "sessions": alignment_sessions,
        "candidate_count": len(candidates),
        "unmatched_count": len(unmatched_rows),
        "unmatched_segments": unmatched_rows,
        "candidates": sorted(
            candidates,
            key=lambda candidate: (
                line_by_id[candidate["line_id"]]["sheet_index"],
                line_by_id[candidate["line_id"]]["excel_row"],
                candidate["rank"],
            ),
        ),
    }
    alignment_path = project_dir / "alignment.json"
    write_json(alignment_path, alignment_payload)
    write_json(manifest_path, manifest)

    review_path = project_dir / "A_line_review.xlsx"
    _write_review_workbook(
        review_path=review_path,
        project_dir=project_dir,
        source_lines=source_data["lines"],
        candidates_by_line=by_line,
        unmatched_rows=unmatched_rows,
        nonverbal_policy=nonverbal_policy,
    )
    (project_dir / "B_unmatched_segments.tsv").unlink(missing_ok=True)
    return {
        "review": review_path,
        "alignment": alignment_path,
        "manifest": manifest_path,
    }


def _technical_flags(segment: dict[str, Any]) -> str:
    flags = []
    metrics = segment.get("metrics") or {}
    if int(metrics.get("clipping_samples") or 0) > 0:
        flags.append("CLIPPING")
    if float(metrics.get("rms_dbfs") or -999.0) < -45.0:
        flags.append("VERY_QUIET")
    if not segment.get("transcript", "").strip():
        flags.append("NO_TRANSCRIPT")
    return ",".join(flags)


def _write_review_workbook(
    *,
    review_path: Path,
    project_dir: Path,
    source_lines: list[dict[str, Any]],
    candidates_by_line: dict[str, list[dict[str, Any]]],
    unmatched_rows: list[dict[str, Any]],
    nonverbal_policy: str = "review",
) -> None:
    nonverbal_policy = _nonverbal_policy(
        {"nonverbal_policy": nonverbal_policy}
    )
    workbook = Workbook()
    lines_sheet = workbook.active
    lines_sheet.title = "Lines"
    candidates_sheet = workbook.create_sheet("Candidates")
    unmatched_sheet = workbook.create_sheet("Unmatched Segments")
    instructions_sheet = workbook.create_sheet("Instructions")

    line_headers = [
        "Line ID",
        "Sheet",
        "Excel Row",
        "Quest",
        "Context",
        "Line to Speak",
        "Acting Note",
        "Facial Emotion",
        "Target Filename",
        "Candidate Count",
        "Candidate 1",
        "Candidate 2",
        "Candidate 3",
        "Suggested Best Segment",
        "Selected Segment",
        "Status",
        "Candidate Summary",
        "User Notes",
    ]
    lines_sheet.append(line_headers)
    for line in source_lines:
        is_nonverbal = is_vocalization_script(line["line"])
        line_candidates = candidates_by_line.get(line["line_id"], [])
        best = line_candidates[0] if line_candidates else None
        reliable_best = next(
            (candidate for candidate in line_candidates if candidate["reliable"]),
            None,
        )
        if is_nonverbal:
            status = (
                "SKIP"
                if nonverbal_policy == "skip"
                else "NONVERBAL_REVIEW"
            )
        else:
            status = (
                "AUTO_OK"
                if reliable_best
                else ("REVIEW" if line_candidates else "MISSING")
            )
        suggested = best["segment_id"] if best else ""
        selected = reliable_best["segment_id"] if reliable_best else ""
        top_candidates = line_candidates[:3]
        summary = "\n".join(
            (
                f"#{candidate['rank']} {candidate['segment_id']} | "
                f"match={candidate['match_score']:.1f} | "
                f"ordered={candidate.get('ordered_similarity', 0.0):.1f} | "
                f"coverage={candidate.get('token_coverage', 0.0):.0%} | "
                f"precision={candidate.get('token_precision', 0.0):.0%} | "
                f"weakest-clause="
                f"{candidate.get('minimum_clause_score', 0.0):.1f} | "
                + (
                    f"joined={candidate.get('fragment_source_count', 0)} | "
                    if candidate.get("fragment_join")
                    else ""
                )
                + f"margin={candidate['confidence_margin']:.1f} | "
                f"{candidate['transcript']}"
            )
            for candidate in line_candidates
        )
        if is_nonverbal and nonverbal_policy == "review":
            summary = (
                "Excluded from text matching. Audition the Unmatched Segments "
                "sheet and copy the intended Segment ID here manually."
            )
        elif is_nonverbal and nonverbal_policy == "skip":
            summary = (
                "Skipped by alignment.nonverbal_policy. Change the status and "
                "select a segment manually if this asset is required."
            )
        elif is_nonverbal and nonverbal_policy == "weak_order":
            summary = (
                "Weak phonetic/duration/order hints only; audition every "
                "candidate before selecting."
                + (f"\n{summary}" if summary else "")
            )
        lines_sheet.append(
            [
                line["line_id"],
                line["sheet"],
                line["excel_row"],
                line["quest"],
                line["context"],
                line["line"],
                line["acting_note"],
                line["emotion"],
                line["target_filename"],
                len(line_candidates),
                *[
                    (
                        top_candidates[index]["segment_id"]
                        if index < len(top_candidates)
                        else ""
                    )
                    for index in range(3)
                ],
                suggested,
                selected,
                status,
                summary,
                "",
            ]
        )
        output_row = lines_sheet.max_row
        link_targets = [
            (11, top_candidates[0] if len(top_candidates) > 0 else None),
            (12, top_candidates[1] if len(top_candidates) > 1 else None),
            (13, top_candidates[2] if len(top_candidates) > 2 else None),
            (14, best),
            (15, reliable_best),
        ]
        for column, candidate in link_targets:
            if candidate:
                _set_segment_hyperlink(
                    lines_sheet.cell(output_row, column),
                    project_dir,
                    candidate["segment_file"],
                )

    candidate_headers = [
        "Line ID",
        "Rank",
        "Segment ID",
        "Segment File",
        "Transcript",
        "Match Score",
        "Ordered Similarity",
        "Token Coverage",
        "Token Precision",
        "Extra Words",
        "Clause Count",
        "Minimum Clause Score",
        "Missing Clauses",
        "Fragment Join",
        "Joined Fragments",
        "Confidence Margin",
        "Technical Score",
        "Reliable",
        "Reliability Reason",
        "Alignment Mode",
        "Primary Match",
        "Segment Match Rank",
        "Take Group",
        "Duration Plausibility",
        "Duplicate Resolution",
        "Transcript Source",
        "Local ASR Rescued",
        "Source WAV",
        "Start Seconds",
        "End Seconds",
        "Duration Seconds",
        "ASR Confidence",
        "Unsafe Untranscribed Merge",
    ]
    candidates_sheet.append(candidate_headers)
    line_order = {
        line["line_id"]: index for index, line in enumerate(source_lines)
    }
    all_candidates = sorted(
        (
            candidate
            for line_candidates in candidates_by_line.values()
            for candidate in line_candidates
        ),
        key=lambda candidate: (
            line_order.get(candidate["line_id"], 10**9),
            candidate["rank"],
        ),
    )
    for candidate in all_candidates:
        candidates_sheet.append(
            [
                candidate["line_id"],
                candidate["rank"],
                candidate["segment_id"],
                candidate["segment_file"],
                candidate["transcript"],
                candidate["match_score"],
                candidate.get("ordered_similarity", 0.0),
                candidate.get("token_coverage", 0.0),
                candidate.get("token_precision", 0.0),
                candidate.get("extra_word_count", 0),
                candidate.get("clause_count", 0),
                candidate.get("minimum_clause_score", 0.0),
                candidate.get("missing_clause_count", 0),
                candidate.get("fragment_join", False),
                candidate.get("fragment_source_count", 0),
                candidate["confidence_margin"],
                candidate["technical_score"],
                candidate["reliable"],
                candidate["reliability_reason"],
                candidate.get("alignment_mode", "sequence"),
                candidate.get("is_primary_match", True),
                candidate.get("segment_match_rank", 1),
                candidate.get("take_group_id") or "",
                candidate.get("duration_plausibility", 0.0),
                candidate.get("duplicate_resolution") or "",
                candidate.get("transcript_source", "session_asr"),
                candidate.get("local_asr_rescued", False),
                candidate["source_audio"],
                candidate["start_seconds"],
                candidate["end_seconds"],
                candidate["duration_seconds"],
                candidate["asr_probability"],
                candidate.get("unsafe_untranscribed_merge", False),
            ]
        )
        file_cell = candidates_sheet.cell(candidates_sheet.max_row, 4)
        _set_segment_hyperlink(
            file_cell,
            project_dir,
            candidate["segment_file"],
        )

    unmatched_headers = [
        "Segment ID",
        "Segment File",
        "Source WAV",
        "Start Seconds",
        "End Seconds",
        "Duration Seconds",
        "Transcript",
        "ASR Confidence",
        "Reason",
        "Suggested Line 1",
        "Suggested Line 1 Text",
        "Suggested Line 1 Score",
        "Suggested Line 2",
        "Suggested Line 2 Text",
        "Suggested Line 2 Score",
        "Technical Flags",
    ]
    unmatched_sheet.append(unmatched_headers)
    for unmatched in unmatched_rows:
        unmatched_sheet.append(
            [
                unmatched["segment_id"],
                unmatched["segment_file"],
                unmatched["source_wav"],
                unmatched["start_seconds"],
                unmatched["end_seconds"],
                unmatched["duration_seconds"],
                unmatched["transcript"],
                unmatched["asr_confidence"],
                unmatched["reason"],
                unmatched["suggested_line_1"],
                unmatched["suggested_line_1_text"],
                unmatched["suggested_line_1_score"],
                unmatched["suggested_line_2"],
                unmatched["suggested_line_2_text"],
                unmatched["suggested_line_2_score"],
                unmatched["technical_flags"],
            ]
        )
        row = unmatched_sheet.max_row
        _set_segment_hyperlink(
            unmatched_sheet.cell(row, 1),
            project_dir,
            unmatched["segment_file"],
        )
        _set_segment_hyperlink(
            unmatched_sheet.cell(row, 2),
            project_dir,
            unmatched["segment_file"],
        )

    instructions = [
        ("Purpose", "Choose one temporary segment for every line that should be exported."),
        (
            "Editable field",
            "Click a linked candidate to audition it. Copy the whole Candidate 1, "
            "Candidate 2, or Candidate 3 cell into Lines!Selected Segment so both "
            "the Segment ID and its file link follow the selection. Segment IDs "
            "from the Candidates or Unmatched Segments sheet are also accepted.",
        ),
        (
            "Unmatched audio",
            "The Unmatched Segments sheet contains audio that could not be mapped "
            "reliably. Its Segment ID and Segment File cells are linked to the WAV.",
        ),
        (
            "Nonverbal lines",
            (
                "NONVERBAL_REVIEW lines are excluded from text matching and "
                "have no automatic selection. Audition Unmatched Segments and "
                "copy the intended Segment ID into Selected Segment."
                if nonverbal_policy == "review"
                else (
                    "Nonverbal lines start as SKIP. Change Status and select a "
                    "segment manually for any nonverbal asset that is required."
                    if nonverbal_policy == "skip"
                    else (
                        "NONVERBAL_REVIEW candidates use weak phonetic, duration, "
                        "and order hints only. Audition every candidate."
                    )
                )
            ),
        ),
        (
            "Duplicate lines",
            "Duplicate Resolution identifies candidates assigned with weak take "
            "order. A reused segment requires export.allow_segment_reuse or the "
            "finalize --allow-segment-reuse option.",
        ),
        (
            "Multi-sentence lines",
            "AUTO_OK requires every sentence-sized clause to be represented in "
            "the transcript. When adjacent ASR segments match successive clauses, "
            "the pipeline also creates a linked Fragment Join candidate spanning "
            "the complete take; the original fragments remain available for "
            "manual review.",
        ),
        (
            "Repeated takes",
            "AUTO_OK rejects transcripts containing excess repeated words, merges "
            "with voiced but untranscribed pieces, and candidates unusually long "
            "for the script text. The last case is labeled POSSIBLE_REPEATED_TAKES "
            "because ASR can collapse repeated performances into one transcript.",
        ),
        (
            "Status",
            "AUTO_OK was filled automatically; REVIEW, NONVERBAL_REVIEW, and "
            "MISSING require attention. Set Status to SKIP to intentionally "
            "omit a line.",
        ),
        (
            "Finalization",
            "The finalizer reads Selected Segment only. Suggested Best is informational.",
        ),
    ]
    instructions_sheet.append(["Topic", "Instruction"])
    for row in instructions:
        instructions_sheet.append(list(row))

    _style_workbook(
        lines_sheet,
        candidates_sheet,
        unmatched_sheet,
        instructions_sheet,
    )
    review_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(review_path)


def _set_segment_hyperlink(cell, project_dir: Path, segment_file: str) -> None:
    segment_path = resolve_project_path(project_dir, segment_file)
    try:
        cell.hyperlink = segment_path.as_uri()
        cell.style = "Hyperlink"
    except ValueError:
        pass


def _style_workbook(
    lines_sheet,
    candidates_sheet,
    unmatched_sheet,
    instructions_sheet,
) -> None:
    navy = "1F4E78"
    pale_blue = "D9EAF7"
    yellow = "FFF2CC"
    green = "E2F0D9"
    red = "FCE4D6"
    gray = "E7E6E6"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9E2F3")
    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(color=white, bold=True)

    for sheet in (
        lines_sheet,
        candidates_sheet,
        unmatched_sheet,
        instructions_sheet,
    ):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)
        sheet.row_dimensions[1].height = 28

    selected_column = 15
    status_column = 16
    candidate_summary_column = 17
    notes_column = 18
    lines_data_last_row = lines_sheet.max_row
    for row in range(2, lines_sheet.max_row + 1):
        lines_sheet.cell(row, selected_column).fill = PatternFill(
            "solid", fgColor=yellow
        )
        lines_sheet.cell(row, notes_column).fill = PatternFill(
            "solid", fgColor=yellow
        )
        lines_sheet.cell(row, candidate_summary_column).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        for column in (4, 5, 6, 7, 8, 11, 12, 13, 14, 15):
            lines_sheet.cell(row, column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    lines_sheet.conditional_formatting.add(
        f"P2:P{lines_data_last_row}",
        FormulaRule(formula=['P2="AUTO_OK"'], fill=PatternFill("solid", fgColor=green)),
    )
    lines_sheet.conditional_formatting.add(
        f"P2:P{lines_data_last_row}",
        FormulaRule(formula=['P2="REVIEW"'], fill=PatternFill("solid", fgColor=yellow)),
    )
    lines_sheet.conditional_formatting.add(
        f"P2:P{lines_data_last_row}",
        FormulaRule(formula=['P2="MISSING"'], fill=PatternFill("solid", fgColor=red)),
    )
    lines_sheet.conditional_formatting.add(
        f"P2:P{lines_data_last_row}",
        FormulaRule(
            formula=['P2="NONVERBAL_REVIEW"'],
            fill=PatternFill("solid", fgColor=pale_blue),
        ),
    )
    lines_sheet.conditional_formatting.add(
        f"P2:P{lines_data_last_row}",
        FormulaRule(formula=['P2="SKIP"'], fill=PatternFill("solid", fgColor=gray)),
    )
    lines_widths = {
        1: 34,
        2: 24,
        3: 10,
        4: 25,
        5: 55,
        6: 55,
        7: 34,
        8: 18,
        9: 42,
        10: 14,
        11: 42,
        12: 42,
        13: 42,
        14: 42,
        15: 42,
        status_column: 22,
        candidate_summary_column: 85,
        notes_column: 35,
    }
    for column, width in lines_widths.items():
        lines_sheet.column_dimensions[get_column_letter(column)].width = width

    for row in range(2, candidates_sheet.max_row + 1):
        for column in (3, 4, 5, 19, 23, 25, 26, 28):
            candidates_sheet.cell(row, column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        candidates_sheet.row_dimensions[row].height = 54
        for column in (6, 7, 12, 16, 17, 24):
            candidates_sheet.cell(row, column).number_format = "0.0"
        for column in (8, 9):
            candidates_sheet.cell(row, column).number_format = "0.0%"
        for column in range(29, 33):
            candidates_sheet.cell(row, column).number_format = "0.000"
    candidate_widths = [
        34,
        8,
        38,
        62,
        70,
        12,
        16,
        14,
        14,
        12,
        12,
        20,
        16,
        14,
        16,
        16,
        14,
        10,
        30,
        14,
        12,
        12,
        30,
        18,
        20,
        18,
        16,
        58,
        14,
        14,
        14,
        14,
        20,
    ]
    for column, width in enumerate(candidate_widths, start=1):
        candidates_sheet.column_dimensions[get_column_letter(column)].width = width

    for row in range(2, unmatched_sheet.max_row + 1):
        for column in (1, 2, 3, 7, 9, 10, 11, 13, 14, 16):
            unmatched_sheet.cell(row, column).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
        unmatched_sheet.row_dimensions[row].height = 54
        for column in range(4, 9):
            unmatched_sheet.cell(row, column).number_format = "0.000"
        for column in (12, 15):
            unmatched_sheet.cell(row, column).number_format = "0.0"
    unmatched_widths = [
        42,
        62,
        58,
        14,
        14,
        16,
        70,
        16,
        26,
        34,
        65,
        18,
        34,
        65,
        18,
        24,
    ]
    for column, width in enumerate(unmatched_widths, start=1):
        unmatched_sheet.column_dimensions[get_column_letter(column)].width = width

    instructions_sheet.column_dimensions["A"].width = 22
    instructions_sheet.column_dimensions["B"].width = 100
    for row in range(2, instructions_sheet.max_row + 1):
        instructions_sheet.cell(row, 1).fill = PatternFill(
            "solid", fgColor=pale_blue
        )
        instructions_sheet.cell(row, 1).font = Font(bold=True)
        instructions_sheet.cell(row, 2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    _add_status_summary_and_chart(
        lines_sheet,
        data_last_row=lines_data_last_row,
        status_column=status_column,
        header_fill=header_fill,
        header_font=header_font,
        status_fills={
            "AUTO_OK": green,
            "REVIEW": yellow,
            "MISSING": red,
            "NONVERBAL_REVIEW": pale_blue,
            "SKIP": gray,
        },
        border=thin_gray,
    )


def _add_status_summary_and_chart(
    lines_sheet,
    *,
    data_last_row: int,
    status_column: int,
    header_fill: PatternFill,
    header_font: Font,
    status_fills: dict[str, str],
    border: Side,
) -> None:
    title_row = data_last_row + 3
    header_row = title_row + 1
    first_status_row = header_row + 1
    status_column_letter = get_column_letter(status_column)

    lines_sheet.merge_cells(
        start_row=title_row,
        start_column=1,
        end_row=title_row,
        end_column=2,
    )
    title_cell = lines_sheet.cell(title_row, 1, "Line Status Summary")
    title_cell.fill = header_fill
    title_cell.font = header_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    lines_sheet.row_dimensions[title_row].height = 24

    lines_sheet.cell(header_row, 1, "Status")
    lines_sheet.cell(header_row, 2, "Line Count")
    for cell in lines_sheet[header_row][0:2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=border)

    statuses = [
        "AUTO_OK",
        "REVIEW",
        "MISSING",
        "NONVERBAL_REVIEW",
        "SKIP",
    ]
    for index, status in enumerate(statuses):
        row = first_status_row + index
        lines_sheet.cell(row, 1, status)
        lines_sheet.cell(
            row,
            2,
            (
                f'=COUNTIF(${status_column_letter}$2:'
                f'${status_column_letter}${data_last_row},A{row})'
            ),
        )
        lines_sheet.cell(row, 1).fill = PatternFill(
            "solid", fgColor=status_fills[status]
        )
        lines_sheet.cell(row, 2).number_format = "#,##0"

    chart = PieChart()
    chart.title = "Review Status Distribution"
    chart.height = 7.0
    chart.width = 11.5
    chart.legend.position = "r"
    data = Reference(
        lines_sheet,
        min_col=2,
        min_row=header_row,
        max_row=first_status_row + len(statuses) - 1,
    )
    labels = Reference(
        lines_sheet,
        min_col=1,
        min_row=first_status_row,
        max_row=first_status_row + len(statuses) - 1,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = True
    chart.series[0].data_points = [
        DataPoint(
            idx=index,
            spPr=GraphicalProperties(solidFill=status_fills[status]),
        )
        for index, status in enumerate(statuses)
    ]
    lines_sheet.add_chart(chart, f"D{title_row}")
