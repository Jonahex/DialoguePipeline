from __future__ import annotations

import csv
import os
import shutil
import tempfile
import wave
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .util import read_json, resolve_project_path


def _header_map(worksheet) -> dict[str, int]:
    return {
        str(cell.value or "").strip(): cell.column
        for cell in worksheet[1]
        if str(cell.value or "").strip()
    }


def _safe_target_filename(value: str, extension: str) -> str:
    value = value.strip()
    if not value:
        raise ValueError("Target filename is blank")
    path = Path(value)
    if path.name != value or any(separator in value for separator in ("/", "\\")):
        raise ValueError(f"Target filename must not contain a path: {value!r}")
    if path.suffix:
        if path.suffix.lower() != extension.lower():
            raise ValueError(
                f"Unexpected target extension {path.suffix!r}; expected {extension!r}"
            )
        return value
    return value + extension


def _segment_lookup(manifest: dict[str, Any]) -> dict[str, dict[str, Any]]:
    lookup = {}
    for session in manifest.get("sessions", []):
        for segment in session.get("segments", []) + session.get(
            "derived_segments", []
        ):
            segment_id = segment["segment_id"]
            if segment_id in lookup:
                raise ValueError(f"Duplicate segment ID in manifest: {segment_id}")
            lookup[segment_id] = segment
    return lookup


def _probe_wave(path: Path) -> dict[str, int]:
    with wave.open(str(path), "rb") as reader:
        return {
            "sample_rate": reader.getframerate(),
            "channels": reader.getnchannels(),
            "bits_per_sample": reader.getsampwidth() * 8,
            "frame_count": reader.getnframes(),
        }


def finalize_review(
    *,
    project_dir: Path,
    project: dict[str, Any],
    review_path: Path,
    output_dir: Path,
    overwrite: bool = False,
    allow_incomplete: bool = False,
    allow_segment_reuse: bool | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    manifest_path = project_dir / "segments_manifest.json"
    if not manifest_path.is_file():
        raise FileNotFoundError(manifest_path)
    manifest = read_json(manifest_path)
    segments = _segment_lookup(manifest)
    workbook = load_workbook(review_path, read_only=True, data_only=True)
    if "Lines" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Review workbook has no Lines sheet: {review_path}")
    worksheet = workbook["Lines"]
    headers = _header_map(worksheet)
    required = {"Line ID", "Target Filename", "Selected Segment", "Status"}
    missing_headers = sorted(required - headers.keys())
    if missing_headers:
        workbook.close()
        raise ValueError("Missing review columns: " + ", ".join(missing_headers))

    extension = str(project.get("export", {}).get("extension", ".wav"))
    expected_format = {
        "sample_rate": int(project.get("export", {}).get("sample_rate", 48000)),
        "channels": int(project.get("export", {}).get("channels", 1)),
        "bits_per_sample": int(
            project.get("export", {}).get("bits_per_sample", 16)
        ),
    }
    if allow_segment_reuse is None:
        allow_segment_reuse = bool(
            project.get("export", {}).get("allow_segment_reuse", False)
        )
    exports = []
    errors = []
    selected_ids = []
    target_names = []

    try:
        for row_number in range(2, worksheet.max_row + 1):
            line_id = str(
                worksheet.cell(row_number, headers["Line ID"]).value or ""
            ).strip()
            target_value = str(
                worksheet.cell(row_number, headers["Target Filename"]).value or ""
            ).strip()
            selected_id = str(
                worksheet.cell(row_number, headers["Selected Segment"]).value or ""
            ).strip()
            status = str(
                worksheet.cell(row_number, headers["Status"]).value or ""
            ).strip().upper()
            if not line_id:
                continue
            if not target_value and not selected_id and not status:
                continue
            if status == "SKIP":
                continue
            if not selected_id:
                errors.append(
                    {
                        "line_id": line_id,
                        "error": "NO_SELECTED_SEGMENT",
                        "detail": f"Lines row {row_number}",
                    }
                )
                continue
            segment = segments.get(selected_id)
            if segment is None:
                errors.append(
                    {
                        "line_id": line_id,
                        "error": "UNKNOWN_SEGMENT",
                        "detail": selected_id,
                    }
                )
                continue
            try:
                target_name = _safe_target_filename(target_value, extension)
            except ValueError as error:
                errors.append(
                    {
                        "line_id": line_id,
                        "error": "INVALID_TARGET_FILENAME",
                        "detail": str(error),
                    }
                )
                continue

            segment_path = resolve_project_path(project_dir, segment["file"])
            if not segment_path.is_file():
                errors.append(
                    {
                        "line_id": line_id,
                        "error": "SEGMENT_FILE_MISSING",
                        "detail": str(segment_path),
                    }
                )
                continue
            wave_format = _probe_wave(segment_path)
            mismatches = {
                key: (wave_format[key], expected)
                for key, expected in expected_format.items()
                if wave_format[key] != expected
            }
            if mismatches:
                errors.append(
                    {
                        "line_id": line_id,
                        "error": "FORMAT_MISMATCH",
                        "detail": repr(mismatches),
                    }
                )
                continue

            destination = output_dir / target_name
            exports.append(
                {
                    "line_id": line_id,
                    "target_filename": target_name,
                    "selected_segment_id": selected_id,
                    "segment_file": str(segment_path),
                    "output_file": str(destination),
                    "source_audio": segment["source_audio"],
                    "start_seconds": segment["start_seconds"],
                    "end_seconds": segment["end_seconds"],
                    "transcript": segment.get("transcript", ""),
                }
            )
            selected_ids.append(selected_id)
            target_names.append(target_name.lower())
    finally:
        workbook.close()

    if not allow_segment_reuse:
        duplicate_segment_ids = {
            segment_id
            for segment_id, count in Counter(selected_ids).items()
            if count > 1
        }
        for segment_id in sorted(duplicate_segment_ids):
            affected = [
                export["line_id"]
                for export in exports
                if export["selected_segment_id"] == segment_id
            ]
            for line_id in affected:
                errors.append(
                    {
                        "line_id": line_id,
                        "error": "SEGMENT_REUSED",
                        "detail": (
                            f"{segment_id} also assigned to another line"
                        ),
                    }
                )
    duplicate_target_names = {
        target_name
        for target_name, count in Counter(target_names).items()
        if count > 1
    }
    for target_name in sorted(duplicate_target_names):
        affected = [
            export["line_id"]
            for export in exports
            if export["target_filename"].lower() == target_name
        ]
        for line_id in affected:
            errors.append(
                {
                    "line_id": line_id,
                    "error": "DUPLICATE_TARGET",
                    "detail": target_name,
                }
            )
    if not overwrite:
        for export in exports:
            destination = Path(export["output_file"])
            if destination.exists():
                errors.append(
                    {
                        "line_id": export["line_id"],
                        "error": "OUTPUT_EXISTS",
                        "detail": str(destination),
                    }
                )

    output_dir.mkdir(parents=True, exist_ok=True)
    export_report = output_dir / "final_export_manifest.tsv"
    error_report = output_dir / "finalization_errors.tsv"
    _write_tsv(
        error_report,
        errors,
        ["line_id", "error", "detail"],
    )

    blocking_errors = bool(errors) and not allow_incomplete
    invalid_line_ids = {error["line_id"] for error in errors if error["line_id"]}
    valid_exports = [
        export for export in exports if export["line_id"] not in invalid_line_ids
    ]
    if not dry_run and not blocking_errors:
        for export in valid_exports:
            source = Path(export["segment_file"])
            destination = Path(export["output_file"])
            destination.parent.mkdir(parents=True, exist_ok=True)
            with tempfile.NamedTemporaryFile(
                dir=destination.parent,
                prefix=f".{destination.name}.",
                suffix=".tmp",
                delete=False,
            ) as handle:
                temporary = Path(handle.name)
            try:
                shutil.copyfile(source, temporary)
                if destination.exists() and not overwrite:
                    raise FileExistsError(destination)
                os.replace(temporary, destination)
            finally:
                if temporary.exists():
                    temporary.unlink()

    _write_tsv(
        export_report,
        valid_exports,
        [
            "line_id",
            "target_filename",
            "selected_segment_id",
            "segment_file",
            "output_file",
            "source_audio",
            "start_seconds",
            "end_seconds",
            "transcript",
        ],
    )
    if blocking_errors:
        raise ValueError(
            f"Finalization stopped with {len(errors)} error(s). "
            f"See {error_report}. Use --allow-incomplete to export valid selections."
        )
    return {
        "export_count": len(valid_exports),
        "error_count": len(errors),
        "export_report": export_report,
        "error_report": error_report,
        "dry_run": dry_run,
        "allow_segment_reuse": allow_segment_reuse,
    }


def _write_tsv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
            delimiter="\t",
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)
