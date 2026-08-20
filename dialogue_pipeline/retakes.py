from __future__ import annotations

import copy
import os
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .review import load_line_review
from .util import read_json, resolve_project_path
from .workbook_io import (
    _effective_spoken_line,
    _find_header_row,
    _find_header_row_for_data_row,
)


SUPPORTED_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}


def _cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _copy_cell(source: Any, destination: Any) -> None:
    destination.value = source.value
    if source.has_style:
        destination._style = copy.copy(source._style)
    if source.hyperlink:
        destination._hyperlink = copy.copy(source.hyperlink)
    if source.comment:
        destination.comment = copy.copy(source.comment)


def _copy_source_row(
    *,
    source: Any,
    source_row: int,
    destination: Any,
    destination_row: int,
) -> None:
    for column in range(1, source.max_column + 1):
        _copy_cell(
            source.cell(source_row, column),
            destination.cell(destination_row, column),
        )

    source_dimension = source.row_dimensions[source_row]
    destination_dimension = destination.row_dimensions[destination_row]
    destination_dimension.height = source_dimension.height
    destination_dimension.hidden = source_dimension.hidden
    destination_dimension.outlineLevel = source_dimension.outlineLevel
    destination_dimension.collapsed = source_dimension.collapsed
    if source_dimension.has_style:
        destination_dimension._style = copy.copy(source_dimension._style)

    for merged_range in source.merged_cells.ranges:
        if (
            merged_range.min_row == source_row
            and merged_range.max_row == source_row
        ):
            destination.merge_cells(
                start_row=destination_row,
                start_column=merged_range.min_col,
                end_row=destination_row,
                end_column=merged_range.max_col,
            )


def _clear_template_rows(worksheet: Any, header_row: int) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row > header_row:
            worksheet.unmerge_cells(str(merged_range))
    if worksheet.max_row > header_row:
        worksheet.delete_rows(
            header_row + 1,
            worksheet.max_row - header_row,
        )
    for row_index in list(worksheet.row_dimensions):
        if row_index > header_row:
            del worksheet.row_dimensions[row_index]


def _source_line_lookup(
    *,
    project_dir: Path,
    project: dict[str, Any],
) -> dict[str, dict[str, Any]]:
    source_path = resolve_project_path(project_dir, project["source_lines"])
    source_data = read_json(source_path)
    return {
        str(line["line_id"]): line
        for line in source_data.get("lines", [])
    }


def export_retake_script(
    *,
    project_dir: Path,
    project: dict[str, Any],
    review_path: Path,
    output_path: Path,
    overwrite: bool = False,
) -> dict[str, Any]:
    """Export RETAKE lines on their original source workbook sheets."""

    project_dir = project_dir.resolve()
    source_workbook = resolve_project_path(project_dir, project["workbook"])
    output_path = output_path.resolve()
    source_suffix = source_workbook.suffix.lower()
    if source_suffix not in SUPPORTED_WORKBOOK_SUFFIXES:
        raise ValueError(
            f"Unsupported source workbook format: {source_workbook.suffix!r}"
        )
    if output_path.suffix.lower() != source_suffix:
        raise ValueError(
            "The retake script must use the same file extension as the source "
            f"workbook ({source_suffix})."
        )
    if output_path == source_workbook:
        raise ValueError("The retake export cannot replace the source workbook.")
    if not source_workbook.is_file():
        raise FileNotFoundError(source_workbook)
    if output_path.exists() and not overwrite:
        raise FileExistsError(output_path)

    review_data = load_line_review(review_path)
    retake_review_lines = [
        line for line in review_data["lines"] if line["status"] == "RETAKE"
    ]
    if not retake_review_lines:
        raise ValueError("No lines are marked for retake.")

    source_lines = _source_line_lookup(
        project_dir=project_dir,
        project=project,
    )
    retake_lines = []
    for review_line in retake_review_lines:
        line_id = str(review_line["line_id"])
        source_line = source_lines.get(line_id)
        if source_line is None:
            raise ValueError(
                f"Retake line is not present in source_lines.json: {line_id}"
            )
        retake_lines.append(source_line)
    retake_lines.sort(
        key=lambda line: (
            int(line["sheet_index"]),
            int(line["excel_row"]),
        )
    )

    workbook = load_workbook(
        filename=source_workbook,
        read_only=False,
        data_only=False,
        keep_vba=source_suffix == ".xlsm",
    )
    temporary_path: Path | None = None
    try:
        for line in retake_lines:
            sheet_name = str(line["sheet"])
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Source workbook no longer contains sheet {sheet_name!r}."
                )
            worksheet = workbook[sheet_name]
            row_number = int(line["excel_row"])
            _, columns = _find_header_row_for_data_row(
                worksheet,
                row_number,
            )
            workbook_line = _effective_spoken_line(
                worksheet,
                row_number,
                columns,
            )
            workbook_filename = _cell_text(
                worksheet.cell(row_number, columns["filename"]).value
            )
            if (
                workbook_line != str(line["line"])
                or workbook_filename != str(line["target_filename"])
            ):
                raise ValueError(
                    "The source workbook changed after this project was created "
                    f"at {sheet_name}!{row_number}."
                )

        canonical_values = {
            "quest": "quest",
            "context": "context",
            "line": "line",
            "acting_note": "acting_note",
            "emotion": "emotion",
            "filename": "target_filename",
        }
        lines_by_sheet: dict[str, list[dict[str, Any]]] = {}
        for line in retake_lines:
            lines_by_sheet.setdefault(str(line["sheet"]), []).append(line)

        original_worksheets = list(workbook.worksheets)
        exported_sheets: list[tuple[Any, str]] = []
        for sheet_name, sheet_lines in lines_by_sheet.items():
            source_sheet = workbook[sheet_name]
            destination_sheet = workbook.copy_worksheet(source_sheet)
            destination_sheet.freeze_panes = source_sheet.freeze_panes
            destination_sheet.auto_filter = copy.copy(source_sheet.auto_filter)
            destination_sheet.sheet_view.showGridLines = (
                source_sheet.sheet_view.showGridLines
            )
            header_row, columns = _find_header_row(destination_sheet)
            _clear_template_rows(destination_sheet, header_row)

            previous_quest: str | None = None
            for offset, line in enumerate(sheet_lines, start=1):
                destination_row = header_row + offset
                _copy_source_row(
                    source=source_sheet,
                    source_row=int(line["excel_row"]),
                    destination=destination_sheet,
                    destination_row=destination_row,
                )
                quest = str(line.get("quest") or "")
                for canonical, source_key in canonical_values.items():
                    target_column = columns.get(canonical)
                    if target_column is None:
                        continue
                    value = str(line.get(source_key) or "")
                    if canonical == "quest":
                        value = quest if quest != previous_quest else ""
                    destination_sheet.cell(
                        destination_row,
                        target_column,
                    ).value = value
                previous_quest = quest

            final_row = header_row + len(sheet_lines)
            if destination_sheet.auto_filter.ref:
                last_column = max(
                    destination_sheet.max_column,
                    max(columns.values()),
                )
                destination_sheet.auto_filter.ref = (
                    f"A{header_row}:"
                    f"{get_column_letter(last_column)}{final_row}"
                )
            exported_sheets.append((destination_sheet, sheet_name))

        for worksheet in original_worksheets:
            workbook.remove(worksheet)
        for worksheet, original_name in exported_sheets:
            worksheet.title = original_name
            worksheet.sheet_state = "visible"
        workbook.active = 0

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            dir=output_path.parent,
            prefix=f".{output_path.stem}.",
            suffix=source_suffix,
            delete=False,
        ) as handle:
            temporary_path = Path(handle.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
        temporary_path = None
    finally:
        workbook.close()
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()

    return {
        "export_count": len(retake_lines),
        "output_path": output_path,
        "source_workbook": source_workbook,
        "sheet_names": list(lines_by_sheet),
    }
