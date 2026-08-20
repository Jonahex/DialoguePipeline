from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook


HEADER_ALIASES = {
    "quest": {"quest"},
    "context": {"context", "topic text"},
    "context_fallback": {"topic"},
    "line": {
        "line to speak",
        "line",
        "dialogue",
        "text",
        "response text",
    },
    "acting_note": {
        "acting note",
        "acting notes",
        "note",
        "script notes",
    },
    "emotion": {"facial emotion", "emotion"},
    "filename": {"filename", "file name", "output filename"},
    "voice_type": {"voice type", "voicetype"},
}


_ODS_NAMESPACES = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}
_ODS_TABLE_NAME = f"{{{_ODS_NAMESPACES['table']}}}name"
_ODS_COLUMN_REPEAT = (
    f"{{{_ODS_NAMESPACES['table']}}}number-columns-repeated"
)
_ODS_ROW_REPEAT = f"{{{_ODS_NAMESPACES['table']}}}number-rows-repeated"
_ODS_STRING_VALUE = f"{{{_ODS_NAMESPACES['office']}}}string-value"
_ODS_VALUE = f"{{{_ODS_NAMESPACES['office']}}}value"
_ODS_DATE_VALUE = f"{{{_ODS_NAMESPACES['office']}}}date-value"
_ODS_TIME_VALUE = f"{{{_ODS_NAMESPACES['office']}}}time-value"
_ODS_BOOLEAN_VALUE = f"{{{_ODS_NAMESPACES['office']}}}boolean-value"
_ODS_SPACE_TAG = f"{{{_ODS_NAMESPACES['text']}}}s"
_ODS_SPACE_COUNT = f"{{{_ODS_NAMESPACES['text']}}}c"
_ODS_TAB_TAG = f"{{{_ODS_NAMESPACES['text']}}}tab"
_ODS_LINE_BREAK_TAG = f"{{{_ODS_NAMESPACES['text']}}}line-break"
_ODS_CELL_TAGS = {
    f"{{{_ODS_NAMESPACES['table']}}}table-cell",
    f"{{{_ODS_NAMESPACES['table']}}}covered-table-cell",
}


@dataclass(frozen=True)
class _OdsCell:
    value: Any = None


class _OdsWorksheet:
    use_legacy_voice_header = False
    allow_blank_line_note_fallback = True

    def __init__(
        self,
        *,
        title: str,
        rows: dict[int, dict[int, str]],
    ) -> None:
        self.title = title
        self._rows = rows
        self.max_row = max(rows, default=0)
        self.max_column = max(
            (
                column
                for row in rows.values()
                for column in row
            ),
            default=0,
        )

    def cell(self, row: int, column: int) -> _OdsCell:
        return _OdsCell(self._rows.get(row, {}).get(column))


class _OdsWorkbook:
    def __init__(self, worksheets: list[_OdsWorksheet]) -> None:
        self.worksheets = worksheets

    def close(self) -> None:
        return None


def _ods_element_text(element: ET.Element) -> str:
    parts = [element.text or ""]
    for child in element:
        if child.tag == _ODS_SPACE_TAG:
            parts.append(" " * max(1, int(child.get(_ODS_SPACE_COUNT, "1"))))
        elif child.tag == _ODS_TAB_TAG:
            parts.append("\t")
        elif child.tag == _ODS_LINE_BREAK_TAG:
            parts.append("\n")
        else:
            parts.append(_ods_element_text(child))
        parts.append(child.tail or "")
    return "".join(parts)


def _ods_cell_text(cell: ET.Element) -> str:
    paragraphs = cell.findall(".//text:p", _ODS_NAMESPACES)
    if paragraphs:
        return "\n".join(
            text
            for paragraph in paragraphs
            if (text := _ods_element_text(paragraph).strip())
        )
    for attribute in (
        _ODS_STRING_VALUE,
        _ODS_VALUE,
        _ODS_DATE_VALUE,
        _ODS_TIME_VALUE,
        _ODS_BOOLEAN_VALUE,
    ):
        value = cell.get(attribute)
        if value is not None:
            return value
    return ""


def _load_ods_workbook(path: Path) -> _OdsWorkbook:
    try:
        with ZipFile(path) as archive:
            content = archive.read("content.xml")
    except (BadZipFile, KeyError) as exc:
        raise ValueError(f"Invalid ODS workbook: {path}") from exc

    try:
        root = ET.fromstring(content)
    except ET.ParseError as exc:
        raise ValueError(f"Invalid ODS XML in workbook: {path}") from exc

    worksheets = []
    for table in root.findall(".//table:table", _ODS_NAMESPACES):
        rows: dict[int, dict[int, str]] = {}
        logical_row = 1
        for row in table.findall("table:table-row", _ODS_NAMESPACES):
            row_values: dict[int, str] = {}
            logical_column = 1
            for cell in row:
                if cell.tag not in _ODS_CELL_TAGS:
                    continue
                repeated_columns = max(
                    1,
                    int(cell.get(_ODS_COLUMN_REPEAT, "1")),
                )
                value = _ods_cell_text(cell)
                if value:
                    for offset in range(repeated_columns):
                        row_values[logical_column + offset] = value
                logical_column += repeated_columns

            repeated_rows = max(1, int(row.get(_ODS_ROW_REPEAT, "1")))
            if row_values:
                for offset in range(repeated_rows):
                    rows[logical_row + offset] = dict(row_values)
            logical_row += repeated_rows

        worksheets.append(
            _OdsWorksheet(
                title=table.get(_ODS_TABLE_NAME) or "Sheet",
                rows=rows,
            )
        )
    if not worksheets:
        raise ValueError(f"ODS workbook contains no worksheets: {path}")
    return _OdsWorkbook(worksheets)


def _normalized_header(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _header_columns(worksheet, row_number: int) -> dict[str, int]:
    found: dict[str, int] = {}
    for column_number in range(1, worksheet.max_column + 1):
        value = _normalized_header(worksheet.cell(row_number, column_number).value)
        for canonical, aliases in HEADER_ALIASES.items():
            if value in aliases:
                found[canonical] = column_number
    return found


def _is_dialogue_header(columns: dict[str, int]) -> bool:
    return "line" in columns and "filename" in columns


def _find_header_row(worksheet) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(worksheet.max_row, 20) + 1):
        found = _header_columns(worksheet, row_number)
        if _is_dialogue_header(found):
            return row_number, found
    raise ValueError(f"Could not find dialogue headers in sheet {worksheet.title!r}")


def _find_header_row_for_data_row(
    worksheet,
    data_row: int,
) -> tuple[int, dict[str, int]]:
    """Return the nearest dialogue header preceding a source data row."""
    found_layout: tuple[int, dict[str, int]] | None = None
    for row_number in range(1, min(data_row, worksheet.max_row) + 1):
        columns = _header_columns(worksheet, row_number)
        if _is_dialogue_header(columns):
            found_layout = (row_number, columns)
    if found_layout is None:
        raise ValueError(
            f"Could not find dialogue headers before "
            f"{worksheet.title}!{data_row}"
        )
    return found_layout


def _cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _effective_spoken_line(
    worksheet,
    row_number: int,
    columns: dict[str, int],
) -> str:
    line_value = worksheet.cell(row_number, columns["line"]).value
    spoken_line = _cell_text(line_value)
    if spoken_line:
        return spoken_line

    # Some Bethesda dialogue exports put a deliberate blank placeholder in
    # "Line to speak" and store the vocalization itself in "Acting Note".
    # Do not turn a genuinely missing line into an instruction: the fallback
    # applies only when the line cell exists and is explicitly blank.
    acting_note_column = columns.get("acting_note")
    if acting_note_column and (
        line_value is not None
        or getattr(worksheet, "allow_blank_line_note_fallback", False)
    ):
        return _cell_text(worksheet.cell(row_number, acting_note_column).value)
    return ""


def parse_workbook(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".ods":
        workbook = _load_ods_workbook(path)
    else:
        workbook = load_workbook(
            filename=path,
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    lines: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    filenames: list[str] = []

    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            _find_header_row(worksheet)
            columns: dict[str, int] = {}
            voice_headers: list[str] = []
            if getattr(worksheet, "use_legacy_voice_header", True):
                first_voice_header = _cell_text(
                    worksheet.cell(1, 2).value
                )
                if first_voice_header:
                    voice_headers.append(first_voice_header)
            sheet_lines: list[str] = []
            current_quest = ""

            for row_number in range(1, worksheet.max_row + 1):
                row_columns = _header_columns(worksheet, row_number)
                if _is_dialogue_header(row_columns):
                    # A worksheet may contain several actor/faction sections,
                    # each with a fresh header row and potentially a different
                    # column layout.
                    columns = row_columns
                    current_quest = ""
                    continue

                possible_voice_header = _cell_text(
                    worksheet.cell(row_number, 2).value
                )
                if possible_voice_header.lower().startswith("you are voicing"):
                    if possible_voice_header not in voice_headers:
                        voice_headers.append(possible_voice_header)

                if not columns:
                    continue

                def get(field: str) -> str:
                    column = columns.get(field)
                    if not column:
                        return ""
                    value = worksheet.cell(row_number, column).value
                    return _cell_text(value)

                voice_type = get("voice_type")
                if voice_type:
                    voice_header = (
                        voice_type
                        if voice_type.lower().startswith("you are voicing")
                        else f"You are voicing {voice_type}"
                    )
                    if voice_header not in voice_headers:
                        voice_headers.append(voice_header)

                quest_raw = get("quest")
                if quest_raw:
                    current_quest = quest_raw
                spoken_line = _effective_spoken_line(
                    worksheet,
                    row_number,
                    columns,
                )
                filename = get("filename")
                if not spoken_line and not filename:
                    continue
                if not spoken_line or not filename:
                    raise ValueError(
                        f"Incomplete dialogue row at {worksheet.title}!{row_number}: "
                        f"line={spoken_line!r}, filename={filename!r}"
                    )

                line_id = f"{worksheet.title}::R{row_number}"
                record = {
                    "line_id": line_id,
                    "sheet": worksheet.title,
                    "sheet_index": sheet_index,
                    "excel_row": row_number,
                    "quest": current_quest,
                    "context": get("context") or get("context_fallback"),
                    "line": spoken_line,
                    "acting_note": get("acting_note"),
                    "emotion": get("emotion"),
                    "target_filename": filename,
                }
                lines.append(record)
                sheet_lines.append(line_id)
                filenames.append(filename)

            sheets.append(
                {
                    "name": worksheet.title,
                    "index": sheet_index,
                    "voice_header": " | ".join(voice_headers),
                    "voice_headers": voice_headers,
                    "line_ids": sheet_lines,
                    "line_count": len(sheet_lines),
                }
            )
    finally:
        workbook.close()

    duplicates = sorted(name for name, count in Counter(filenames).items() if count > 1)
    if duplicates:
        raise ValueError(
            "Duplicate target filenames in workbook: " + ", ".join(duplicates[:20])
        )

    return {
        "schema_version": 1,
        "source_workbook": str(path.resolve()),
        "line_count": len(lines),
        "sheet_count": len(sheets),
        "sheets": sheets,
        "lines": lines,
    }


def lines_for_session(
    source_data: dict[str, Any], session: dict[str, Any]
) -> list[dict[str, Any]]:
    sheet_order = {name: index for index, name in enumerate(session.get("sheets", []))}
    allowed_rows = set(session.get("excel_rows") or [])
    allowed_ids = set(session.get("line_ids") or [])

    selected = []
    for line in source_data["lines"]:
        if allowed_ids and line["line_id"] not in allowed_ids:
            continue
        if line["sheet"] not in sheet_order:
            continue
        if allowed_rows and line["excel_row"] not in allowed_rows:
            continue
        selected.append(line)

    return sorted(
        selected,
        key=lambda line: (
            sheet_order[line["sheet"]],
            line["excel_row"],
        ),
    )
