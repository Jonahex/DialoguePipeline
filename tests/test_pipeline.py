from __future__ import annotations

import math
import struct
import subprocess
import threading
import wave
from pathlib import Path
from types import SimpleNamespace
from zipfile import ZIP_STORED, ZipFile

import numpy as np
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from dialogue_pipeline.alignment import (
    _apply_duplicate_line_policy,
    _boundary_clause_consensus_transcript,
    _boundary_noise_cleanup_actions,
    _boundary_voice_trim_actions,
    _candidate_reliability,
    _complete_subspan_recovery_actions,
    _edge_vocalization_extension_actions,
    _exact_line_scores,
    _expand_alignment_actions,
    _has_unsafe_untranscribed_merge,
    _intra_segment_trim_actions,
    _multisentence_fragment_join_actions,
    _reroute_actions_to_exact_asr_primary_matches,
    _word_gap_boundaries,
    SpanCatalog,
    TranscriptEvaluator,
    align_project,
    order_independent_align,
    sentence_fidelity,
    text_similarity,
    transcript_fidelity,
)
from dialogue_pipeline.alignment_settings import (
    AlignmentSettings,
    default_alignment_config,
)
from dialogue_pipeline.audio import (
    cut_pcm_wav,
    first_quiet_pcm_boundary,
    open_pcm_wav,
    pcm_wav_shape,
    prepare_pcm_segmentation_source,
    quietest_pcm_boundary,
)
from dialogue_pipeline.cancellation import (
    ProcessingCancelled,
    cancellation_scope,
    check_processing_cancelled,
)
from dialogue_pipeline.finalize import finalize_review
from dialogue_pipeline.project import (
    create_project,
    editable_project_settings,
    infer_sessions,
    migrate_project_config,
)
from dialogue_pipeline.retakes import export_retake_script
from dialogue_pipeline.review import (
    add_base_segment_candidate,
    build_line_review,
    delete_edited_candidate,
    load_line_review,
    preserve_manual_selections,
    prune_line_candidates,
    save_edited_candidate,
    save_line_review,
    segment_edit_source,
    transcribe_edited_candidate,
)
from dialogue_pipeline.segmentation import (
    materialize_trimmed_segment,
    prevent_region_overlaps,
    refresh_project_audio,
    segment_project,
    split_regions_on_word_gaps,
)
from dialogue_pipeline.transcription import (
    _automatic_batch_size,
    _best_ordered_script_score,
    _decode_clips_batched,
    _likely_silence_hallucination,
    _prompt_candidates,
    transcribe_candidate_span,
    transcribe_candidate_spans,
    transcribe_project,
    transcribe_segments_project,
)
from dialogue_pipeline.ui import (
    AudioPlayer,
    DialogueReviewApp,
    _base_segment_candidate_usage,
    _candidate_selection_display,
    _candidate_take_groups,
    _context_display_text,
    _first_line_base_segment_key,
    _format_excel_rows,
    _initial_segment_window,
    _mapping_sheet_action_names,
    _parse_excel_rows,
    _panned_sample_window,
    _playback_samples,
    _project_settings_from_values,
    _read_pcm_wav,
    _selected_segment_score,
    _selected_line_ids_by_segment,
    _sounddevice_output_options,
    _uses_unmatched_candidates,
    _validate_session_mappings,
    _zoomed_sample_window,
)
from dialogue_pipeline.util import (
    default_model_cache_root,
    read_json,
    resolve_model_cache_root,
    sha256_file,
    stable_hash,
    verbal_script_text,
    write_json,
)
from dialogue_pipeline.workbook_io import parse_workbook


def _alignment_settings(**groups: Any) -> AlignmentSettings:
    return AlignmentSettings.from_value(groups)


def _write_tone(
    path: Path,
    duration_seconds: float = 1.0,
    amplitude: int = 8000,
) -> None:
    sample_rate = 48000
    time = np.arange(int(sample_rate * duration_seconds)) / sample_rate
    samples = np.rint(np.sin(2 * math.pi * 440 * time) * amplitude).astype("<i2")
    with wave.open(str(path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(2)
        writer.setframerate(sample_rate)
        writer.writeframes(samples.tobytes())


def _write_extensible_pcm_tone(path: Path) -> np.ndarray:
    sample_rate = 96000
    samples = np.array([-12000, -4000, 4000, 12000], dtype="<i2")
    channels = 1
    bits_per_sample = 16
    block_alignment = channels * bits_per_sample // 8
    byte_rate = sample_rate * block_alignment
    pcm_subformat_guid = bytes.fromhex("0100000000001000800000aa00389b71")
    fmt_chunk = (
        struct.pack(
            "<HHIIHHH",
            0xFFFE,
            channels,
            sample_rate,
            byte_rate,
            block_alignment,
            bits_per_sample,
            22,
        )
        + struct.pack("<HI", bits_per_sample, 0x4)
        + pcm_subformat_guid
    )
    data = samples.tobytes()
    riff_size = 4 + 8 + len(fmt_chunk) + 8 + len(data)
    path.write_bytes(
        b"RIFF"
        + struct.pack("<I", riff_size)
        + b"WAVEfmt "
        + struct.pack("<I", len(fmt_chunk))
        + fmt_chunk
        + b"data"
        + struct.pack("<I", len(data))
        + data
    )
    return samples


def test_audio_player_uses_in_process_shared_wasapi_playback(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    audio_path = tmp_path / "candidate.wav"
    _write_tone(audio_path, duration_seconds=0.05)
    settings_calls: list[dict[str, Any]] = []
    shared_settings = object()
    stream_calls: list[dict[str, Any]] = []

    class FakeStream:
        started = False
        stopped = False
        closed = False

        def start(self) -> None:
            self.started = True

        def stop(self) -> None:
            self.stopped = True

        def close(self) -> None:
            self.closed = True

    stream = FakeStream()

    monkeypatch.setattr(
        "dialogue_pipeline.ui.sys",
        SimpleNamespace(platform="win32"),
    )
    monkeypatch.setattr(
        "dialogue_pipeline.ui.sd.query_hostapis",
        lambda: (
            {
                "name": "MME",
                "default_output_device": 5,
            },
            {
                "name": "Windows WASAPI",
                "default_output_device": 15,
            },
        ),
    )

    def fake_wasapi_settings(**kwargs: Any) -> object:
        settings_calls.append(kwargs)
        return shared_settings

    monkeypatch.setattr(
        "dialogue_pipeline.ui.sd.WasapiSettings",
        fake_wasapi_settings,
    )
    monkeypatch.setattr(
        "dialogue_pipeline.ui.sd.OutputStream",
        lambda **kwargs: (stream_calls.append(kwargs), stream)[1],
    )
    player = AudioPlayer()

    playback_id = player.play(audio_path)

    assert stream.started is True
    assert len(stream_calls) == 1
    assert stream_calls[0] == {
        "samplerate": 48000,
        "channels": 1,
        "dtype": "float32",
        "callback": player._fill_output,
        "device": 15,
        "extra_settings": shared_settings,
    }
    assert settings_calls == [{"exclusive": False, "auto_convert": True}]

    output = np.zeros((480, 1), dtype=np.float32)
    player._fill_output(output, 480, None, None)
    assert np.any(output != 0)
    position, duration, playing, paused = player.status(playback_id)
    assert position == pytest.approx(0.01)
    assert duration == pytest.approx(0.05)
    assert playing is True
    assert paused is False

    assert player.pause(playback_id) is True
    assert player.pause(playback_id) is False
    output.fill(1)
    player._fill_output(output, 480, None, None)
    assert np.all(output == 0)
    assert player.status(playback_id)[0] == pytest.approx(0.01)

    assert player.seek(playback_id, 0.02) is True
    assert player.resume(playback_id) is True
    player._fill_output(output, 480, None, None)
    assert player.status(playback_id)[0] == pytest.approx(0.03)

    player.stop()
    output.fill(1)
    player._fill_output(output, 480, None, None)
    assert np.all(output == 0)

    player.close()
    assert stream.stopped is True
    assert stream.closed is True


def test_sounddevice_output_options_requires_wasapi_default(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        "dialogue_pipeline.ui.sys",
        SimpleNamespace(platform="win32"),
    )
    monkeypatch.setattr(
        "dialogue_pipeline.ui.sd.query_hostapis",
        lambda: ({"name": "MME", "default_output_device": 5},),
    )

    with pytest.raises(RuntimeError, match="No default Windows WASAPI"):
        _sounddevice_output_options()


def test_read_pcm_wav_supports_unsigned_8_bit_samples(tmp_path: Path) -> None:
    audio_path = tmp_path / "eight_bit.wav"
    with wave.open(str(audio_path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(1)
        writer.setframerate(48000)
        writer.writeframes(b"\x80")

    samples, sample_rate = _read_pcm_wav(audio_path)

    assert samples.dtype == np.uint8
    assert samples.tolist() == [128]
    assert sample_rate == 48000


def test_extensible_pcm_wav_is_supported_for_waveform_and_cutting(
    tmp_path: Path,
) -> None:
    source = tmp_path / "extensible.wav"
    expected_samples = _write_extensible_pcm_tone(source)

    assert pcm_wav_shape(source) == {
        "sample_rate": 96000,
        "channels": 1,
        "bits_per_sample": 16,
        "frame_count": 4,
    }
    with open_pcm_wav(source) as reader:
        assert reader.getnchannels() == 1
        assert reader.getsampwidth() == 2
        assert reader.getframerate() == 96000
        assert reader.getnframes() == 4
        assert np.frombuffer(reader.readframes(4), dtype="<i2").tolist() == (
            expected_samples.tolist()
        )

    waveform_samples, sample_rate = _read_pcm_wav(source)
    assert sample_rate == 96000
    assert waveform_samples.tolist() == expected_samples.tolist()

    destination = tmp_path / "cut.wav"
    metrics = cut_pcm_wav(
        source,
        destination,
        start_sample=1,
        end_sample=3,
        fade_ms=0.0,
    )
    assert metrics["frame_count"] == 2
    with wave.open(str(destination), "rb") as reader:
        assert np.frombuffer(reader.readframes(2), dtype="<i2").tolist() == [
            -4000,
            4000,
        ]


def test_playback_samples_downmixes_and_resamples() -> None:
    stereo = np.array(
        [
            [-32768, 32767],
            [16384, 16384],
        ],
        dtype=np.int16,
    )

    result = _playback_samples(stereo, 24000, output_sample_rate=48000)

    assert result.dtype == np.float32
    assert result.shape == (4,)
    assert result[0] == pytest.approx(-1 / 65536)
    assert result[-1] == pytest.approx(0.5)


def test_candidate_take_groups_use_base_lineage_and_custom_edit_source() -> None:
    candidates = [
        {
            "segment_id": "session__t00001_0000000010_0000000090",
            "session_id": "session",
            "base_indices": [0],
            "score": 88.0,
        },
        {
            "segment_id": "session__s00002",
            "session_id": "session",
            "base_indices": [1],
            "score": 91.0,
        },
        {
            "segment_id": "session__s00001",
            "session_id": "session",
            "base_indices": [0],
            "score": 88.0,
        },
        {
            "segment_id": "session__ecustom",
            "session_id": "session",
            "base_indices": [0, 1],
            "score": 88.0,
            "manual_edit": True,
            "edited_from_segment_id": "session__t00001_0000000010_0000000090",
        },
    ]

    groups = _candidate_take_groups(candidates)

    assert [[item["segment_id"] for item in group] for group in groups] == [
        [
            "session__t00001_0000000010_0000000090",
            "session__s00001",
            "session__ecustom",
        ],
        ["session__s00002"],
    ]


def test_candidate_take_groups_parse_legacy_base_segment_ids() -> None:
    groups = _candidate_take_groups(
        [
            {"segment_id": "session_name__s00007", "score": 80.0},
            {
                "segment_id": "session_name__eold",
                "score": 80.0,
                "manual_edit": True,
                "edited_from_segment_id": "session_name__s00007",
            },
        ]
    )

    assert len(groups) == 1


def test_candidate_take_groups_cluster_shifted_spans_around_acoustic_roots() -> None:
    segments = []
    cursor = 0.0
    gaps_after = [0.1, 0.0, 1.5, 0.2, 0.0, 0.05, 1.5, 0.4, 0.05, 1.8]
    for gap_after in gaps_after:
        segments.append(
            {
                "start_seconds": cursor,
                "end_seconds": cursor + 1.0,
            }
        )
        cursor += 1.0 + gap_after

    def candidate(segment_id: str, start: int, end: int, score: float) -> dict:
        return {
            "segment_id": segment_id,
            "session_id": "session",
            "base_indices": list(range(start, end + 1)),
            "score": score,
        }

    candidates = [
        candidate("session__m00003_00006", 2, 5, 95.65),
        candidate("session__m00007_00009", 6, 8, 95.65),
        candidate("session__m00008_00010", 7, 9, 95.65),
        candidate("session__m00004_00007", 3, 6, 95.65),
        candidate("session__m00001_00003", 0, 2, 92.98),
        candidate("session__m00001_00002", 0, 1, 92.98),
        candidate("session__m00004_00006", 3, 5, 95.65),
        candidate("session__m00004_00008", 3, 7, 94.21),
        candidate("session__m00008_00009", 7, 8, 95.65),
    ]

    groups = _candidate_take_groups(
        candidates,
        {"session": segments},
    )
    groups_by_root = {
        group[0]["segment_id"]: {
            candidate["segment_id"] for candidate in group
        }
        for group in groups
    }

    assert set(groups_by_root) == {
        "session__m00001_00003",
        "session__m00004_00007",
        "session__m00008_00010",
    }
    assert "session__m00003_00006" in groups_by_root[
        "session__m00004_00007"
    ]
    assert "session__m00007_00009" in groups_by_root[
        "session__m00008_00010"
    ]
    assert "session__m00001_00002" in groups_by_root[
        "session__m00001_00003"
    ]


def test_candidate_take_groups_keep_repetitions_inside_one_base_separate() -> None:
    segments = [{"start_seconds": 10.0, "end_seconds": 13.0}]

    def candidate(
        segment_id: str,
        start: float,
        end: float,
        score: float,
        *,
        repeated: bool = False,
    ) -> dict:
        return {
            "segment_id": segment_id,
            "session_id": "session",
            "base_indices": [0],
            "start_seconds": start,
            "end_seconds": end,
            "score": score,
            "intra_segment_trim": True,
            "repeated_take_trim": repeated,
        }

    candidates = [
        candidate("session__take1", 10.0, 11.5, 100.0, repeated=True),
        candidate("session__take1_shifted", 9.95, 11.6, 96.0, repeated=True),
        candidate("session__take2", 11.5, 13.0, 99.0, repeated=True),
        candidate("session__whole", 10.0, 13.0, 70.0),
        {
            **candidate("session__custom", 11.7, 12.8, 98.0),
            "manual_edit": True,
            "edited_from_segment_id": "session__take2",
        },
    ]

    groups = _candidate_take_groups(candidates, {"session": segments})
    groups_by_root = {
        group[0]["segment_id"]: {
            candidate["segment_id"] for candidate in group
        }
        for group in groups
    }

    assert set(groups_by_root) == {"session__take1", "session__take2"}
    assert "session__take1_shifted" in groups_by_root["session__take1"]
    assert "session__custom" in groups_by_root["session__take2"]


def test_manual_only_normal_line_keeps_unmatched_take_pool_visible() -> None:
    line = {
        "type": "normal",
        "status": "REVIEW",
        "candidates": [
            {
                "segment_id": "session__ecustom",
                "manual_edit": True,
            }
        ],
    }

    assert _uses_unmatched_candidates(line) is True


def test_base_segment_usage_marks_partial_and_merged_candidates() -> None:
    review_data = {
        "lines": [
            {
                "line_id": "current",
                "candidates": [
                    {
                        "segment_id": "session__partial",
                        "session_id": "session",
                        "base_indices": [2],
                    },
                    {
                        "segment_id": "session__merged",
                        "session_id": "session",
                        "base_indices": [4, 5],
                    },
                ],
            },
            {
                "line_id": "other",
                "candidates": [
                    {
                        "segment_id": "session__s00002",
                        "session_id": "session",
                        "base_indices": [1],
                    },
                    {
                        "segment_id": "session__overlap",
                        "session_id": "session",
                        "base_indices": [5, 6],
                    },
                ],
            },
        ]
    }

    current, other = _base_segment_candidate_usage(review_data, "current")

    assert current == {("session", 2), ("session", 4), ("session", 5)}
    assert other == {("session", 1), ("session", 5), ("session", 6)}


def test_first_line_base_segment_prefers_earliest_candidate_by_time() -> None:
    line = {
        "selected_segment_id": "selected",
        "suggested_segment_id": "suggested",
        "candidates": [
            {
                "segment_id": "suggested",
                "session_id": "session",
                "base_indices": [8],
                "start_seconds": 12.0,
                "score": 70.0,
            },
            {
                "segment_id": "selected",
                "session_id": "session",
                "base_indices": [4, 5],
                "start_seconds": 20.0,
                "score": 99.0,
            },
            {
                "segment_id": "earliest",
                "session_id": "session",
                "base_indices": [2, 3],
                "start_seconds": 5.0,
                "score": 50.0,
            },
        ],
    }

    assert _first_line_base_segment_key(line) == ("session", 2)


def test_add_base_segment_candidate_persists_and_survives_reprocessing(
    tmp_path: Path,
) -> None:
    source_line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 3,
        "line": "A manually added take.",
        "target_filename": "manual_take",
    }
    review_path = tmp_path / "line_review.json"
    review = build_line_review(
        source_lines=[source_line],
        candidates_by_line={},
        unmatched_segments=[],
    )
    save_line_review(review_path, review)
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "sessions": [
                {
                    "session_id": "session",
                    "audio": "source.wav",
                    "segments": [
                        {
                            "segment_id": "session__s00001",
                            "kind": "base",
                            "session_id": "session",
                            "source_audio": "source.wav",
                            "base_indices": [0],
                            "start_seconds": 1.0,
                            "end_seconds": 2.25,
                            "file": "segments/session/session__s00001.wav",
                            "transcript": "A manually added take.",
                            "metrics": {"duration_seconds": 1.25},
                        }
                    ],
                }
            ]
        },
    )

    candidate = add_base_segment_candidate(
        project_dir=tmp_path,
        review_path=review_path,
        review_data=review,
        line_id=source_line["line_id"],
        segment_id="session__s00001",
    )

    assert candidate["manually_added_base_segment"] is True
    assert candidate["base_indices"] == [0]
    assert review["lines"][0]["status"] == "REVIEW"
    assert load_line_review(review_path)["lines"][0]["candidates"] == [candidate]

    regenerated = build_line_review(
        source_lines=[source_line],
        candidates_by_line={},
        unmatched_segments=[],
    )
    preserved = preserve_manual_selections(regenerated, review)
    assert preserved["lines"][0]["candidates"][0]["segment_id"] == (
        "session__s00001"
    )

    with pytest.raises(ValueError, match="already represented"):
        add_base_segment_candidate(
            project_dir=tmp_path,
            review_path=review_path,
            review_data=review,
            line_id=source_line["line_id"],
            segment_id="session__s00001",
        )


def test_refresh_project_audio_recuts_existing_spans_without_reprocessing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.segmentation as segmentation_module

    def fake_probe_audio(
        path: Path,
        *,
        include_hash: bool = True,
    ) -> dict[str, Any]:
        with wave.open(str(path), "rb") as reader:
            result = {
                "path": str(path.resolve()),
                "size_bytes": path.stat().st_size,
                "duration_seconds": reader.getnframes() / reader.getframerate(),
                "codec": "pcm_s16le",
                "sample_format": "s16",
                "sample_rate": reader.getframerate(),
                "channels": reader.getnchannels(),
                "channel_layout": "mono",
                "bits_per_sample": reader.getsampwidth() * 8,
            }
        if include_hash:
            result["sha256"] = sha256_file(path)
        return result

    monkeypatch.setattr(segmentation_module, "probe_audio", fake_probe_audio)

    source = tmp_path / "source.wav"
    _write_tone(source, duration_seconds=2.0, amplitude=4000)
    old_source_hash = sha256_file(source)
    base_file = tmp_path / "segments" / "session" / "session__s00001.wav"
    derived_file = (
        tmp_path
        / "segments"
        / "session"
        / "session__t00001_0000024000_0000072000.wav"
    )
    base_metrics = cut_pcm_wav(
        source,
        base_file,
        start_sample=0,
        end_sample=48000,
        fade_ms=0.0,
    )
    derived_metrics = cut_pcm_wav(
        source,
        derived_file,
        start_sample=24000,
        end_sample=72000,
        fade_ms=0.0,
    )
    project = {
        "audio_dir": ".",
        "audio_inventory": "audio_inventory.json",
        "sessions": [
            {
                "id": "session",
                "enabled": True,
                "audio": "source.wav",
            }
        ],
        "export": {
            "sample_rate": 48000,
            "channels": 1,
            "bits_per_sample": 16,
        },
        "segmentation": {
            "fade_ms": 0.0,
            "voice_boundary_detection_enabled": False,
        },
    }
    write_json(
        tmp_path / "audio_inventory.json",
        {"files": [fake_probe_audio(source)]},
    )
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "schema_version": 1,
            "settings": project["segmentation"],
            "sessions": [
                {
                    "session_id": "session",
                    "cache_key": "preserved-segmentation-cache-key",
                    "audio": "source.wav",
                    "working_audio": "source.wav",
                    "normalized_source": False,
                    "source_sha256": old_source_hash,
                    "sample_rate": 48000,
                    "source_frames": 96000,
                    "voice_boundary_detection": {"enabled": False},
                    "segments": [
                        {
                            "segment_id": "session__s00001",
                            "kind": "base",
                            "session_id": "session",
                            "source_audio": "source.wav",
                            "source_sha256": old_source_hash,
                            "base_indices": [0],
                            "start_sample": 0,
                            "end_sample": 48000,
                            "start_seconds": 0.0,
                            "end_seconds": 1.0,
                            "file": base_file.relative_to(tmp_path).as_posix(),
                            "transcript": "Keep this transcript",
                            "segment_asr": {"primary": {"transcript": "Keep this transcript"}},
                            "voice_bounds": None,
                            "metrics": base_metrics,
                        }
                    ],
                    "derived_segments": [
                        {
                            "segment_id": (
                                "session__t00001_0000024000_0000072000"
                            ),
                            "kind": "trimmed",
                            "session_id": "session",
                            "source_audio": "source.wav",
                            "source_sha256": old_source_hash,
                            "base_indices": [0],
                            "start_sample": 24000,
                            "end_sample": 72000,
                            "start_seconds": 0.5,
                            "end_seconds": 1.5,
                            "file": derived_file.relative_to(tmp_path).as_posix(),
                            "transcript": "Keep the derived transcript",
                            "candidate_asr": {
                                "transcript": "Keep the derived transcript"
                            },
                            "metrics": derived_metrics,
                        }
                    ],
                }
            ],
        },
    )
    review_payload = {
        "schema_version": 1,
        "lines": [],
        "unmatched_segments": [],
    }
    write_json(tmp_path / "line_review.json", review_payload)

    old_base_hash = sha256_file(base_file)
    _write_tone(source, duration_seconds=2.1, amplitude=12000)
    with pytest.raises(ValueError, match="duration changed"):
        refresh_project_audio(project_dir=tmp_path, project=project)
    assert sha256_file(base_file) == old_base_hash
    assert (
        read_json(tmp_path / "audio_inventory.json")["files"][0]["sha256"]
        == old_source_hash
    )

    _write_tone(source, duration_seconds=2.0, amplitude=12000)
    new_source_hash = sha256_file(source)
    assert new_source_hash != old_source_hash

    result = refresh_project_audio(project_dir=tmp_path, project=project)

    assert result["changed_source_count"] == 1
    assert result["segment_count"] == 1
    assert result["derived_segment_count"] == 1
    inventory = read_json(tmp_path / "audio_inventory.json")
    assert inventory["files"][0]["sha256"] == new_source_hash

    manifest_session = read_json(tmp_path / "segments_manifest.json")["sessions"][0]
    assert manifest_session["source_sha256"] == new_source_hash
    assert manifest_session["cache_key"] == "preserved-segmentation-cache-key"
    base_segment = manifest_session["segments"][0]
    derived_segment = manifest_session["derived_segments"][0]
    assert base_segment["source_sha256"] == new_source_hash
    assert derived_segment["source_sha256"] == new_source_hash
    assert base_segment["transcript"] == "Keep this transcript"
    assert base_segment["segment_asr"]["primary"]["transcript"] == (
        "Keep this transcript"
    )
    assert derived_segment["candidate_asr"]["transcript"] == (
        "Keep the derived transcript"
    )
    assert base_segment["metrics"]["rms_dbfs"] > base_metrics["rms_dbfs"] + 8.0
    assert derived_segment["metrics"]["rms_dbfs"] > (
        derived_metrics["rms_dbfs"] + 8.0
    )
    assert read_json(tmp_path / "line_review.json") == review_payload

    with wave.open(str(source), "rb") as reader:
        source_samples = np.frombuffer(
            reader.readframes(reader.getnframes()),
            dtype="<i2",
        )
    with wave.open(str(base_file), "rb") as reader:
        base_samples = np.frombuffer(
            reader.readframes(reader.getnframes()),
            dtype="<i2",
        )
    with wave.open(str(derived_file), "rb") as reader:
        derived_samples = np.frombuffer(
            reader.readframes(reader.getnframes()),
            dtype="<i2",
        )
    np.testing.assert_array_equal(base_samples, source_samples[:48000])
    np.testing.assert_array_equal(derived_samples, source_samples[24000:72000])


def _write_pattern(path: Path) -> None:
    sample_rate = 48000
    total = np.zeros(sample_rate * 5, dtype="<i2")
    for start, end in ((0.25, 0.9), (1.4, 2.05), (3.0, 3.75)):
        first = int(start * sample_rate)
        last = int(end * sample_rate)
        time = np.arange(last - first) / sample_rate
        total[first:last] = np.rint(
            np.sin(2 * math.pi * 440 * time) * 8000
        ).astype("<i2")
    with wave.open(str(path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(2)
        writer.setframerate(sample_rate)
        writer.writeframes(total.tobytes())


def test_materialize_trimmed_segment_supports_multi_base_bounds(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.wav"
    _write_tone(source, duration_seconds=3.0)
    base_segments = [
        {
            "start_sample": 0,
            "end_sample": 72000,
            "asr_probability": 0.9,
        },
        {
            "start_sample": 72000,
            "end_sample": 144000,
            "asr_probability": 0.8,
        },
    ]
    session_entry = {
        "session_id": "session",
        "audio": "source.wav",
        "source_sha256": "source-hash",
        "sample_rate": 48000,
        "derived_segments": [],
    }

    bounded = materialize_trimmed_segment(
        project_dir=tmp_path,
        project={"segmentation": {"fade_ms": 0.0}},
        session_entry=session_entry,
        base_segments=base_segments,
        base_index=0,
        count=2,
        start_sample=24000,
        end_sample=120000,
        transcript="Complete take",
        asr_probability=None,
    )

    assert bounded["kind"] == "bounded"
    assert bounded["base_indices"] == [0, 1]
    assert bounded["start_sample"] == 24000
    assert bounded["end_sample"] == 120000
    assert bounded["metrics"]["duration_seconds"] == pytest.approx(2.0)
    assert (tmp_path / bounded["file"]).is_file()


def test_word_gap_boundary_snaps_after_release_burst(tmp_path: Path) -> None:
    sample_rate = 48000
    audio_path = tmp_path / "boundary.wav"
    samples = np.full(round(0.50 * sample_rate), 1000, dtype="<i2")
    samples[round(0.15 * sample_rate) : round(0.24 * sample_rate)] = 200
    samples[round(0.24 * sample_rate) : round(0.29 * sample_rate)] = 5000
    samples[round(0.29 * sample_rate) : round(0.39 * sample_rate)] = 20
    with wave.open(str(audio_path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(2)
        writer.setframerate(sample_rate)
        writer.writeframes(samples.tobytes())

    proposed = round(0.27 * sample_rate)
    snapped = quietest_pcm_boundary(
        audio_path,
        proposed_sample=proposed,
        minimum_sample=round(0.12 * sample_rate),
        maximum_sample=round(0.42 * sample_rate),
    )
    assert round(0.30 * sample_rate) <= snapped <= round(
        0.38 * sample_rate
    )

    boundaries = _word_gap_boundaries(
        {
            "file": audio_path.name,
            "start_sample": 0,
            "end_sample": samples.shape[0],
            "start_seconds": 0.0,
            "end_seconds": samples.shape[0] / sample_rate,
            "segment_asr": {
                "primary": {
                    "words": [
                        {"word": " not", "start": 0.0, "end": 0.12},
                        {"word": " I", "start": 0.42, "end": 0.50},
                    ]
                }
            },
        },
        minimum_gap=0.10,
        maximum_boundaries=1,
        segmentation_settings={},
        project_dir=tmp_path,
    )
    assert boundaries[0][1] == pytest.approx(snapped / sample_rate)

    regions = split_regions_on_word_gaps(
        [
            {
                "speech_start": 0.0,
                "speech_end": 0.5,
                "start": 0.0,
                "end": 0.5,
            }
        ],
        {
            "segments": [
                {
                    "start": 0.0,
                    "end": 0.5,
                    "text": "not I",
                    "words": [
                        {"word": " not", "start": 0.0, "end": 0.12},
                        {"word": " I", "start": 0.42, "end": 0.50},
                    ],
                }
            ]
        },
        duration_seconds=0.5,
        settings={
            "word_split_enabled": True,
            "word_split_gap_seconds": 0.10,
            "word_split_min_region_seconds": 0.10,
            "word_split_max_boundaries": 1,
            "word_split_max_segment_seconds": 8.0,
            "minimum_segment_seconds": 0.05,
            "pre_padding_seconds": 0.0,
            "post_padding_seconds": 0.0,
            "word_split_snap_enabled": True,
        },
        audio_path=audio_path,
        sample_rate=sample_rate,
    )
    assert len(regions) == 2
    assert regions[0]["speech_end"] == pytest.approx(snapped / sample_rate)
    assert regions[1]["speech_start"] == pytest.approx(snapped / sample_rate)


def test_required_quiet_word_gap_does_not_fall_back_to_loud_midpoint(
    tmp_path: Path,
) -> None:
    sample_rate = 48000
    audio_path = tmp_path / "loud_gap.wav"
    samples = np.full(round(1.0 * sample_rate), 4000, dtype="<i2")
    with wave.open(str(audio_path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(2)
        writer.setframerate(sample_rate)
        writer.writeframes(samples.tobytes())

    assert quietest_pcm_boundary(
        audio_path,
        proposed_sample=round(0.50 * sample_rate),
        minimum_sample=round(0.30 * sample_rate),
        maximum_sample=round(0.70 * sample_rate),
        require_quiet=True,
    ) is None
    assert _word_gap_boundaries(
        {
            "file": audio_path.name,
            "start_sample": 0,
            "end_sample": samples.shape[0],
            "start_seconds": 0.0,
            "end_seconds": 1.0,
            "segment_asr": {
                "primary": {
                    "words": [
                        {"word": " previous", "start": 0.0, "end": 0.30},
                        {"word": " target", "start": 0.70, "end": 1.0},
                    ]
                }
            },
        },
        minimum_gap=0.30,
        maximum_boundaries=1,
        project_dir=tmp_path,
    ) == []


def test_first_quiet_pcm_boundary_finds_gap_before_breath(tmp_path: Path) -> None:
    sample_rate = 48000
    audio_path = tmp_path / "breath-boundary.wav"
    samples = np.full(sample_rate, 1200, dtype="<i2")
    samples[round(0.20 * sample_rate) : round(0.32 * sample_rate)] = 5000
    samples[round(0.32 * sample_rate) : round(0.38 * sample_rate)] = 10
    samples[round(0.38 * sample_rate) : round(0.70 * sample_rate)] = 800
    with wave.open(str(audio_path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(2)
        writer.setframerate(sample_rate)
        writer.writeframes(samples.tobytes())

    boundary = first_quiet_pcm_boundary(
        audio_path,
        start_sample=round(0.20 * sample_rate),
        end_sample=round(0.70 * sample_rate),
    )

    assert boundary is not None
    assert round(0.32 * sample_rate) <= boundary <= round(0.38 * sample_rate)


def test_first_quiet_pcm_boundary_rejects_tail_without_quiet_gap(
    tmp_path: Path,
) -> None:
    sample_rate = 48000
    audio_path = tmp_path / "no-quiet-boundary.wav"
    samples = np.full(sample_rate, 1200, dtype="<i2")
    with wave.open(str(audio_path), "wb") as writer:
        writer.setnchannels(1)
        writer.setsampwidth(2)
        writer.setframerate(sample_rate)
        writer.writeframes(samples.tobytes())

    assert (
        first_quiet_pcm_boundary(
            audio_path,
            start_sample=round(0.20 * sample_rate),
            end_sample=round(0.70 * sample_rate),
        )
        is None
    )


def test_boundary_voice_trim_creates_clean_candidate_and_marks_original(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (sample_rate, sample_rate * 9),
    )
    snap_calls = []

    def fake_quiet_boundary(_path: Path, **kwargs) -> int:
        snap_calls.append(kwargs)
        return int(kwargs["start_sample"]) + round(0.03 * sample_rate)

    monkeypatch.setattr(
        alignment_module,
        "first_quiet_pcm_boundary",
        fake_quiet_boundary,
    )
    line_text = "Never mind, a dumb joke. I meant nothing by it."
    lines = [{"line_id": "target", "line": line_text}]
    action = {
        "type": "assigned",
        "start_index": 0,
        "count": 2,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 80.0,
        "order_hint": 0.0,
        "top_matches": [],
    }
    base_segments = [
        {
            "start_sample": 0,
            "end_sample": sample_rate * 5,
            "start_seconds": 0.0,
            "end_seconds": 5.0,
        },
        {
            "start_sample": sample_rate * 5,
            "end_sample": sample_rate * 10,
            "start_seconds": 5.0,
            "end_seconds": 10.0,
        },
    ]
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=lines,
        base_segments=base_segments,
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    assert action["unclean_boundary_audio"] is True
    assert len(cleaned) == 1
    assert cleaned[0]["boundary_voice_trim"] is True
    assert cleaned[0]["unclean_boundary_audio"] is False
    assert cleaned[0]["trim_start_sample"] == sample_rate - round(
        0.08 * sample_rate
    )
    proposed_end = sample_rate * 9 + round(0.12 * sample_rate)
    assert cleaned[0]["trim_end_sample"] == proposed_end + round(
        0.03 * sample_rate
    )
    assert snap_calls[0]["start_sample"] == proposed_end
    assert snap_calls[0]["end_sample"] == round(9.35 * sample_rate)


def test_boundary_voice_trim_preserves_quiet_first_word(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    def fake_voice_bounds(*_args, **kwargs):
        if float(kwargs.get("threshold", 0.5)) < 0.5:
            return (0, round(0.25 * sample_rate))
        return (sample_rate, sample_rate * 10)

    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        fake_voice_bounds,
    )
    line_text = (
        "It is a place of love first and foremost. "
        "It should not be left broken like a widow's heart."
    )
    action = {
        "type": "assigned",
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 100.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[{"line_id": "target", "line": line_text}],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": sample_rate * 10,
                "start_seconds": 0.0,
                "end_seconds": 10.0,
                "transcript_source": "segment_asr",
                "words": [
                    {"word": " It", "start": 0.0, "end": 0.25},
                    {"word": " heart", "start": 8.5, "end": 9.2},
                ],
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator(
            [{"line_id": "target", "line": line_text}],
            {},
        ),
    )

    assert cleaned == []
    assert "unclean_boundary_audio" not in action


def test_boundary_voice_trim_ignores_zero_clamped_stretched_first_word(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (
            round(0.93 * sample_rate),
            round(3.10 * sample_rate),
        ),
    )
    line_text = "What now, milk-drinker?"
    line = {"line_id": "target", "line": line_text}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 50.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[line],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": round(3.68 * sample_rate),
                "start_seconds": 0.0,
                "end_seconds": 3.68,
                "segment_asr": {
                    "primary": {
                        "words": [
                            {"word": " What", "start": 0.0, "end": 1.04},
                            {
                                "word": " drinker",
                                "start": 2.25,
                                "end": 2.92,
                            },
                        ]
                    }
                },
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
    )

    assert cleaned[0]["trim_start_sample"] == round(0.85 * sample_rate)


def test_boundary_voice_trim_does_not_trust_shifted_asr_word_start(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (
            round(0.70 * sample_rate),
            round(3.0 * sample_rate),
        ),
    )
    line_text = "Am I going crazy?"
    line = {"line_id": "target", "line": line_text}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 100.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[line],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": round(3.0 * sample_rate),
                "start_seconds": 0.0,
                "end_seconds": 3.0,
                "segment_asr": {
                    "primary": {
                        "words": [
                            {"word": " Am", "start": 0.24, "end": 0.76},
                            {"word": " crazy", "start": 2.5, "end": 3.0},
                        ]
                    }
                },
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
    )

    assert cleaned[0]["trim_start_sample"] == round(0.62 * sample_rate)


def test_boundary_voice_trim_ignores_word_crossing_derived_cut(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    raw_start = sample_rate
    raw_end = sample_rate * 4
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (
            raw_start + round(0.60 * sample_rate),
            raw_end,
        ),
    )
    line_text = "Yeah, yeah."
    line = {"line_id": "target", "line": line_text}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 100.0,
        "trim_start_sample": raw_start,
        "trim_end_sample": raw_end,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[line],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": raw_end,
                "start_seconds": 0.0,
                "end_seconds": 4.0,
                "segment_asr": {
                    "primary": {
                        "words": [
                            {"word": " yeah", "start": 0.8, "end": 1.05},
                            {"word": " yeah", "start": 1.7, "end": 2.1},
                        ]
                    }
                },
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
    )

    assert cleaned[0]["trim_start_sample"] == raw_start + round(
        0.52 * sample_rate
    )


def test_boundary_voice_trim_preserves_initial_sibilant(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (
            round(0.70 * sample_rate),
            round(2.80 * sample_rate),
        ),
    )
    monkeypatch.setattr(
        alignment_module,
        "pcm_leading_sibilant_start",
        lambda *_args, **_kwargs: round(0.42 * sample_rate),
    )
    line = {"line_id": "target", "line": "Speak, I haven't got all day."}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line["line"],
        "duration_plausibility": 100.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[line],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": sample_rate * 3,
                "start_seconds": 0.0,
                "end_seconds": 3.0,
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
    )

    assert cleaned[0]["trim_start_sample"] == round(0.34 * sample_rate)


def test_boundary_voice_trim_cleans_secondary_match(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (sample_rate, sample_rate * 4),
    )
    primary = {"line_id": "other", "line": "Entirely different words."}
    target = {"line_id": "target", "line": "Has to be. Look at her!"}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 20.0,
        "transcript": target["line"],
        "duration_plausibility": 100.0,
        "top_matches": [
            {
                "line_index": 1,
                "match_score": 100.0,
                "confidence_margin": 50.0,
            }
        ],
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[primary, target],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": sample_rate * 5,
                "start_seconds": 0.0,
                "end_seconds": 5.0,
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([primary, target], {}),
    )

    assert action["unclean_boundary_audio"] is True
    assert len(cleaned) == 1
    assert cleaned[0]["line_index"] == 1
    assert cleaned[0]["trim_start_sample"] == round(0.92 * sample_rate)


def test_boundary_voice_trim_bounds_incomplete_match_fallback(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (
            round(0.60 * sample_rate),
            round(4.0 * sample_rate),
        ),
    )
    line = {"line_id": "target", "line": "Has to be. Look at her!"}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 60.0,
        "transcript": "Has to be.",
        "duration_plausibility": 100.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[line],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": sample_rate * 4,
                "start_seconds": 0.0,
                "end_seconds": 4.0,
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
    )

    assert len(cleaned) == 1
    assert cleaned[0]["trim_start_sample"] == round(0.52 * sample_rate)


def test_boundary_voice_trim_removes_detached_short_tail(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (
            round(0.20 * sample_rate),
            round(9.20 * sample_rate),
        ),
    )
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_regions",
        lambda *_args, **_kwargs: [
            (round(0.20 * sample_rate), round(8.0 * sample_rate)),
            (round(9.0 * sample_rate), round(9.20 * sample_rate)),
        ],
    )
    line = {"line_id": "target", "line": "Ancient artists left many clues."}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line["line"],
        "duration_plausibility": 100.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[line],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": sample_rate * 10,
                "start_seconds": 0.0,
                "end_seconds": 10.0,
                "segment_asr": {
                    "primary": {
                        "words": [
                            {"word": " Ancient", "start": 0.2, "end": 0.8},
                            {"word": " clues.", "start": 7.2, "end": 7.7},
                        ]
                    }
                },
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
    )

    assert cleaned[0]["detached_trailing_voice_trim"] is True
    assert cleaned[0]["trim_end_sample"] <= round(8.35 * sample_rate)


def test_boundary_voice_trim_recovers_detached_failed_prefix(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    line_text = (
        "I could, but I'd have to import them from Cyrodiil. "
        "They're not too common around here."
    )
    line = {"line_id": "target", "line": line_text}
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        lambda *_args, **_kwargs: (
            round(0.22 * sample_rate),
            round(7.17 * sample_rate),
        ),
    )
    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_regions",
        lambda *_args, **_kwargs: [
            (round(0.22 * sample_rate), round(1.60 * sample_rate)),
            (round(2.14 * sample_rate), round(5.57 * sample_rate)),
            (round(5.76 * sample_rate), round(7.17 * sample_rate)),
        ],
    )
    monkeypatch.setattr(
        alignment_module,
        "first_quiet_pcm_boundary",
        lambda _path, **kwargs: int(kwargs["start_sample"]),
    )
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 79.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=[line],
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": round(8.90 * sample_rate),
                "start_seconds": 0.0,
                "end_seconds": 8.90,
                "segment_asr": {
                    "primary": {
                        "words": [
                            {"word": " I", "start": 0.0, "end": 0.26},
                            {"word": " here", "start": 6.68, "end": 6.94},
                        ]
                    }
                },
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
    )

    detached = next(
        item for item in cleaned if item["detached_leading_voice_trim"]
    )
    assert detached["trim_start_sample"] == round(2.06 * sample_rate)
    assert detached["unclean_boundary_audio"] is False
    assert action["detached_edge_speech"] is True


def test_boundary_voice_trim_uses_quiet_gap_after_strict_tail_vad(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    thresholds = []

    def fake_voice_bounds(
        _path: Path,
        *,
        start_sample: int,
        end_sample: int,
        threshold: float,
    ) -> tuple[int, int]:
        thresholds.append(threshold)
        end_seconds = 9.8 if threshold < 0.7 else 9.0
        return start_sample + sample_rate, round(
            start_sample + end_seconds * sample_rate
        )

    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        fake_voice_bounds,
    )
    quiet_boundary = round(sample_rate * 9.25)
    monkeypatch.setattr(
        alignment_module,
        "first_quiet_pcm_boundary",
        lambda *_args, **_kwargs: quiet_boundary,
    )
    line_text = "A bet's a bet. Here's your gold."
    lines = [{"line_id": "target", "line": line_text}]
    action = {
        "type": "assigned",
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 100.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
        },
        lines=lines,
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": sample_rate * 10,
                "start_seconds": 0.0,
                "end_seconds": 10.0,
                "words": [
                    {"word": " gold.", "start": 7.5, "end": 8.8},
                ],
            }
        ],
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    assert thresholds == [0.5, 0.7]
    assert len(cleaned) == 1
    assert cleaned[0]["trim_end_sample"] == quiet_boundary
    assert action["boundary_voice_trailing_seconds"] == pytest.approx(1.0)


def test_boundary_voice_trim_reuses_segmentation_voice_bounds(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000

    def unexpected_live_vad(*_args, **_kwargs):
        raise AssertionError("alignment should reuse segmentation VAD metadata")

    monkeypatch.setattr(
        alignment_module,
        "pcm_voice_bounds",
        unexpected_live_vad,
    )
    line_text = "A bet's a bet. Here's your gold."
    lines = [{"line_id": "target", "line": line_text}]
    action = {
        "type": "assigned",
        "start_index": 0,
        "count": 2,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 100.0,
    }
    cleaned = _boundary_voice_trim_actions(
        [action],
        project_dir=tmp_path,
        session_entry={
            "sample_rate": sample_rate,
            "working_audio": "source.wav",
            "audio": "source.wav",
            "voice_boundary_detection": {
                "enabled": True,
                "vad_threshold": 0.5,
                "breath_vad_threshold": 0.7,
            },
        },
        lines=lines,
        base_segments=[
            {
                "start_sample": 0,
                "end_sample": sample_rate * 5,
                "start_seconds": 0.0,
                "end_seconds": 5.0,
                "voice_bounds": {
                    "source": "silero_vad",
                    "speech": {
                        "start_sample": sample_rate,
                        "end_sample": round(sample_rate * 4.8),
                    },
                    "strict_speech": {
                        "start_sample": sample_rate,
                        "end_sample": round(sample_rate * 4.5),
                    },
                },
            },
            {
                "start_sample": sample_rate * 5,
                "end_sample": sample_rate * 10,
                "start_seconds": 5.0,
                "end_seconds": 10.0,
                "segment_asr": {
                    "primary": {
                        "words": [
                            {
                                "word": " gold.",
                                "start": 2.5,
                                "end": 3.8,
                            },
                        ]
                    }
                },
                "voice_bounds": {
                    "source": "silero_vad",
                    "speech": {
                        "start_sample": round(sample_rate * 5.2),
                        "end_sample": round(sample_rate * 9.8),
                    },
                    "strict_speech": {
                        "start_sample": round(sample_rate * 5.2),
                        "end_sample": sample_rate * 9,
                    },
                },
            },
        ],
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
        segmentation_settings={
            "voice_boundary_detection_enabled": True,
            "voice_boundary_vad_threshold": 0.5,
            "voice_boundary_breath_vad_threshold": 0.7,
        },
    )

    assert len(cleaned) == 1
    assert cleaned[0]["trim_start_sample"] == sample_rate - round(
        0.08 * sample_rate
    )
    assert cleaned[0]["trim_end_sample"] == round(9.35 * sample_rate)


def test_shared_model_cache_resolution(tmp_path: Path, monkeypatch) -> None:
    shared = tmp_path / "shared-model-cache"
    monkeypatch.setenv("DIALOGUE_VA_MODEL_CACHE", str(shared))
    assert default_model_cache_root() == shared.resolve()
    assert resolve_model_cache_root(tmp_path, {}) == shared.resolve()
    assert resolve_model_cache_root(
        tmp_path,
        {"model_cache": "project-override"},
    ) == (tmp_path / "project-override").resolve()


def test_float_wav_normalization(tmp_path: Path) -> None:
    integer_source = tmp_path / "integer.wav"
    float_source = tmp_path / "float.wav"
    normalized = tmp_path / "normalized.wav"
    _write_tone(integer_source, duration_seconds=0.25)
    subprocess.run(
        [
            "ffmpeg",
            "-hide_banner",
            "-loglevel",
            "error",
            "-y",
            "-i",
            str(integer_source),
            "-ar",
            "44100",
            "-c:a",
            "pcm_f32le",
            str(float_source),
        ],
        check=True,
    )

    working, shape, was_normalized = prepare_pcm_segmentation_source(
        float_source,
        normalized,
        sample_rate=48000,
        channels=1,
        bits_per_sample=16,
    )
    assert working == normalized
    assert was_normalized
    assert shape["sample_rate"] == 48000
    assert shape["channels"] == 1
    assert shape["bits_per_sample"] == 16
    with wave.open(str(normalized), "rb") as reader:
        assert reader.getframerate() == 48000
        assert reader.getsampwidth() == 2


def test_sample_workbook_schema() -> None:
    workbook_path = (
        Path(__file__).parents[1] / "MaleElfYoung" / "ARG1RMElfYoung.xlsm"
    )
    result = parse_workbook(workbook_path)
    assert result["sheet_count"] == 21
    assert result["line_count"] == 457
    assert len({line["target_filename"] for line in result["lines"]}) == 457


def test_workbook_parser_supports_repeated_sections_and_note_vocalizations(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "sectioned-lines.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Main"
    headers = [
        "Quest",
        "Context",
        "Line to speak",
        "Acting Note",
        "Filecutting Note",
        "Filename",
    ]
    worksheet.cell(1, 2).value = "You are voicing Actor One"
    for column, header in enumerate(headers, start=1):
        worksheet.cell(2, column).value = header
    worksheet.append(["Quest One", "Greeting", "Hello.", "", "", "one.wav"])

    worksheet.cell(5, 2).value = "You are voicing Actor Two"
    second_headers = [
        "Filename",
        "Acting Note",
        "Line to speak",
        "Context",
        "Quest",
        "Filecutting Note",
    ]
    for column, header in enumerate(second_headers, start=1):
        worksheet.cell(6, column).value = header
    worksheet.append(["two.wav", "Huh...", " ", "Idle", "Quest Two", ""])
    workbook.save(workbook_path)
    workbook.close()

    result = parse_workbook(workbook_path)

    assert result["line_count"] == 2
    assert [line["line_id"] for line in result["lines"]] == [
        "Main::R3",
        "Main::R7",
    ]
    assert [line["quest"] for line in result["lines"]] == [
        "Quest One",
        "Quest Two",
    ]
    assert result["lines"][1]["line"] == "Huh..."
    assert result["lines"][1]["acting_note"] == "Huh..."
    assert result["sheets"][0]["voice_headers"] == [
        "You are voicing Actor One",
        "You are voicing Actor Two",
    ]
    assert "Line to speak" not in {line["line"] for line in result["lines"]}


def test_workbook_parser_supports_creation_kit_ods_context_fallback(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "dialogue.ods"
    content = """<?xml version="1.0" encoding="UTF-8"?>
<office:document-content
    xmlns:office="urn:oasis:names:tc:opendocument:xmlns:office:1.0"
    xmlns:table="urn:oasis:names:tc:opendocument:xmlns:table:1.0"
    xmlns:text="urn:oasis:names:tc:opendocument:xmlns:text:1.0">
  <office:body>
    <office:spreadsheet>
      <table:table table:name="dialogueExport">
        <table:table-row>
          <table:table-cell><text:p>QUEST</text:p></table:table-cell>
          <table:table-cell><text:p>TOPIC</text:p></table:table-cell>
          <table:table-cell><text:p>TOPIC TEXT</text:p></table:table-cell>
          <table:table-cell><text:p>RESPONSE TEXT</text:p></table:table-cell>
          <table:table-cell><text:p>SCRIPT NOTES</text:p></table:table-cell>
          <table:table-cell><text:p>EMOTION</text:p></table:table-cell>
          <table:table-cell><text:p>FILENAME</text:p></table:table-cell>
          <table:table-cell><text:p>VOICE TYPE</text:p></table:table-cell>
        </table:table-row>
        <table:table-row>
          <table:table-cell><text:p>QuestOne</text:p></table:table-cell>
          <table:table-cell><text:p>InternalTopic</text:p></table:table-cell>
          <table:table-cell><text:p>Player-facing prompt</text:p></table:table-cell>
          <table:table-cell><text:p>First response.</text:p></table:table-cell>
          <table:table-cell><text:p>Friendly</text:p></table:table-cell>
          <table:table-cell><text:p>Neutral 50</text:p></table:table-cell>
          <table:table-cell><text:p>first_line</text:p></table:table-cell>
          <table:table-cell><text:p>VoiceTypeOne</text:p></table:table-cell>
        </table:table-row>
        <table:table-row>
          <table:table-cell/>
          <table:table-cell><text:p>FallbackTopic</text:p></table:table-cell>
          <table:table-cell/>
          <table:table-cell><text:p>Second response.</text:p></table:table-cell>
          <table:table-cell/>
          <table:table-cell><text:p>Anger 25</text:p></table:table-cell>
          <table:table-cell><text:p>second_line</text:p></table:table-cell>
          <table:table-cell><text:p>VoiceTypeOne</text:p></table:table-cell>
        </table:table-row>
        <table:table-row>
          <table:table-cell/>
          <table:table-cell/>
          <table:table-cell/>
          <table:table-cell/>
          <table:table-cell><text:p>Hmm...</text:p></table:table-cell>
          <table:table-cell><text:p>Neutral 0</text:p></table:table-cell>
          <table:table-cell><text:p>third_line</text:p></table:table-cell>
          <table:table-cell><text:p>VoiceTypeOne</text:p></table:table-cell>
        </table:table-row>
        <table:table-row table:number-rows-repeated="1048572">
          <table:table-cell table:number-columns-repeated="8"/>
        </table:table-row>
      </table:table>
    </office:spreadsheet>
  </office:body>
</office:document-content>
"""
    with ZipFile(workbook_path, "w") as archive:
        archive.writestr(
            "mimetype",
            "application/vnd.oasis.opendocument.spreadsheet",
            compress_type=ZIP_STORED,
        )
        archive.writestr("content.xml", content)

    result = parse_workbook(workbook_path)

    assert result["line_count"] == 3
    assert [line["context"] for line in result["lines"]] == [
        "Player-facing prompt",
        "FallbackTopic",
        "",
    ]
    assert result["lines"][1]["quest"] == "QuestOne"
    assert result["lines"][2]["line"] == "Hmm..."
    assert result["lines"][2]["acting_note"] == "Hmm..."
    assert result["sheets"][0]["voice_headers"] == [
        "You are voicing VoiceTypeOne"
    ]
    assert result["sheets"][0]["line_ids"][-1] == "dialogueExport::R4"


def test_new_project_settings_are_validated_and_merged(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.project as project_module

    workbook = tmp_path / "lines.xlsx"
    workbook.touch()
    audio_dir = tmp_path / "audio"
    audio_dir.mkdir()
    _write_tone(audio_dir / "Actor.wav")
    monkeypatch.setattr(
        project_module,
        "parse_workbook",
        lambda _path: {
            "sheet_count": 1,
            "line_count": 0,
            "sheets": [
                {
                    "name": "Actor",
                    "line_count": 0,
                    "voice_header": "",
                }
            ],
            "lines": [],
        },
    )
    values = editable_project_settings()
    values["language"] = "de"
    values["transcription"].update(
        {
            "model": "medium",
            "device": "cuda",
            "compute_type": "float16",
            "batch_size": "8",
            "batch_size_max": "24",
        }
    )
    values["segment_transcription"]["enabled"] = False
    settings = _project_settings_from_values(
        values,
        editable_project_settings(),
    )

    project = create_project(
        workbook_path=workbook,
        audio_dir=audio_dir,
        project_dir=tmp_path / "work",
        project_settings=settings,
    )

    assert project["language"] == "de"
    assert project["transcription"]["model"] == "medium"
    assert project["transcription"]["batch_size"] == 8
    assert project["transcription"]["batch_size_max"] == 24
    assert project["transcription"]["vad_filter"] is True
    assert project["segment_transcription"]["enabled"] is False
    assert project["segment_transcription"]["prompt_fallback_enabled"] is True
    assert set(settings) == {
        "language",
        "transcription",
        "segment_transcription",
        "segmentation",
        "alignment",
        "export",
    }

    invalid = editable_project_settings()
    invalid["transcription"]["batch_size"] = "zero"
    with pytest.raises(ValueError, match="Batch size"):
        _project_settings_from_values(
            invalid,
            editable_project_settings(),
        )


def test_create_button_shows_existing_settings_before_reprocessing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.ui as ui_module

    project_dir = tmp_path / "existing"
    project_dir.mkdir()
    write_json(project_dir / "project.json", {"schema_version": 1})
    app = DialogueReviewApp.__new__(DialogueReviewApp)
    selected = []
    shown = []

    def settings_dialog(settings, *, reprocessing):
        shown.append((settings, reprocessing))
        return settings

    app.ask_project_settings = settings_dialog
    app.run_existing_project = lambda path, settings: selected.append(
        (path, settings)
    )
    app.run_new_project = lambda **_kwargs: pytest.fail(
        "Existing projects must not be initialized again"
    )
    monkeypatch.setattr(
        ui_module.filedialog,
        "askdirectory",
        lambda **_kwargs: str(project_dir),
    )
    monkeypatch.setattr(
        ui_module.filedialog,
        "askopenfilename",
        lambda **_kwargs: pytest.fail(
            "Existing projects must not ask for a workbook"
        ),
    )

    app.choose_new_project()

    assert len(shown) == 1
    assert shown[0][1] is True
    assert shown[0][0] == editable_project_settings({"schema_version": 1})
    assert selected == [(project_dir.resolve(), shown[0][0])]


def test_mapping_row_filter_round_trip_and_validation() -> None:
    assert _format_excel_rows([24, 18, 19, 20, 12, 12]) == "12, 18-20, 24"
    assert _parse_excel_rows("12, 18 - 20, 24") == [12, 18, 19, 20, 24]
    assert _parse_excel_rows("  ") == []
    with pytest.raises(ValueError, match="ends before"):
        _parse_excel_rows("24-18")
    with pytest.raises(ValueError, match="Invalid Excel row"):
        _parse_excel_rows("12, all")

    source_data = {
        "sheets": [
            {"name": "First"},
            {"name": "Second"},
        ],
        "lines": [
            {"line_id": "First::R12", "sheet": "First", "excel_row": 12},
            {"line_id": "Second::R18", "sheet": "Second", "excel_row": 18},
        ],
    }
    sessions = [
        {
            "id": "main",
            "enabled": True,
            "sheets": ["First"],
            "excel_rows": [12],
            "line_ids": [],
        },
        {
            "id": "unused",
            "enabled": False,
            "sheets": [],
            "excel_rows": [],
            "line_ids": [],
        },
    ]
    _validate_session_mappings(sessions, source_data)

    sessions[0]["excel_rows"] = [999]
    with pytest.raises(ValueError, match="contains no script lines"):
        _validate_session_mappings(sessions, source_data)
    sessions[0]["excel_rows"] = []
    sessions[0]["sheets"] = ["Missing"]
    with pytest.raises(ValueError, match="missing script sheet"):
        _validate_session_mappings(sessions, source_data)


def test_mapping_sheet_actions_follow_selection_and_bulk_state() -> None:
    assert _mapping_sheet_action_names(
        mapped_sheets=["First"],
        available_sheets=["First", "Second"],
        selected_sheet="First",
    ) == ("Remove", "Add All")
    assert _mapping_sheet_action_names(
        mapped_sheets=["First"],
        available_sheets=["First", "Second"],
        selected_sheet="Second",
    ) == ("Add", "Add All")
    assert _mapping_sheet_action_names(
        mapped_sheets=["First", "Second"],
        available_sheets=["First", "Second"],
        selected_sheet="Second",
    ) == ("Remove", "Remove All")

    session = {"id": "actor", "sheets": ["Second"]}
    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app._mapping_project = {"sessions": [session]}
    app._mapping_source_data = {
        "sheets": [{"name": "First"}, {"name": "Second"}]
    }
    app._mapping_session_id = "actor"
    refreshes = []
    app._refresh_mapping_sheet_tree = lambda: refreshes.append(True)

    app._toggle_all_mapping_sheets()
    assert session["sheets"] == ["Second", "First"]
    app._toggle_all_mapping_sheets()
    assert session["sheets"] == []
    assert len(refreshes) == 2


def test_new_project_pauses_for_mapping_review_after_inventory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.ui as ui_module

    project_dir = tmp_path / "project"
    calls = []
    monkeypatch.setattr(
        ui_module,
        "create_project",
        lambda **_kwargs: calls.append("inventory") or {},
    )
    monkeypatch.setattr(
        ui_module,
        "transcribe_project",
        lambda **_kwargs: pytest.fail("Transcription must wait for mapping review"),
    )
    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app.show_progress = lambda path, *, title: calls.append((path, title))
    app._inventory_finished = lambda path, *, reprocessing: calls.append(
        (path, reprocessing)
    )
    app._start_worker = lambda function, callback: callback(function())

    app.run_new_project(
        workbook_path=tmp_path / "lines.xlsx",
        audio_dir=tmp_path / "audio",
        project_dir=project_dir,
        project_settings={},
    )

    assert calls == [
        (project_dir, "Building project inventory"),
        "inventory",
        (project_dir.resolve(), False),
    ]


def test_reprocessing_pauses_for_mapping_review_before_transcription(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.ui as ui_module

    project_dir = tmp_path / "project"
    project_dir.mkdir()
    write_json(project_dir / "project.json", {"schema_version": 1})
    monkeypatch.setattr(
        ui_module,
        "transcribe_project",
        lambda **_kwargs: pytest.fail("Transcription must wait for mapping review"),
    )
    calls = []
    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app.root = object()
    app.show_progress = lambda path, *, title: calls.append((path, title))
    app._inventory_finished = lambda path, *, reprocessing: calls.append(
        (path, reprocessing)
    )
    app._start_worker = lambda function, callback: callback(function())

    app.run_existing_project(
        project_dir,
        editable_project_settings({"schema_version": 1}),
    )

    assert calls == [
        (project_dir, "Preparing project for reprocessing"),
        (project_dir.resolve(), True),
    ]


def test_confirmed_mapping_is_persisted_before_processing(tmp_path: Path) -> None:
    session = {
        "id": "actor",
        "enabled": True,
        "audio": "Actor.wav",
        "sheets": ["Second", "First"],
        "excel_rows": [],
        "line_ids": [],
        "needs_mapping_review": True,
    }
    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app.root = object()
    app.project_dir = tmp_path
    app._mapping_project = {"schema_version": 1, "sessions": [session]}
    app._mapping_source_data = {
        "sheets": [{"name": "First"}, {"name": "Second"}],
        "lines": [
            {"line_id": "Second::R18", "sheet": "Second", "excel_row": 18}
        ],
    }
    app._mapping_inventory = {}
    app._mapping_reprocessing = False
    app._mapping_session_id = "actor"
    app._mapping_row_values = {"actor": "18-19"}
    app._mapping_line_id_values = {"actor": []}
    app._remember_mapping_rows = lambda: None
    continued = []
    app._continue_project_processing = (
        lambda path, *, reprocessing: continued.append((path, reprocessing))
    )

    app._confirm_session_mappings()

    saved = read_json(tmp_path / "project.json")
    assert saved["sessions"][0]["sheets"] == ["Second", "First"]
    assert saved["sessions"][0]["excel_rows"] == [18, 19]
    assert saved["sessions"][0]["needs_mapping_review"] is False
    assert continued == [(tmp_path, False)]


def test_forced_metadata_refresh_backs_up_and_preserves_existing_settings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.project as project_module

    project_dir = tmp_path / "work"
    project_dir.mkdir()
    workbook = tmp_path / "lines.xlsx"
    workbook.touch()
    audio_dir = tmp_path / "audio"
    audio_dir.mkdir()
    audio = audio_dir / "Actor.wav"
    audio.touch()
    write_json(
        project_dir / "project.json",
        {
            "schema_version": 1,
            "language": "fr",
            "alignment": {"span_search": {"max_segments": 6}},
            "sessions": [
                {
                    "id": "actor",
                    "audio": "../audio/Actor.wav",
                    "enabled": False,
                    "sheets": ["Manual"],
                    "excel_rows": [],
                    "line_ids": [],
                    "pass": "main",
                    "needs_mapping_review": False,
                }
            ],
        },
    )
    source_data = {
        "sheet_count": 1,
        "line_count": 0,
        "sheets": [
            {
                "name": "Actor",
                "line_count": 0,
                "voice_header": "You are voicing Actor",
            }
        ],
        "lines": [],
    }
    monkeypatch.setattr(project_module, "parse_workbook", lambda _path: source_data)
    monkeypatch.setattr(
        project_module,
        "probe_audio",
        lambda path, include_hash: {
            "path": str(path.resolve()),
            "sha256": "hash",
        },
    )

    refreshed = create_project(
        workbook_path=workbook,
        audio_dir=audio_dir,
        project_dir=project_dir,
        force=True,
    )

    assert refreshed["language"] == "fr"
    assert AlignmentSettings.from_value(refreshed["alignment"])[
        "max_merge_segments"
    ] == 6
    assert refreshed["sessions"][0]["sheets"] == ["Manual"]
    assert refreshed["sessions"][0]["enabled"] is False
    assert read_json(project_dir / "project.before-force.json")["language"] == "fr"


def test_create_button_collects_settings_only_for_new_project(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.ui as ui_module

    project_dir = tmp_path / "new"
    workbook = tmp_path / "lines.xlsx"
    audio_dir = tmp_path / "audio"
    settings = {"language": "en", "transcription": {"model": "small"}}
    selected_directories = iter((str(project_dir), str(audio_dir)))
    monkeypatch.setattr(
        ui_module.filedialog,
        "askdirectory",
        lambda **_kwargs: next(selected_directories),
    )
    monkeypatch.setattr(
        ui_module.filedialog,
        "askopenfilename",
        lambda **_kwargs: str(workbook),
    )
    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app.ask_project_settings = (
        lambda current, *, reprocessing: (
            settings
            if not reprocessing and current == editable_project_settings()
            else pytest.fail("Unexpected settings dialog state")
        )
    )
    captured = {}
    app.run_new_project = lambda **kwargs: captured.update(kwargs)
    app.run_existing_project = lambda _path, _settings: pytest.fail(
        "A new folder must initialize a project"
    )

    app.choose_new_project()

    assert captured == {
        "workbook_path": workbook,
        "audio_dir": audio_dir,
        "project_dir": project_dir.resolve(),
        "project_settings": settings,
    }


def test_existing_project_reprocess_pipeline_preserves_folder(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.ui as ui_module

    project_dir = tmp_path / "existing"
    project_dir.mkdir()
    write_json(
        project_dir / "project.json",
        {"schema_version": 1, "custom_setting": "keep-me"},
    )
    sentinel = project_dir / "do-not-delete.txt"
    sentinel.write_text("preserved", encoding="utf-8")
    calls = []

    def record(phase: str):
        def run(*, project_dir: Path, project: dict, **_kwargs):
            calls.append((phase, project_dir, project["custom_setting"]))

        return run

    monkeypatch.setattr(ui_module, "transcribe_project", record("transcribe"))
    monkeypatch.setattr(ui_module, "segment_project", record("segment"))
    monkeypatch.setattr(
        ui_module,
        "transcribe_segments_project",
        record("transcribe_segments"),
    )
    monkeypatch.setattr(ui_module, "align_project", record("align"))
    monkeypatch.setattr(
        ui_module,
        "create_project",
        lambda **_kwargs: pytest.fail(
            "Reprocessing must not recreate project.json"
        ),
    )

    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app.root = object()
    progress = []
    finished = []
    mapping_review = []
    app.show_progress = lambda path, *, title: progress.append((path, title))
    app._pipeline_finished = lambda path: finished.append(path)
    app._inventory_finished = lambda path, *, reprocessing: mapping_review.append(
        (path, reprocessing)
    )
    app._start_worker = lambda function, callback: callback(function())

    project_settings = editable_project_settings(
        {
            "schema_version": 1,
            "custom_setting": "keep-me",
        }
    )
    project_settings["language"] = "fr"
    app.run_existing_project(project_dir, project_settings)

    assert calls == []
    assert progress == [(project_dir, "Preparing project for reprocessing")]
    assert mapping_review == [(project_dir.resolve(), True)]
    saved_project = read_json(project_dir / "project.json")
    assert saved_project["custom_setting"] == "keep-me"
    assert saved_project["language"] == "fr"

    app._continue_project_processing(project_dir, reprocessing=True)

    assert [phase for phase, _path, _setting in calls] == [
        "transcribe",
        "segment",
        "transcribe_segments",
        "align",
    ]
    assert all(path == project_dir.resolve() for _phase, path, _setting in calls)
    assert all(setting == "keep-me" for _phase, _path, setting in calls)
    assert progress == [
        (project_dir, "Preparing project for reprocessing"),
        (project_dir, "Reprocessing project"),
    ]
    assert finished == [project_dir.resolve()]
    assert sentinel.read_text(encoding="utf-8") == "preserved"


def test_processing_cancellation_scope_is_thread_local() -> None:
    event = threading.Event()
    with cancellation_scope(event):
        check_processing_cancelled()
        event.set()
        with pytest.raises(ProcessingCancelled):
            check_processing_cancelled()
    check_processing_cancelled()


def test_project_creation_stops_before_writing_when_already_cancelled(
    tmp_path: Path,
) -> None:
    workbook = tmp_path / "lines.xlsx"
    workbook.touch()
    audio_dir = tmp_path / "audio"
    audio_dir.mkdir()
    event = threading.Event()
    event.set()

    with cancellation_scope(event), pytest.raises(ProcessingCancelled):
        create_project(
            workbook_path=workbook,
            audio_dir=audio_dir,
            project_dir=tmp_path / "work",
        )

    assert not (tmp_path / "work").exists()


def test_processing_screen_cancel_signals_worker_and_returns_to_start() -> None:
    class FakeRoot:
        def after(self, _delay, _callback):
            return None

    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app.root = FakeRoot()
    app._cancel_event = threading.Event()
    app._cancel_button = None
    app._cancel_status = None
    app._worker_messages = None
    app._worker_thread = None
    app._progress = None
    app._log_text = None
    starts = []
    completions = []
    app.show_start = lambda: starts.append(True)
    app._pipeline_failed = lambda error: pytest.fail(str(error))

    app.cancel_processing()
    app._start_worker(
        lambda: check_processing_cancelled(),
        lambda result: completions.append(result),
    )
    assert app._worker_thread is not None
    app._worker_thread.join(timeout=2)
    app._poll_worker()

    assert starts == [True]
    assert completions == []
    assert app._worker_messages is None


def test_text_similarity() -> None:
    assert text_similarity("Where are you going?", "where are you going") > 95


def test_order_independent_alignment_handles_reordered_lines_and_takes() -> None:
    lines = [
        {"line_id": "a", "line": "Hello there."},
        {"line_id": "b", "line": "Where are you going?"},
        {"line_id": "c", "line": "Goodbye friend."},
    ]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 1.0,
            "transcript": "Goodbye friend.",
            "asr_probability": 0.98,
        },
        {
            "start_seconds": 1.5,
            "end_seconds": 2.5,
            "transcript": "Hello there.",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 3.0,
            "end_seconds": 4.0,
            "transcript": "Goodbye friend!",
            "asr_probability": 0.97,
        },
        {
            "start_seconds": 4.5,
            "end_seconds": 5.8,
            "transcript": "Where are you going?",
            "asr_probability": 0.96,
        },
    ]
    settings = _alignment_settings(
        span_search={
            "max_segments": 2,
            "max_gap_seconds": 2.0,
            "max_duration_seconds": 20.0,
            "minimum_score": 45.0,
            "candidate_top_k": 3,
        },
        ranking={
            "noise_penalty": 2.2,
            "duration_hint_weight": 1.0,
            "order_hint_weight": 0.0,
        },
    )

    actions = order_independent_align(segments, lines, settings)

    assert [action["line_index"] for action in actions] == [2, 0, 2, 1]
    assert all(
        action["top_matches"][0]["line_index"] == action["line_index"]
        for action in actions
    )


def test_ordered_scoring_separates_consecutive_multiclause_takes() -> None:
    line_text = (
        "Ugh, you reek. How long have you been trudging along through the mud "
        "without a proper bath? This is exactly why I miss Leyawiin."
    )
    transcripts = [
        "Huh?",
        "You reek.",
        "Ugh!",
        "How long have you been trudging along through the mud without a "
        "proper bath?",
        "This.",
        "is exactly why I miss Leowin.",
        "Whoa",
        "You're weak!",
        "How long have you been trudging along through the mud without a "
        "proper bath?",
        "This",
        "is exactly why I miss Leowin.",
        "Ugh, you reek.",
        "mmm",
        "How long have you been trudging along through the mud without a "
        "proper bath?",
        "This is exactly why I miss Leowin.",
    ]
    durations = [
        0.6,
        1.7,
        1.2,
        5.2,
        1.1,
        3.1,
        1.0,
        1.5,
        5.4,
        1.0,
        3.0,
        2.6,
        1.6,
        5.3,
        3.0,
    ]
    segments = []
    cursor = 0.0
    for transcript, duration in zip(transcripts, durations):
        segments.append(
            {
                "start_seconds": cursor,
                "end_seconds": cursor + duration,
                "transcript": transcript,
                "asr_probability": 0.95,
            }
        )
        cursor += duration + 0.1

    actions = order_independent_align(
        segments,
        [{"line_id": "line", "line": line_text}],
        _alignment_settings(
            span_search={
                "max_segments": 8,
                "max_gap_seconds": 2.5,
                "max_duration_seconds": 35.0,
                "minimum_score": 45.0,
            },
            ranking={
                "noise_penalty": 2.2,
                "duration_hint_weight": 1.0,
                "order_hint_weight": 0.0,
            },
        ),
    )

    assert [
        (action["start_index"], action["count"])
        for action in actions
    ] == [(0, 6), (6, 5), (11, 4)]


def test_text_similarity_penalizes_reordered_take_text() -> None:
    expected = (
        "Ugh, you reek. How long have you been trudging along through the "
        "mud without a proper bath? This is exactly why I miss Leyawiin."
    )
    reordered = (
        "How long have you been trudging along through the mud without a "
        "proper bath? This is exactly why I miss Leyawiin. Ugh, you reek."
    )

    assert text_similarity(expected, expected) > 99
    assert text_similarity(expected, reordered) < 90


def test_text_similarity_handles_repeated_compound_words() -> None:
    assert text_similarity("Run away! Run away!", "Runaway Runaway") >= 72


@pytest.mark.parametrize(
    ("expected", "observed"),
    [
        ("C'mon... get up!", "Come on, get up!"),
        ("Could've asked.", "Could have asked."),
        ("I've had enough.", "I have had enough."),
        ("You're finished!", "You are finished!"),
        ("We won't survive.", "We will not survive."),
        ("You gotta be joking.", "You got to be joking."),
    ],
)
def test_spoken_form_normalization_accepts_equivalent_phrasing(
    expected: str,
    observed: str,
) -> None:
    fidelity = transcript_fidelity(expected, observed)

    assert text_similarity(expected, observed) == pytest.approx(100.0)
    assert fidelity["ordered_similarity"] == pytest.approx(100.0)
    assert fidelity["token_coverage"] == 1.0
    assert fidelity["token_precision"] == 1.0
    assert _candidate_reliability(
        line={"line": expected},
        match_score=100.0,
        margin=20.0,
        settings={},
        observed=observed,
        duration_plausibility=80.0,
    ) == (True, "")


def test_square_bracket_metadata_is_ignored_for_spoken_matching() -> None:
    expected = "[if Ternimir Pass is blocked] The road to Leyawiin is blocked."
    observed = "The road to Leyawiin is blocked."

    assert verbal_script_text(expected) == observed
    assert text_similarity(expected, observed) == pytest.approx(100.0)
    assert transcript_fidelity(expected, observed) == transcript_fidelity(
        observed,
        observed,
    )
    assert sentence_fidelity(expected, observed)["missing_clause_count"] == 0
    assert _candidate_reliability(
        line={"line": expected},
        match_score=100.0,
        margin=20.0,
        settings={},
        observed=observed,
        duration_plausibility=80.0,
    ) == (True, "")


def test_segment_prompt_matching_ignores_square_bracket_metadata() -> None:
    target = {
        "line_id": "target",
        "line": "[only after the quest is complete] The road is open.",
    }
    other = {"line_id": "other", "line": "Nothing has changed."}

    assert _best_ordered_script_score([target, other], "The road is open.") == 100.0
    assert _prompt_candidates(
        [target, other],
        "The road is open.",
        maximum_words=4,
        top_k=1,
    ) == [target]


def test_missing_single_sentence_boundary_prevents_auto_acceptance() -> None:
    expected = (
        "I live with that woman, and there is no way anyone who's heard "
        "her speak would want to bed her, much less live with her."
    )
    missing_start = (
        "And there is no way anyone who's heard her speak would want to "
        "bed her, much less live with her."
    )
    missing_end = (
        "I live with that woman, and there's no way anyone who's heard "
        "her speak would want to bed her."
    )

    start_fidelity = transcript_fidelity(expected, missing_start)
    end_fidelity = transcript_fidelity(expected, missing_end)
    assert start_fidelity["token_coverage"] > 0.75
    assert start_fidelity["prefix_token_coverage"] < 0.60
    assert end_fidelity["token_coverage"] > 0.75
    assert end_fidelity["suffix_token_coverage"] < 0.60
    assert _candidate_reliability(
        line={"line": expected},
        match_score=text_similarity(expected, missing_start),
        margin=40.0,
        settings={},
        observed=missing_start,
        duration_plausibility=80.0,
    ) == (False, "MISSING_LINE_START")
    assert _candidate_reliability(
        line={"line": expected},
        match_score=text_similarity(expected, missing_end),
        margin=40.0,
        settings={},
        observed=missing_end,
        duration_plausibility=80.0,
    ) == (False, "MISSING_LINE_END")


def test_single_missing_or_extra_opening_word_prevents_auto_acceptance() -> None:
    missing_expected = "Good, I've had enough of you for one day."
    missing_observed = "I've had enough of you for one day."
    prepended_expected = (
        "About time. Do you know how long I've been waiting for this?"
    )
    prepended_observed = (
        "Innkeeper. About time. Do you know how long I've been waiting "
        "for this?"
    )

    missing_fidelity = transcript_fidelity(
        missing_expected,
        missing_observed,
    )
    prepended_fidelity = transcript_fidelity(
        prepended_expected,
        prepended_observed,
    )
    assert missing_fidelity["prefix_missing_token_count"] == 1
    assert missing_fidelity["leading_missing_token_count"] == 1
    assert prepended_fidelity["leading_extra_token_count"] == 1
    assert _candidate_reliability(
        line={"line": missing_expected},
        match_score=text_similarity(missing_expected, missing_observed),
        margin=20.0,
        settings={},
        observed=missing_observed,
        duration_plausibility=80.0,
    ) == (False, "MISSING_LINE_START")
    assert _candidate_reliability(
        line={"line": prepended_expected},
        match_score=text_similarity(prepended_expected, prepended_observed),
        margin=20.0,
        settings={},
        observed=prepended_observed,
        duration_plausibility=80.0,
    ) == (False, "EXTRA_LINE_START")


def test_repeated_short_clauses_keep_their_earliest_positions() -> None:
    line = (
        "No, no! A staff! That's it! Should have turned you into a staff! "
        "Then you'd actually be useful."
    )

    fidelity = sentence_fidelity(line, line)

    assert fidelity["missing_clause_count"] == 0
    assert fidelity["clauses_in_order"] is True
    assert fidelity["clause_positions"] == sorted(
        fidelity["clause_positions"]
    )


def test_exact_asr_can_reroute_join_to_better_script_line() -> None:
    actions = [
        {
            "start_index": 10,
            "count": 3,
            "line_index": 0,
            "primary_line_index": 0,
            "match_score": 94.0,
            "confidence_margin": 5.0,
            "segment_match_rank": 1,
            "is_primary_match": True,
            "fragment_join": True,
            "fragment_source_count": 3,
            "fragment_join_provisional": True,
        }
    ]

    rerouted = _reroute_actions_to_exact_asr_primary_matches(
        actions,
        materialized_by_span={
            (10, 3): {"segment_id": "session__m00011_00013"}
        },
        exact_scores_by_segment={
            "session__m00011_00013": [94.0, 100.0]
        },
        minimum_score=45.0,
    )

    corrected = next(
        action for action in rerouted if action["line_index"] == 1
    )
    assert corrected["is_primary_match"] is True
    assert corrected["primary_line_index"] == 1
    assert corrected["fragment_join"] is True
    assert corrected["match_score"] == 100.0


def test_exact_line_scores_exclude_nonverbal_and_vocalization_lines() -> None:
    lines = [
        {"line": "(cough)"},
        {"line": "oof"},
        {"line": "Oof, that hurt."},
    ]
    evaluator = TranscriptEvaluator(lines, {})

    scores = _exact_line_scores(lines, evaluator, "oof")

    assert scores[:2] == [0.0, 0.0]
    assert scores[2] > 0.0


def test_flat_alignment_settings_are_rejected() -> None:
    configured = default_alignment_config()
    configured["span_search"]["max_segments"] = 9
    configured["max_merge_segments"] = 7

    with pytest.raises(
        ValueError,
        match="Flat alignment settings are no longer supported",
    ):
        AlignmentSettings.from_value(configured)


def test_grouped_alignment_settings_apply_overrides() -> None:
    configured = default_alignment_config()
    configured["span_search"]["max_segments"] = 9

    settings = AlignmentSettings.from_value(configured)

    assert settings["max_merge_segments"] == 9
    assert settings["fragment_join_max_segments"] == 8
    assert settings["duplicate_line_policy"] == "weak_order"
    assert settings["auto_reject_clipping"] is True


def test_project_migration_rejects_flat_alignment_settings() -> None:
    with pytest.raises(
        ValueError,
        match="Flat alignment settings are no longer supported",
    ):
        migrate_project_config(
            {
                "schema_version": 1,
                "settings_version": 3,
                "alignment": {"max_merge_segments": 6},
            }
        )


def test_grouped_alignment_settings_expose_every_effective_default() -> None:
    configured = default_alignment_config()

    settings = AlignmentSettings.from_value(configured)

    assert settings["max_merge_segments"] == 8
    assert settings["fragment_join_max_segments"] == 8
    assert (
        configured["recovery"]["trim_candidates_per_segment"]
        == settings["intra_segment_trim_max_actions_per_segment"]
    )
    assert (
        configured["recovery"]["complete"]["maximum_length_ratio"]
        == settings["fragment_join_complete_max_length_ratio"]
    )


def test_grouped_v1_defaults_migrate_back_to_historical_limits() -> None:
    project = {
        "schema_version": 1,
        "segmentation": {
            "silence_noise_db": -45.0,
            "silence_detection_min_seconds": 0.35,
            "split_gap_seconds": 0.35,
        },
        "alignment": {
            "span_search": {"max_segments": 10},
            "recovery": {"max_segments": 10},
        },
    }

    migrated = migrate_project_config(project)
    settings = AlignmentSettings.from_value(migrated["alignment"])

    assert migrated["settings_version"] == 3
    assert settings["max_merge_segments"] == 8
    assert settings["fragment_join_max_segments"] == 8
    assert len(settings) > 0
    assert migrated["segmentation"] == {
        "silence_noise_db": -40.0,
        "silence_detection_min_seconds": 0.20,
        "split_gap_seconds": 0.20,
        "word_split_snap_enabled": True,
        "word_split_snap_search_seconds": 0.20,
        "word_split_snap_window_seconds": 0.02,
        "word_split_snap_max_rms_dbfs": -42.0,
        "voice_boundary_detection_enabled": True,
        "voice_boundary_vad_threshold": 0.50,
        "voice_boundary_breath_vad_threshold": 0.70,
    }


def test_v2_alignment_vad_thresholds_migrate_to_segmentation() -> None:
    project = {
        "schema_version": 1,
        "settings_version": 2,
        "segmentation": {},
        "alignment": {
            "recovery": {
                "audio_boundaries": {
                    "snap_word_gaps": False,
                    "snap_search_seconds": 0.15,
                    "vad_threshold": 0.55,
                    "breath_vad_threshold": 0.75,
                }
            }
        },
    }

    migrated = migrate_project_config(project)

    assert migrated["settings_version"] == 3
    assert (
        migrated["segmentation"]["voice_boundary_vad_threshold"] == 0.55
    )
    assert migrated["segmentation"]["word_split_snap_enabled"] is False
    assert (
        migrated["segmentation"]["word_split_snap_search_seconds"] == 0.15
    )
    assert (
        migrated["segmentation"]["voice_boundary_breath_vad_threshold"]
        == 0.75
    )
    assert (
        "vad_threshold"
        not in migrated["alignment"]["recovery"]["audio_boundaries"]
    )


def test_ambiguous_generic_recording_includes_both_sheets_and_needs_review(
    tmp_path: Path,
) -> None:
    source_data = {
        "sheets": [
            {
                "name": "Any NPC using voice type ARG1RM",
                "line_count": 97,
                "voice_header": "You are voicing ARG1RMElfSly",
            },
            {
                "name": "Лист1",
                "line_count": 150,
                "voice_header": "You are voicing ARG1RMElfSly (Combat)",
            },
            {
                "name": "Member of faction ARGBanditFact",
                "line_count": 17,
                "voice_header": "You are voicing ARGBanditFaction",
            },
            {
                "name": "Лист2",
                "line_count": 64,
                "voice_header": "You are voicing ARGBanditFaction (Combat)",
            },
        ]
    }
    names = [
        "ARG_MaleElfSly_Generic_POST-PROC.wav",
        "ARG_MaleElfSly_Combat_POST-PROC.wav",
        "ARG_MaleElfSly_BanditGeneric_POST-PROC.wav",
        "ARG_MaleElfSly_BanditCombat_POST-PROC.wav",
    ]
    audio_files = [tmp_path / name for name in names]

    sessions = infer_sessions(audio_files, source_data, tmp_path)
    mappings = {session["id"]: session for session in sessions}

    assert mappings["arg_maleelfsly_generic_post_proc"]["sheets"] == [
        "Any NPC using voice type ARG1RM",
        "Лист1",
    ]
    assert mappings["arg_maleelfsly_generic_post_proc"][
        "needs_mapping_review"
    ] is True
    assert mappings["arg_maleelfsly_combat_post_proc"]["sheets"] == ["Лист1"]
    assert mappings["arg_maleelfsly_banditgeneric_post_proc"]["sheets"] == [
        "Member of faction ARGBanditFact"
    ]
    assert mappings["arg_maleelfsly_banditcombat_post_proc"]["sheets"] == [
        "Лист2"
    ]
    assert not any(
        session["needs_mapping_review"]
        for session_id, session in mappings.items()
        if session_id != "arg_maleelfsly_generic_post_proc"
    )


def test_exact_short_match_still_requires_resolved_ambiguity() -> None:
    settings = _alignment_settings(
        reliability={
            "short": {
                "minimum_score": 88.0,
                "minimum_margin": 15.0,
            }
        }
    )
    line = {"line": "Goodbye."}

    assert _candidate_reliability(
        line=line,
        match_score=100.0,
        margin=5.0,
        settings=settings,
        observed="Goodbye",
    ) == (False, "SHORT_LINE_AMBIGUOUS")
    assert _candidate_reliability(
        line=line,
        match_score=100.0,
        margin=5.0,
        settings=settings,
        observed="Goodbye",
        ambiguity_resolved=True,
    ) == (True, "")


def test_clipping_blocks_automatic_reliability() -> None:
    assert _candidate_reliability(
        line={"line": "This transcript is otherwise complete."},
        match_score=100.0,
        margin=100.0,
        settings={},
        observed="This transcript is otherwise complete.",
        clipping_samples=1,
        technical_score=65.0,
    ) == (False, "TECHNICAL_CLIPPING")


def test_short_line_auto_reliability_requires_order_and_no_extra_words() -> None:
    settings = _alignment_settings(
        reliability={
            "short": {
                "minimum_score": 88.0,
                "minimum_margin": 15.0,
                "minimum_ordered_score": 70.0,
                "minimum_token_coverage": 1.0,
                "minimum_token_precision": 1.0,
            }
        }
    )
    line = {"line": "Not... remember..."}

    assert text_similarity(line["line"], "Remember Not") < 88.0
    assert _candidate_reliability(
        line=line,
        match_score=95.0,
        margin=40.0,
        settings=settings,
        observed="Remember Not",
    ) == (False, "SHORT_LINE_ORDER_MISMATCH")
    assert _candidate_reliability(
        line=line,
        match_score=95.0,
        margin=40.0,
        settings=settings,
        observed="Not remember not",
    ) == (False, "SHORT_LINE_EXTRA_WORDS")
    assert _candidate_reliability(
        line=line,
        match_score=100.0,
        margin=40.0,
        settings=settings,
        observed="Not remember",
    ) == (True, "")


def test_suspicious_duration_prevents_exact_match_auto_acceptance() -> None:
    line = {"line": "Pathetic!"}

    assert _candidate_reliability(
        line=line,
        match_score=100.0,
        margin=40.0,
        settings=_alignment_settings(
            reliability={"minimum_duration_plausibility": 25.0}
        ),
        observed="Pathetic",
        duration_plausibility=20.2,
    ) == (False, "POSSIBLE_REPEATED_TAKES")
    assert _candidate_reliability(
        line=line,
        match_score=100.0,
        margin=40.0,
        settings=_alignment_settings(
            reliability={"minimum_duration_plausibility": 25.0}
        ),
        observed="Pathetic",
        duration_plausibility=38.4,
    ) == (True, "")


@pytest.mark.parametrize(
    ("expected", "repeated"),
    [
        (
            "You'll break before I will!",
            "Break Before I Will You'll Break Before I Will Will "
            "You'll Break Before I Will",
        ),
        (
            "Right! Bring it on!",
            "Bring It On Right Bring It On Right Bring It On",
        ),
    ],
)
def test_transcribed_repeated_takes_prevent_auto_acceptance(
    expected: str,
    repeated: str,
) -> None:
    reliable, reason = _candidate_reliability(
        line={"line": expected},
        match_score=text_similarity(expected, repeated),
        margin=40.0,
        settings={},
        observed=repeated,
        duration_plausibility=50.0,
    )

    assert reliable is False
    assert reason == "LOW_MATCH_SCORE"


def test_transcript_fidelity_tolerates_minor_asr_spelling_errors() -> None:
    fidelity = transcript_fidelity(
        "Good afternoon.",
        "Good afternon",
    )

    assert fidelity["ordered_similarity"] >= 90.0
    assert fidelity["token_coverage"] == 1.0
    assert fidelity["token_precision"] == 1.0
    assert fidelity["extra_word_count"] == 0


@pytest.mark.parametrize(
    ("expected", "partial"),
    [
        (
            "You've been avoiding the sun, haven't you? "
            "Can't say I blame you with this heat.",
            "Can't say I blame you with this heat",
        ),
        (
            "Got too much to do. Feels like I barely started...",
            "Feels like I barely started",
        ),
        (
            "Ever hear the story about the frozen Hist? "
            "Wonder if any of it's true...",
            "Ever hear the story about the frozen Hist",
        ),
        (
            "Been some strange storms lately. "
            "You wouldn't know anything about that, would you?",
            "You wouldn't know anything about that would you",
        ),
        ("Maybe over here... No, nothing.", "Maybe over here"),
    ],
)
def test_missing_sentence_prevents_auto_acceptance(
    expected: str,
    partial: str,
) -> None:
    sentence = sentence_fidelity(expected, partial)

    assert sentence["clause_count"] == 2
    assert sentence["missing_clause_count"] >= 1
    assert _candidate_reliability(
        line={"line": expected},
        match_score=max(90.0, text_similarity(expected, partial)),
        margin=40.0,
        settings=_alignment_settings(
            reliability={
                "normal": {
                    "minimum_score": 70.0,
                    "minimum_margin": 5.0,
                },
                "clauses": {"minimum_score": 55.0},
            }
        ),
        observed=partial,
    ) == (False, "MISSING_SENTENCE")


def test_reordered_sentences_prevent_auto_acceptance() -> None:
    expected = (
        "Hmm... I don't suppose we'll be able to count myself among that "
        "latter group. Tell me, how goes the war?"
    )
    observed = (
        "Tell me how goes the war Hmm I don't suppose we'll be able to "
        "count myself among that latter group"
    )

    sentence = sentence_fidelity(expected, observed)

    assert sentence["missing_clause_count"] == 0
    assert sentence["clauses_in_order"] is False
    assert _candidate_reliability(
        line={"line": expected},
        match_score=94.0,
        margin=40.0,
        settings={},
        observed=observed,
        duration_plausibility=70.0,
    ) == (False, "SENTENCE_ORDER_MISMATCH")


def test_adjacent_sentence_fragments_create_complete_take_candidates() -> None:
    line_text = (
        "You've been avoiding the sun, haven't you? "
        "Can't say I blame you with this heat."
    )
    lines = [{"line_id": "line", "line": line_text}]
    transcripts = [
        "Been avoiding the sun haven't you",
        "Can't say I blame you with this heat",
        "You've been avoiding the sun haven't you",
        "Can't say I blame you with this heat",
    ]
    segments = [
        {
            "start_seconds": index * 2.0,
            "end_seconds": index * 2.0 + 1.5,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": index,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(line_text, transcript),
            "transcript": transcript,
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
        for index, transcript in enumerate(transcripts)
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 2,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    assert [(action["start_index"], action["count"]) for action in joined] == [
        (0, 2),
        (2, 2),
    ]
    assert all(action["fragment_join"] for action in joined)
    assert all(action["fragment_source_count"] == 2 for action in joined)
    assert all(
        sentence_fidelity(line_text, action["transcript"])[
            "missing_clause_count"
        ]
        == 0
        for action in joined
    )
    limited = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 2,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            },
            recovery={"max_candidates_per_line": 1},
        ),
    )
    assert len(limited) == 1


def test_boundary_completion_can_join_repeated_short_clauses() -> None:
    line_text = "No! Don't be dead... Don't be dead..."
    lines = [{"line_id": "line", "line": line_text}]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 0.8,
            "transcript": "No!",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 0.9,
            "end_seconds": 2.9,
            "transcript": "Don't be dead! Don't be dead!",
            "asr_probability": 0.95,
        },
    ]
    actions = [
        {
            "start_index": index,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(
                line_text,
                segment["transcript"],
            ),
            "transcript": segment["transcript"],
            "duration_plausibility": 70.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
        for index, segment in enumerate(segments)
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    assert len(joined) == 1
    assert joined[0]["start_index"] == 0
    assert joined[0]["count"] == 2
    assert joined[0]["match_score"] == 100.0


def test_fragment_join_keeps_bounded_high_score_fallbacks() -> None:
    lines = [
        {
            "line_id": "target",
            "line": "You win... I... I surrender!",
        },
        {"line_id": "win", "line": "You win."},
        {"line_id": "aye", "line": "Aye."},
    ]
    transcripts = [
        "I surrender.",
        "You win.",
        "Aye.",
        "I surrender.",
    ]
    segments = [
        {
            "start_seconds": index * 1.1,
            "end_seconds": index * 1.1 + 1.0,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": index,
            "count": 1,
            "line_index": line_index,
            "match_score": text_similarity(
                lines[line_index]["line"],
                transcript,
            ),
            "transcript": transcript,
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
        for index, (line_index, transcript) in enumerate(
            zip([2, 1, 2, 0], transcripts)
        )
    ]
    words = [
        {"word": " Win", "start": 1.2, "end": 1.8},
        {"word": " I", "start": 2.3, "end": 2.5},
        {"word": " I", "start": 2.5, "end": 2.7},
        {"word": " surrender", "start": 3.4, "end": 3.9},
    ]
    transcription = {
        "segments": [{"start": 1.2, "end": 3.9, "words": words}]
    }

    without_fallback = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 4,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            },
            recovery={"fallback_candidates_per_line": 0},
        ),
        transcription=transcription,
    )
    with_fallback = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 4,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            },
            recovery={
                "fallback_candidates_per_line": 2,
                "fallback_minimum_score": 89.0,
            },
        ),
        transcription=transcription,
    )

    assert not any(
        action["start_index"] == 1 and action["count"] == 3
        for action in without_fallback
    )
    restored = next(
        action
        for action in with_fallback
        if action["start_index"] == 1 and action["count"] == 3
    )
    assert restored["fragment_join_fallback"] is True
    assert restored["match_score"] >= 89.0
    assert (
        sum(
            bool(action.get("fragment_join_fallback"))
            for action in with_fallback
        )
        == 2
    )


def test_oversized_base_segment_can_recover_pause_bounded_trim() -> None:
    line_text = (
        "Oh! Is that all? Not that I ever say anything true while drunk."
    )
    lines = [{"line_id": "target", "line": line_text}]
    word_text = [
        "Oh",
        "is",
        "that",
        "all",
        "Not",
        "that",
        "I",
        "ever",
        "say",
        "anything",
        "true",
        "while",
        "drunk",
    ]
    words = [
        {
            "word": f" {word}",
            "start": index * 0.2,
            "end": index * 0.2 + 0.18,
        }
        for index, word in enumerate(word_text)
    ]
    words.extend(
        [
            {"word": " Yes", "start": 3.2, "end": 3.5},
            {"word": " appended", "start": 3.5, "end": 3.9},
            {"word": " sentence", "start": 3.9, "end": 4.3},
        ]
    )
    full_transcript = " ".join(
        str(word["word"]).strip() for word in words
    )
    base_start_sample = 1000
    segment = {
        "segment_id": "session__s00001",
        "start_sample": base_start_sample,
        "end_sample": base_start_sample + 5 * 48000,
        "start_seconds": 10.0,
        "end_seconds": 15.0,
        "transcript": full_transcript,
        "asr_probability": 0.95,
        "segment_asr": {
            "primary": {
                "transcript": full_transcript,
                "words": words,
            }
        },
    }
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": text_similarity(line_text, full_transcript),
        "transcript": full_transcript,
        "duration_plausibility": 60.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=lines,
        base_segments=[segment],
        sample_rate=48000,
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    assert len(trimmed) == 1
    assert trimmed[0]["intra_segment_trim"] is True
    assert trimmed[0]["trim_start_sample"] == base_start_sample
    assert trimmed[0]["trim_end_sample"] < segment["end_sample"]
    assert trimmed[0]["transcript"].endswith("drunk")
    assert "appended" not in trimmed[0]["transcript"]
    assert trimmed[0]["match_score"] == 100.0
    expanded = _expand_alignment_actions(
        trimmed,
        lines=lines,
        settings={},
    )
    assert expanded[0]["intra_segment_trim"] is True
    assert (
        expanded[0]["trim_end_sample"]
        == trimmed[0]["trim_end_sample"]
    )


def test_trim_recovery_scans_bases_inside_a_merged_primary_action() -> None:
    line_text = (
        "Nothing too substantial. A friendly game of cards or dice."
    )
    target_words = [
        "Nothing",
        "too",
        "substantial",
        "A",
        "friendly",
        "game",
        "of",
        "cards",
        "or",
        "dice",
    ]
    words = [
        {
            "word": f" {word}",
            "start": index * 0.2,
            "end": index * 0.2 + 0.18,
        }
        for index, word in enumerate(target_words)
    ]
    words.extend(
        [
            {"word": " Nothing", "start": 3.0, "end": 3.4},
            {"word": " too", "start": 3.4, "end": 3.7},
            {"word": " substantial", "start": 3.7, "end": 4.2},
        ]
    )
    full_transcript = " ".join(
        str(word["word"]).strip() for word in words
    )
    segments = [
        {
            "segment_id": "session__s00001",
            "start_sample": 0,
            "end_sample": 48000,
            "start_seconds": 0.0,
            "end_seconds": 1.0,
            "transcript": "Background",
            "asr_probability": 0.2,
            "segment_asr": {"primary": {"words": []}},
        },
        {
            "segment_id": "session__s00002",
            "start_sample": 48000,
            "end_sample": 288000,
            "start_seconds": 1.0,
            "end_seconds": 6.0,
            "transcript": full_transcript,
            "asr_probability": 0.95,
            "segment_asr": {
                "primary": {
                    "transcript": full_transcript,
                    "words": words,
                }
            },
        },
    ]
    action = {
        "start_index": 0,
        "count": 2,
        "line_index": 0,
        "match_score": text_similarity(line_text, full_transcript),
        "transcript": full_transcript,
        "duration_plausibility": 60.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=[{"line_id": "target", "line": line_text}],
        base_segments=segments,
        sample_rate=48000,
        settings={},
        evaluator=TranscriptEvaluator(
            [{"line_id": "target", "line": line_text}],
            {},
        ),
    )

    recovered = next(
        item for item in trimmed if item["start_index"] == 1
    )
    assert recovered["count"] == 1
    assert recovered["trim_end_sample"] < segments[1]["end_sample"]
    assert recovered["match_score"] == 100.0
    assert "Nothing too substantial Nothing" not in recovered["transcript"]


@pytest.mark.parametrize(
    ("prefix", "suffix"),
    [
        ("I suppose I should start", ""),
        ("", "The next script line starts here"),
    ],
)
def test_trim_recovery_finds_complete_take_beside_partial_dialogue(
    prefix: str,
    suffix: str,
) -> None:
    sample_rate = 48000
    line_text = "I suppose I should start making space for the display."
    line_tokens = line_text.rstrip(".").split()
    tokens = [*prefix.split(), *line_tokens, *suffix.split()]
    line_start = len(prefix.split())
    line_end = line_start + len(line_tokens) - 1
    words = []
    cursor = 0.0
    for index, token in enumerate(tokens):
        if index == line_start and line_start:
            cursor += 0.25
        if index == line_end + 1:
            cursor += 0.25
        words.append(
            {"word": f" {token}", "start": cursor, "end": cursor + 0.20}
        )
        cursor += 0.20
    transcript = " ".join(tokens)
    segment = {
        "segment_id": "session__s00001",
        "start_sample": 0,
        "end_sample": round((cursor + 0.20) * sample_rate),
        "start_seconds": 0.0,
        "end_seconds": cursor + 0.20,
        "transcript": transcript,
        "segment_asr": {
            "primary": {"transcript": transcript, "words": words}
        },
    }
    line = {"line_id": "target", "line": line_text}
    evaluator = TranscriptEvaluator([line], {})
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": evaluator.match(0, transcript),
        "transcript": transcript,
        "duration_plausibility": 30.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=[line],
        base_segments=[segment],
        sample_rate=sample_rate,
        settings={},
        evaluator=evaluator,
    )

    recovered = next(item for item in trimmed if item["match_score"] == 100.0)
    assert recovered["repeated_take_trim"] is True
    assert (recovered["trim_start_sample"] > 0) is bool(prefix)
    assert (
        recovered["trim_end_sample"] < segment["end_sample"]
    ) is bool(suffix)


def test_trim_recovery_deduplicates_fuzzy_windows_between_repeated_takes() -> None:
    sample_rate = 48000
    line_text = (
        "I can send Seeks-Treasure and a few volunteers there at once. "
        "We'll work out other details when we have time."
    )
    spoken_tokens = (
        "I can send six treasure and a few volunteers there at once "
        "we'll work out other details when we have time"
    ).split()
    words = []
    cursor = 0.0
    for take in range(2):
        if take:
            cursor += 0.50
        for token in spoken_tokens:
            words.append(
                {
                    "word": f" {token}",
                    "start": cursor,
                    "end": cursor + 0.20,
                }
            )
            cursor += 0.20
    transcript = " ".join(spoken_tokens * 2)
    segment = {
        "segment_id": "session__s00001",
        "start_sample": 0,
        "end_sample": round((cursor + 0.25) * sample_rate),
        "start_seconds": 0.0,
        "end_seconds": cursor + 0.25,
        "transcript": transcript,
        "segment_asr": {
            "primary": {"transcript": transcript, "words": words}
        },
    }
    line = {"line_id": "target", "line": line_text}
    evaluator = TranscriptEvaluator([line], {})
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": evaluator.match(0, transcript),
        "transcript": transcript,
        "duration_plausibility": 20.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=[line],
        base_segments=[segment],
        sample_rate=sample_rate,
        settings={},
        evaluator=evaluator,
    )

    assert len(trimmed) == 2
    assert all(item["match_score"] > 95.0 for item in trimmed)
    assert trimmed[0]["trim_end_sample"] == trimmed[1]["trim_start_sample"]
    assert all(item["repeated_take_trim"] for item in trimmed)


def test_trim_recovery_splits_adjacent_repeated_line_without_word_gap() -> None:
    line_text = "Wonderful to see you again."
    repeated_words = [
        {"word": " wonderful", "start": 0.16, "end": 0.80},
        {"word": " to", "start": 0.80, "end": 1.02},
        {"word": " see", "start": 1.02, "end": 1.30},
        {"word": " you", "start": 1.30, "end": 1.54},
        {"word": " again", "start": 1.54, "end": 1.90},
        # Large-v3 can stretch the first word of the second take across
        # the pause, leaving no timestamp gap at the repeated boundary.
        {"word": " wonderful", "start": 1.90, "end": 3.46},
        {"word": " to", "start": 3.46, "end": 3.78},
        {"word": " see", "start": 3.78, "end": 4.10},
        {"word": " you", "start": 4.10, "end": 4.40},
        {"word": " again", "start": 4.40, "end": 4.78},
    ]
    segment = {
        "segment_id": "session__s00001",
        "start_sample": 0,
        "end_sample": round(5.65 * 48000),
        "start_seconds": 0.0,
        "end_seconds": 5.65,
        "transcript": (
            "wonderful to see you again "
            "wonderful to see you again"
        ),
        "asr_probability": 0.9,
        "segment_asr": {
            "primary": {
                "transcript": (
                    "wonderful to see you again "
                    "wonderful to see you again"
                ),
                "words": repeated_words,
            }
        },
    }
    lines = [{"line_id": "target", "line": line_text}]
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 86.0,
        "transcript": segment["transcript"],
        "duration_plausibility": 40.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=lines,
        base_segments=[segment],
        sample_rate=48000,
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    assert len(trimmed) == 2
    first, second = trimmed
    expected_boundary = round(((1.90 + 3.46) / 2.0) * 48000)
    assert first["trim_start_sample"] == 0
    assert first["trim_end_sample"] == expected_boundary
    assert second["trim_start_sample"] == expected_boundary
    assert second["trim_end_sample"] == segment["end_sample"]
    assert all(item["match_score"] == 100.0 for item in trimmed)


def test_trim_recovery_discovers_repeated_single_word_for_other_primary_line() -> None:
    sample_rate = 48000
    lines = [
        {"line_id": "R21", "line": "Goodbye."},
        {"line_id": "other", "line": "Safe travels."},
    ]
    words = [
        {"word": " Goodbye", "start": 0.0, "end": 0.53},
        {"word": " Goodbye", "start": 0.81, "end": 1.55},
    ]
    segment = {
        "segment_id": "session__s00043",
        "start_sample": 0,
        "end_sample": round(2.2 * sample_rate),
        "start_seconds": 0.0,
        "end_seconds": 2.2,
        "transcript": "Goodbye Goodbye",
        "segment_asr": {
            "primary": {
                "transcript": "Goodbye Goodbye",
                "words": words,
            }
        },
    }
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 1,
        "match_score": 20.0,
        "transcript": segment["transcript"],
        "duration_plausibility": 20.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=lines,
        base_segments=[segment],
        sample_rate=sample_rate,
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    goodbye_takes = [item for item in trimmed if item["line_index"] == 0]
    assert len(goodbye_takes) == 2
    assert all(item["repeated_take_trim"] for item in goodbye_takes)


def test_trim_recovery_keeps_three_repetitions_before_other_primary_line() -> None:
    sample_rate = 48000
    lines = [
        {"line_id": "R22", "line": "See you later."},
        {"line_id": "R23", "line": "Safe travels."},
    ]
    words = [
        {"word": " See", "start": 0.0, "end": 0.358},
        {"word": " you", "start": 0.358, "end": 0.458},
        {"word": " later", "start": 0.458, "end": 0.758},
        {"word": " See", "start": 0.898, "end": 1.818},
        {"word": " you", "start": 1.818, "end": 1.998},
        {"word": " later", "start": 1.998, "end": 2.318},
        {"word": " See", "start": 2.458, "end": 3.218},
        {"word": " you", "start": 3.218, "end": 3.358},
        {"word": " later", "start": 3.358, "end": 3.638},
        {"word": " Safe", "start": 4.698, "end": 4.998},
        {"word": " travels", "start": 4.998, "end": 5.378},
    ]
    transcript = "See you later See you later See you later Safe travels"
    segment = {
        "segment_id": "session__s00048",
        "start_sample": 0,
        "end_sample": round(6.125 * sample_rate),
        "start_seconds": 0.0,
        "end_seconds": 6.125,
        "transcript": transcript,
        "segment_asr": {
            "primary": {"transcript": transcript, "words": words}
        },
    }
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 1,
        "match_score": 50.0,
        "transcript": transcript,
        "duration_plausibility": 20.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=lines,
        base_segments=[segment],
        sample_rate=sample_rate,
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    r22_takes = [item for item in trimmed if item["line_index"] == 0]
    assert len(r22_takes) == 3
    assert all(item["repeated_take_trim"] for item in r22_takes)


def test_trim_recovery_allows_compound_word_asr_in_repeated_take() -> None:
    sample_rate = 48000
    lines = [
        {"line_id": "PatronR6", "line": "Mud hopper stew, extra spicy."},
        {"line_id": "other", "line": "Yule pie."},
    ]
    words = [
        {"word": " Mudhopper", "start": 0.0, "end": 0.912},
        {"word": " stew", "start": 0.912, "end": 1.131},
        {"word": " Extra", "start": 1.432, "end": 2.272},
        {"word": " spicy", "start": 2.272, "end": 2.792},
        {"word": " Mudhopper", "start": 3.912, "end": 4.872},
        {"word": " stew", "start": 4.872, "end": 5.131},
        {"word": " Extra", "start": 5.592, "end": 6.551},
        {"word": " spicy", "start": 6.551, "end": 7.151},
    ]
    transcript = "Mudhopper stew Extra spicy Mudhopper stew Extra spicy"
    segment = {
        "segment_id": "session__s00024",
        "start_sample": 0,
        "end_sample": round(8.223 * sample_rate),
        "start_seconds": 0.0,
        "end_seconds": 8.223,
        "transcript": transcript,
        "segment_asr": {
            "primary": {"transcript": transcript, "words": words}
        },
    }
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 1,
        "match_score": 20.0,
        "transcript": transcript,
        "duration_plausibility": 20.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=lines,
        base_segments=[segment],
        sample_rate=sample_rate,
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    patron_takes = [item for item in trimmed if item["line_index"] == 0]
    assert len(patron_takes) == 2
    assert all(item["repeated_take_trim"] for item in patron_takes)
    assert all(item["match_score"] > 95.0 for item in patron_takes)


def test_repeated_take_trim_uses_ordered_vad_gaps(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    line_text = "See you later."
    words = [
        {"word": " See", "start": 0.0, "end": 0.35},
        {"word": " you", "start": 0.35, "end": 0.46},
        {"word": " later", "start": 0.46, "end": 0.76},
        {"word": " See", "start": 0.90, "end": 1.82},
        {"word": " you", "start": 1.82, "end": 2.0},
        {"word": " later", "start": 2.0, "end": 2.32},
        {"word": " See", "start": 2.46, "end": 3.22},
        {"word": " you", "start": 3.22, "end": 3.36},
        {"word": " later", "start": 3.36, "end": 3.64},
    ]
    segment = {
        "segment_id": "session__s00001",
        "file": "segment.wav",
        "start_sample": 0,
        "end_sample": round(4.2 * sample_rate),
        "start_seconds": 0.0,
        "end_seconds": 4.2,
        "transcript": "See you later See you later See you later",
        "segment_asr": {
            "primary": {
                "transcript": "See you later See you later See you later",
                "words": words,
            }
        },
    }
    monkeypatch.setattr(
        alignment_module,
        "_segment_voice_regions_for_trimming",
        lambda *_args, **_kwargs: [
            (round(0.25 * sample_rate), round(1.20 * sample_rate)),
            (round(1.42 * sample_rate), round(2.70 * sample_rate)),
            (round(2.92 * sample_rate), round(3.84 * sample_rate)),
        ],
    )
    monkeypatch.setattr(
        alignment_module,
        "quietest_pcm_boundary",
        lambda _path, **kwargs: int(kwargs["proposed_sample"]),
    )
    line = {"line_id": "target", "line": line_text}
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 60.0,
        "transcript": segment["transcript"],
        "duration_plausibility": 30.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=[line],
        base_segments=[segment],
        sample_rate=sample_rate,
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
        project_dir=tmp_path,
    )

    expected_boundaries = {
        round(((1.20 + 1.42) / 2.0) * sample_rate),
        round(((2.70 + 2.92) / 2.0) * sample_rate),
    }
    actual_boundaries = {
        int(item["trim_start_sample"])
        for item in trimmed
        if int(item["trim_start_sample"]) > 0
    } | {
        int(item["trim_end_sample"])
        for item in trimmed
        if int(item["trim_end_sample"]) < int(segment["end_sample"])
    }
    assert actual_boundaries == expected_boundaries
    assert all(item["repeated_take_trim"] for item in trimmed)


def test_acoustic_take_trim_recovers_asr_collapsed_repetitions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.alignment as alignment_module

    sample_rate = 48000
    line_text = "Seems like they're fighting over something."
    line = {"line_id": "target", "line": line_text}
    segment = {
        "segment_id": "session__s00001",
        "file": "segment.wav",
        "start_sample": 0,
        "end_sample": 12 * sample_rate,
        "start_seconds": 0.0,
        "end_seconds": 12.0,
        # ASR collapsed three audible takes into one transcript and one word
        # sequence, so textual repetition detection has no split points.
        "transcript": line_text,
        "segment_asr": {
            "primary": {
                "transcript": line_text,
                "words": [
                    {"word": " Seems", "start": 0.2, "end": 0.6},
                    {"word": " like", "start": 0.6, "end": 0.9},
                    {"word": " they're", "start": 0.9, "end": 1.2},
                    {"word": " fighting", "start": 1.2, "end": 1.7},
                    {"word": " over", "start": 1.7, "end": 2.0},
                    {"word": " something", "start": 2.0, "end": 2.6},
                ],
            }
        },
    }
    monkeypatch.setattr(
        alignment_module,
        "_segment_voice_regions_for_trimming",
        lambda *_args, **_kwargs: [
            (round(0.3 * sample_rate), round(3.0 * sample_rate)),
            (round(4.2 * sample_rate), round(7.0 * sample_rate)),
            (round(8.2 * sample_rate), round(11.0 * sample_rate)),
        ],
    )
    monkeypatch.setattr(
        alignment_module,
        "quietest_pcm_boundary",
        lambda _path, **kwargs: int(kwargs["proposed_sample"]),
    )
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": line_text,
        "duration_plausibility": 25.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=[line],
        base_segments=[segment],
        sample_rate=sample_rate,
        settings={},
        evaluator=TranscriptEvaluator([line], {}),
        project_dir=tmp_path,
    )

    acoustic = [item for item in trimmed if item.get("acoustic_take_trim")]
    assert len(acoustic) >= 3
    assert all(item["repeated_take_trim"] for item in acoustic)
    assert all(
        item["transcript_source"] == "acoustic_take_trim_preview"
        for item in acoustic
    )


def test_fragment_join_can_trim_shared_take_boundary_segment() -> None:
    line_text = "It's... not over... yet..."
    lines = [{"line_id": "target", "line": line_text}]

    def segment(
        index: int,
        transcript: str,
        words: list[dict[str, Any]],
        duration: float,
    ) -> dict[str, Any]:
        start_seconds = sum(
            (1.0, 1.0, 2.0, 1.0)[:index]
        )
        start_sample = round(start_seconds * 48000)
        return {
            "segment_id": f"session__s{index + 1:05d}",
            "start_sample": start_sample,
            "end_sample": start_sample + round(duration * 48000),
            "start_seconds": start_seconds,
            "end_seconds": start_seconds + duration,
            "transcript": transcript,
            "asr_probability": 0.95,
            "segment_asr": {
                "primary": {
                    "transcript": transcript,
                    "words": words,
                }
            },
        }

    segments = [
        segment(
            0,
            "It's",
            [{"word": " It's", "start": 0.0, "end": 0.7}],
            1.0,
        ),
        segment(
            1,
            "not over",
            [
                {"word": " not", "start": 0.0, "end": 0.4},
                {"word": " over", "start": 0.4, "end": 0.9},
            ],
            1.0,
        ),
        segment(
            2,
            "Yet. It's not over.",
            [
                {"word": " Yet", "start": 0.0, "end": 0.3},
                {"word": " It's", "start": 0.44, "end": 0.74},
                {"word": " not", "start": 0.74, "end": 1.04},
                {"word": " over", "start": 1.04, "end": 1.44},
            ],
            2.0,
        ),
        segment(
            3,
            "Yet",
            [{"word": " Yet", "start": 0.0, "end": 0.5}],
            1.0,
        ),
    ]
    actions = [
        {
            "start_index": 0,
            "count": 2,
            "line_index": 0,
            "match_score": text_similarity(line_text, "It's not over"),
            "transcript": "It's not over",
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [],
        },
        *[
            {
                "start_index": index,
                "count": 1,
                "line_index": 0,
                "match_score": text_similarity(
                    line_text,
                    segments[index]["transcript"],
                ),
                "transcript": segments[index]["transcript"],
                "duration_plausibility": 80.0,
                "order_hint": 0.0,
                "top_matches": [],
            }
            for index in (2, 3)
        ],
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 3,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    first_take = next(
        action
        for action in joined
        if action["start_index"] == 0
        and action["count"] == 3
        and action.get("trimmed_edge_join")
    )
    second_take = next(
        action
        for action in joined
        if action["start_index"] == 2
        and action["count"] == 2
        and action.get("trimmed_edge_join")
    )
    assert first_take["trim_end_sample"] < segments[2]["end_sample"]
    assert second_take["trim_start_sample"] > segments[2]["start_sample"]
    assert first_take["match_score"] == 100.0
    assert second_take["match_score"] == 100.0


def test_fragment_join_trims_repeated_prefix_from_complete_multisegment_take() -> None:
    line_text = (
        "Your best bet is Master Meretor. You can usually find him in the "
        "castle. Just be aware he's a bit... peculiar."
    )
    lines = [{"line_id": "target", "line": line_text}]
    first_tokens = (
        "Your best bet is Master Meretor "
        "Your best bet is Master Meretor "
        "You can usually find him in the castle"
    ).split()
    first_words = []
    cursor = 0.0
    for index, token in enumerate(first_tokens):
        if index == 6:
            cursor += 0.70
        elif index == 12:
            cursor += 0.50
        first_words.append(
            {"word": f" {token}", "start": cursor, "end": cursor + 0.20}
        )
        cursor += 0.20

    first_duration = cursor + 0.20
    second_text = "Just be aware he's a bit peculiar"
    second_tokens = second_text.split()
    second_words = [
        {
            "word": f" {token}",
            "start": index * 0.25,
            "end": index * 0.25 + 0.20,
        }
        for index, token in enumerate(second_tokens)
    ]
    second_duration = len(second_tokens) * 0.25 + 0.20
    sample_rate = 48000
    segments = [
        {
            "segment_id": "session__s00001",
            "start_sample": 0,
            "end_sample": round(first_duration * sample_rate),
            "start_seconds": 0.0,
            "end_seconds": first_duration,
            "transcript": " ".join(first_tokens),
            "asr_probability": 0.95,
            "segment_asr": {
                "primary": {
                    "transcript": " ".join(first_tokens),
                    "words": first_words,
                }
            },
        },
        {
            "segment_id": "session__s00002",
            "start_sample": round(first_duration * sample_rate),
            "end_sample": round(
                (first_duration + second_duration) * sample_rate
            ),
            "start_seconds": first_duration,
            "end_seconds": first_duration + second_duration,
            "transcript": second_text,
            "asr_probability": 0.95,
            "segment_asr": {
                "primary": {
                    "transcript": second_text,
                    "words": second_words,
                }
            },
        },
    ]
    evaluator = TranscriptEvaluator(lines, {})
    full_text = " ".join(segment["transcript"] for segment in segments)
    source_action = {
        "start_index": 0,
        "count": 2,
        "line_index": 0,
        "match_score": evaluator.match(0, full_text),
        "transcript": full_text,
        "duration_plausibility": 60.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    joined = _multisentence_fragment_join_actions(
        [source_action],
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 2,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 20.0,
            }
        ),
        evaluator=evaluator,
        span_catalog=SpanCatalog(segments, {}),
    )

    recovered = next(
        action
        for action in joined
        if action["start_index"] == 0
        and action["count"] == 2
        and action.get("trimmed_edge_join")
    )
    assert recovered["trim_start_sample"] > segments[0]["start_sample"]
    assert recovered["trim_end_sample"] == segments[1]["end_sample"]
    assert recovered["match_score"] == 100.0
    assert recovered["repeated_take_trim"] is True


def test_boundary_noise_cleanup_recovers_clean_single_base_span() -> None:
    lines = [{"line_id": "target", "line": "Better get going then."}]
    segments = [
        {
            "segment_id": "session__s00001",
            "start_sample": 0,
            "end_sample": 36000,
            "start_seconds": 0.0,
            "end_seconds": 0.75,
            "transcript": "Pfft.",
            "asr_probability": 0.5,
            "metrics": {"duration_seconds": 0.75, "rms_dbfs": -47.0},
            "segment_asr": {
                "primary": {
                    "transcript": "Pfft.",
                    "words": [
                        {"word": " Pfft.", "start": 0.0, "end": 0.7}
                    ],
                }
            },
        },
        {
            "segment_id": "session__s00002",
            "start_sample": 36000,
            "end_sample": 132000,
            "start_seconds": 0.75,
            "end_seconds": 2.75,
            "transcript": "Better get going then.",
            "asr_probability": 0.95,
            "metrics": {"duration_seconds": 2.0, "rms_dbfs": -18.0},
            "segment_asr": {
                "primary": {
                    "transcript": "Better get going then.",
                    "words": [],
                }
            },
        },
    ]
    evaluator = TranscriptEvaluator(lines, {})
    action = {
        "type": "assigned",
        "start_index": 0,
        "count": 2,
        "line_index": 0,
        "match_score": 96.0,
        "transcript": "Pfft. Better get going then.",
        "duration_plausibility": 80.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    cleaned = _boundary_noise_cleanup_actions(
        [action],
        lines=lines,
        base_segments=segments,
        settings={},
        evaluator=evaluator,
        span_catalog=SpanCatalog(segments, {}),
    )

    assert len(cleaned) == 1
    assert cleaned[0]["start_index"] == 1
    assert cleaned[0]["count"] == 1
    assert cleaned[0]["match_score"] == 100.0
    assert cleaned[0]["boundary_noise_cleanup"] is True


def test_complete_subspan_recovery_restores_exact_base_hidden_by_merge() -> None:
    line_text = "Is there anything else you wish to know?"
    lines = [
        {"line_id": "target", "line": line_text},
        {"line_id": "next", "line": "Yes."},
    ]
    segments = [
        {
            "segment_id": "session__s00001",
            "start_sample": 0,
            "end_sample": 96000,
            "start_seconds": 0.0,
            "end_seconds": 2.0,
            "transcript": line_text,
            "asr_probability": 0.99,
        },
        {
            "segment_id": "session__s00002",
            "start_sample": 100800,
            "end_sample": 120000,
            "start_seconds": 2.1,
            "end_seconds": 2.5,
            "transcript": "",
            "asr_probability": 0.0,
        },
        {
            "segment_id": "session__s00003",
            "start_sample": 124800,
            "end_sample": 153600,
            "start_seconds": 2.6,
            "end_seconds": 3.2,
            "transcript": "Yes.",
            "asr_probability": 0.99,
        },
    ]
    evaluator = TranscriptEvaluator(lines, {})
    source_action = {
        "type": "assigned",
        "start_index": 0,
        "count": 3,
        "line_index": 0,
        "match_score": text_similarity(line_text, f"{line_text} Yes."),
        "transcript": f"{line_text} Yes.",
        "duration_plausibility": 80.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    recovered = _complete_subspan_recovery_actions(
        [source_action],
        lines=lines,
        settings={},
        evaluator=evaluator,
        span_catalog=SpanCatalog(segments, {}),
    )

    clean = next(
        action
        for action in recovered
        if action["start_index"] == 0 and action["count"] == 1
    )
    assert clean["transcript"] == line_text
    assert clean["match_score"] == 100.0
    assert clean["complete_subspan_recovery"] is True


def test_edge_cue_recovery_attaches_adjacent_vocalization_segment() -> None:
    lines = [
        {
            "line_id": "target",
            "line": (
                "(cackle) You have no idea how satisfying it is "
                "to see you in a cell."
            ),
        }
    ]
    segments = [
        {
            "segment_id": "session__s00001",
            "start_sample": 0,
            "end_sample": 144000,
            "start_seconds": 0.0,
            "end_seconds": 3.0,
            "transcript": "Ha ha ha ha!",
            "asr_probability": 0.8,
            "metrics": {"duration_seconds": 3.0, "rms_dbfs": -20.0},
            "segment_asr": {
                "primary": {
                    "transcript": "Ha ha ha ha!",
                    "words": [],
                }
            },
        },
        {
            "segment_id": "session__s00002",
            "start_sample": 144000,
            "end_sample": 480000,
            "start_seconds": 3.0,
            "end_seconds": 10.0,
            "transcript": (
                "You have no idea how satisfying it is "
                "to see you in a cell."
            ),
            "asr_probability": 0.95,
            "metrics": {"duration_seconds": 7.0, "rms_dbfs": -18.0},
            "segment_asr": {
                "primary": {
                    "transcript": (
                        "You have no idea how satisfying it is "
                        "to see you in a cell."
                    ),
                    "words": [],
                }
            },
        },
    ]
    evaluator = TranscriptEvaluator(lines, {})
    action = {
        "type": "assigned",
        "start_index": 1,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "transcript": segments[1]["transcript"],
        "duration_plausibility": 80.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    extended = _edge_vocalization_extension_actions(
        [action],
        lines=lines,
        base_segments=segments,
        settings={},
        evaluator=evaluator,
        span_catalog=SpanCatalog(segments, {}),
    )

    assert len(extended) == 1
    assert extended[0]["start_index"] == 0
    assert extended[0]["count"] == 2
    assert extended[0]["edge_vocalization_extension"] is True
    assert (
        extended[0]["forced_review_reason"]
        == "EDGE_VOCALIZATION_UNVERIFIED"
    )


def test_intra_segment_trim_does_not_remove_scripted_edge_cue() -> None:
    line_text = "(laugh) See! What did I tell you?"
    lines = [{"line_id": "target", "line": line_text}]
    words = [
        {"word": " Ha!", "start": 0.0, "end": 0.4},
        {"word": " See!", "start": 1.0, "end": 1.4},
        {"word": " What", "start": 1.6, "end": 1.8},
        {"word": " did", "start": 1.8, "end": 2.0},
        {"word": " I", "start": 2.0, "end": 2.1},
        {"word": " tell", "start": 2.1, "end": 2.3},
        {"word": " you?", "start": 2.3, "end": 2.6},
    ]
    segment = {
        "segment_id": "session__s00001",
        "start_sample": 0,
        "end_sample": 144000,
        "start_seconds": 0.0,
        "end_seconds": 3.0,
        "transcript": "Ha! See! What did I tell you?",
        "asr_probability": 0.95,
        "segment_asr": {
            "primary": {
                "transcript": "Ha! See! What did I tell you?",
                "words": words,
            }
        },
    }
    action = {
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 98.0,
        "transcript": segment["transcript"],
        "duration_plausibility": 80.0,
        "order_hint": 0.0,
        "top_matches": [],
    }

    trimmed = _intra_segment_trim_actions(
        [action],
        lines=lines,
        base_segments=[segment],
        sample_rate=48000,
        settings={},
        evaluator=TranscriptEvaluator(lines, {}),
    )

    assert all(item["trim_start_sample"] == 0 for item in trimmed)


def test_secondary_near_complete_match_can_seed_fragment_join() -> None:
    lines = [
        {"line_id": "short", "line": "You should leave."},
        {"line_id": "target", "line": "You should leave. Now."},
        {"line_id": "now", "line": "Now."},
    ]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 1.2,
            "transcript": "You should leave.",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 1.2,
            "end_seconds": 2.0,
            "transcript": "Now.",
            "asr_probability": 0.95,
        },
    ]
    actions = [
        {
            "start_index": 0,
            "count": 1,
            "line_index": 0,
            "match_score": 100.0,
            "transcript": segments[0]["transcript"],
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [
                {"line_index": 0, "match_score": 100.0},
                {
                    "line_index": 1,
                    "match_score": text_similarity(
                        lines[1]["line"],
                        segments[0]["transcript"],
                    ),
                },
            ],
        },
        {
            "start_index": 1,
            "count": 1,
            "line_index": 2,
            "match_score": 100.0,
            "transcript": segments[1]["transcript"],
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [
                {"line_index": 2, "match_score": 100.0}
            ],
        },
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    recovered = next(action for action in joined if action["line_index"] == 1)
    assert recovered["transcript"] == "You should leave. Now."
    assert recovered["fragment_join_provisional"] is False


def test_missing_long_line_can_recover_from_fragments_assigned_elsewhere() -> None:
    lines = [
        {
            "line_id": "target",
            "line": (
                "Akhori, dongo do-daibethe! Good to see you. "
                "How are things at the docks?"
            ),
        },
        {"line_id": "name", "line": "Akhori."},
        {"line_id": "yoku", "line": "Dongo do-daibethe."},
        {"line_id": "greeting", "line": "Good to see you."},
        {
            "line_id": "question",
            "line": "How are things at the docks?",
        },
    ]
    transcripts = [
        "Akhori.",
        "Dongo do-daibethe.",
        "Good to see you.",
        "How are things at the docks?",
    ]
    segments = [
        {
            "start_seconds": index * 1.2,
            "end_seconds": index * 1.2 + 1.0,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": index,
            "count": 1,
            "line_index": index + 1,
            "match_score": 100.0,
            "transcript": transcript,
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [
                {"line_index": index + 1, "match_score": 100.0}
            ],
        }
        for index, transcript in enumerate(transcripts)
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 4,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    recovered = next(action for action in joined if action["line_index"] == 0)
    assert recovered["start_index"] == 0
    assert recovered["count"] == 4
    assert recovered["fragment_line_seed"] is True
    assert recovered["fragment_join_provisional"] is False


def test_uncertain_short_boundary_audio_is_kept_for_review() -> None:
    line_text = "No... escape..."
    lines = [
        {"line_id": "target", "line": line_text},
        {"line_id": "other", "line": "Oh, come on! I need it more!"},
    ]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 0.9,
            "transcript": "Oh",
            "asr_probability": 0.08,
        },
        {
            "start_seconds": 0.9,
            "end_seconds": 2.7,
            "transcript": "Escape.",
            "asr_probability": 0.43,
        },
    ]
    actions = [
        {
            "start_index": 0,
            "count": 1,
            "line_index": 1,
            "match_score": text_similarity(
                lines[1]["line"],
                segments[0]["transcript"],
            ),
            "transcript": segments[0]["transcript"],
            "duration_plausibility": 50.0,
            "order_hint": 0.0,
            "top_matches": [],
        },
        {
            "start_index": 1,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(
                line_text,
                segments[1]["transcript"],
            ),
            "transcript": segments[1]["transcript"],
            "duration_plausibility": 50.0,
            "order_hint": 0.0,
            "top_matches": [],
        },
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    recovered = next(action for action in joined if action["line_index"] == 0)
    assert recovered["start_index"] == 0
    assert recovered["count"] == 2
    assert recovered["forced_review_reason"] == "UNCERTAIN_BOUNDARY_AUDIO"


def test_fragment_join_uses_shortest_text_bounded_span() -> None:
    line_text = (
        "Slow, but picking up. Lost a lot of good customers when Darius "
        "let half his workers go."
    )
    lines = [{"line_id": "line", "line": line_text}]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 1.5,
            "transcript": "Slow but picking up",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 1.8,
            "end_seconds": 5.0,
            "transcript": (
                "Lost a lot of good customers when Darius let half his "
                "workers go"
            ),
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 5.3,
            "end_seconds": 6.0,
            "transcript": "",
            "asr_probability": None,
        },
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": 0,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(
                line_text,
                segments[0]["transcript"],
            ),
            "transcript": segments[0]["transcript"],
            "duration_plausibility": 50.0,
            "order_hint": 0.0,
            "top_matches": [],
        },
        {
            "type": "assigned",
            "start_index": 1,
            "count": 2,
            "line_index": 0,
            "match_score": text_similarity(
                line_text,
                segments[1]["transcript"],
            ),
            "transcript": segments[1]["transcript"],
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [],
        },
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 3,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            },
            recovery={"max_segments": 3},
        ),
    )

    assert [(action["start_index"], action["count"]) for action in joined] == [
        (0, 2)
    ]


def test_fragment_join_recovers_unselected_preceding_sentence_from_word_span() -> None:
    line_text = (
        "Some old sailor superstition, maybe. Taunting death. "
        "Seems more like tempting fate, but what would I know? "
        "I've never been one for the sea."
    )
    lines = [{"line_id": "line", "line": line_text}]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 1.0,
            "transcript": "Old sailor superstition maybe",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 1.2,
            "end_seconds": 2.0,
            "transcript": "Death Seems",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 2.2,
            "end_seconds": 5.0,
            "transcript": (
                "More like tempting fate but what would I know "
                "I've never been one for the sea"
            ),
            "asr_probability": 0.95,
        },
    ]
    normalized_words = line_text.translate(
        str.maketrans({".": "", ",": "", "?": ""})
    ).split()
    word_duration = 4.8 / len(normalized_words)
    transcription = {
        "segments": [
            {
                "start": 0.0,
                "end": 5.0,
                "words": [
                    {
                        "start": index * word_duration,
                        "end": (index + 1) * word_duration,
                        "word": f" {word}",
                        "probability": 0.99,
                    }
                    for index, word in enumerate(normalized_words)
                ],
            }
        ]
    }
    partial = f"{segments[1]['transcript']} {segments[2]['transcript']}"
    actions = [
        {
            "type": "assigned",
            "start_index": 1,
            "count": 2,
            "line_index": 0,
            "match_score": text_similarity(line_text, partial),
            "transcript": partial,
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 3,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            },
            recovery={"max_segments": 3},
        ),
        transcription=transcription,
    )

    assert len(joined) == 1
    assert joined[0]["start_index"] == 0
    assert joined[0]["count"] == 3
    assert "Taunting" in joined[0]["transcript"]
    assert (
        sentence_fidelity(line_text, joined[0]["transcript"])[
            "missing_clause_count"
        ]
        == 0
    )


def test_fragment_join_prefers_complete_segment_asr_over_stale_session_words() -> None:
    line_text = "Maybe over here... No, nothing."
    lines = [{"line_id": "line", "line": line_text}]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 1.0,
            "transcript": "Maybe over here",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 1.2,
            "end_seconds": 2.0,
            "transcript": "No nothing",
            "asr_probability": 0.95,
        },
    ]
    transcription = {
        "segments": [
            {
                "start": 0.0,
                "end": 2.0,
                "words": [
                    {
                        "start": 0.0,
                        "end": 0.3,
                        "word": " Maybe",
                        "probability": 0.99,
                    },
                    {
                        "start": 0.3,
                        "end": 0.6,
                        "word": " over",
                        "probability": 0.99,
                    },
                    {
                        "start": 0.6,
                        "end": 0.9,
                        "word": " here",
                        "probability": 0.99,
                    },
                ],
            }
        ]
    }
    actions = [
        {
            "type": "assigned",
            "start_index": 0,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(
                line_text,
                segments[0]["transcript"],
            ),
            "transcript": segments[0]["transcript"],
            "duration_plausibility": 60.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
        transcription=transcription,
    )

    assert len(joined) == 1
    assert joined[0]["start_index"] == 0
    assert joined[0]["count"] == 2
    assert joined[0]["transcript"] == "Maybe over here No nothing"
    assert joined[0]["transcript_source"] == "segment_asr_span"


def test_fragment_join_recovers_single_clause_split_at_internal_pause() -> None:
    line_text = "Only my consciousness remains, trapped in this dream."
    lines = [{"line_id": "line", "line": line_text}]
    transcripts = [
        "Only my consciousness remains",
        "trapped in this dream",
    ]
    segments = [
        {
            "start_seconds": index * 2.0,
            "end_seconds": index * 2.0 + 1.5,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": 0,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(line_text, transcripts[0]),
            "transcript": transcripts[0],
            "duration_plausibility": 60.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    assert [(action["start_index"], action["count"]) for action in joined] == [
        (0, 2)
    ]


def test_fragment_join_can_recover_four_segments_around_partial_seed() -> None:
    line_text = (
        "The tree here is an illusion. What point would there be? "
        "No. He's swayed you!"
    )
    lines = [{"line_id": "line", "line": line_text}]
    transcripts = [
        "The tree here is an illusion",
        "What point would there be",
        "No",
        "He's swayed you",
    ]
    segments = [
        {
            "start_seconds": index * 1.2,
            "end_seconds": index * 1.2 + 1.0,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    partial = f"{transcripts[1]} {transcripts[2]}"
    actions = [
        {
            "type": "assigned",
            "start_index": 1,
            "count": 2,
            "line_index": 0,
            "match_score": text_similarity(line_text, partial),
            "transcript": partial,
            "duration_plausibility": 70.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 4,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            },
            recovery={"max_segments": 2},
        ),
    )

    assert any(
        action["start_index"] == 0 and action["count"] == 4
        for action in joined
    )


def test_fragment_join_sends_plausible_incomplete_span_to_exact_asr() -> None:
    line_text = "Bah. Fine. I can see you're no fool. What do you want?"
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 0.8,
            "transcript": "Ugh, fine",
            "asr_probability": 0.9,
        },
        {
            "start_seconds": 1.0,
            "end_seconds": 2.2,
            "transcript": "I can see you're no fool",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 2.4,
            "end_seconds": 3.2,
            "transcript": "What do you want",
            "asr_probability": 0.95,
        },
    ]
    partial = " ".join(segment["transcript"] for segment in segments[:2])
    actions = [
        {
            "type": "assigned",
            "start_index": 0,
            "count": 2,
            "line_index": 0,
            "match_score": text_similarity(line_text, partial),
            "transcript": partial,
            "duration_plausibility": 70.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=[{"line_id": "line", "line": line_text}],
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 3,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    recovered = next(
        action
        for action in joined
        if action["start_index"] == 0 and action["count"] == 3
    )
    assert recovered["fragment_join_provisional"] is True


def test_fragment_join_recovers_single_sentence_missing_end() -> None:
    line_text = (
        "I live with that woman, and there is no way anyone who's heard "
        "her speak would want to bed her, much less live with her."
    )
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 6.0,
            "transcript": (
                "I live with that woman, and there's no way anyone who's "
                "heard her speak would want to bed her."
            ),
            "asr_probability": 0.98,
        },
        {
            "start_seconds": 6.2,
            "end_seconds": 8.0,
            "transcript": "Much less live with her.",
            "asr_probability": 0.96,
        },
    ]
    partial = segments[0]["transcript"]
    actions = [
        {
            "type": "assigned",
            "start_index": 0,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(line_text, partial),
            "transcript": partial,
            "duration_plausibility": 75.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=[{"line_id": "Palioth::R53", "line": line_text}],
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 10,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 20.0,
            }
        ),
    )

    recovered = next(
        action
        for action in joined
        if action["start_index"] == 0 and action["count"] == 2
    )
    fidelity = transcript_fidelity(line_text, recovered["transcript"])
    assert fidelity["prefix_token_coverage"] == 1.0
    assert fidelity["suffix_token_coverage"] == 1.0


def test_fragment_join_recovers_seven_fragments_with_ten_segment_limit() -> None:
    line_text = (
        "I live with that woman, and there is no way anyone who's heard "
        "her speak would want to bed her, much less live with her."
    )
    transcripts = [
        "I live with that woman.",
        "And there is no way.",
        "Anyone",
        "who's heard her speak",
        "would want to bed her",
        "much less",
        "live with her",
    ]
    segments = [
        {
            "start_seconds": index * 1.2,
            "end_seconds": index * 1.2 + 1.0,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    partial = " ".join(transcripts[1:])
    actions = [
        {
            "type": "assigned",
            "start_index": 1,
            "count": 6,
            "line_index": 0,
            "match_score": text_similarity(line_text, partial),
            "transcript": partial,
            "duration_plausibility": 75.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=[{"line_id": "Palioth::R53", "line": line_text}],
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_segments": 10,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 20.0,
            }
        ),
    )

    assert any(
        action["start_index"] == 0 and action["count"] == 7
        for action in joined
    )


def test_inline_edge_cues_do_not_affect_fidelity_but_require_review() -> None:
    line_text = "(laugh) See! What did I tell you?"
    observed = "See! What did I tell you?"

    assert text_similarity(line_text, observed) == pytest.approx(100.0)
    assert sentence_fidelity(line_text, observed)["missing_clause_count"] == 0
    assert _candidate_reliability(
        line={"line": line_text},
        match_score=100.0,
        margin=20.0,
        settings={},
        observed=observed,
    ) == (False, "EDGE_VOCALIZATION_UNVERIFIED")


def test_fragment_join_searches_neighbor_before_worse_selected_candidate() -> None:
    line_text = "All right, okay... Sounds good."
    lines = [{"line_id": "line", "line": line_text}]
    transcripts = [
        "Oh all right okay",
        "Sounds good",
        "Alright, okay. Sounds good. Alright, okay.",
    ]
    segments = [
        {
            "start_seconds": index * 2.0,
            "end_seconds": index * 2.0 + 1.5,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": 2,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(line_text, transcripts[2]),
            "transcript": transcripts[2],
            "duration_plausibility": 45.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    assert any(
        action["start_index"] == 0 and action["count"] == 2
        for action in joined
    )


def test_fragment_join_skips_line_with_complete_contraction_variant() -> None:
    line_text = "Could've been worse."
    lines = [{"line_id": "line", "line": line_text}]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 1.0,
            "transcript": "Could have been worse",
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 1.2,
            "end_seconds": 1.8,
            "transcript": "Seems",
            "asr_probability": 0.95,
        },
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": 0,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(
                line_text,
                segments[0]["transcript"],
            ),
            "transcript": segments[0]["transcript"],
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
            }
        ),
    )

    assert joined == []


def test_fragment_join_recovers_later_takes_when_line_has_complete_take() -> None:
    line_text = (
        "This place is nothing but a rest stop for eccentrics and recluses."
    )
    lines = [{"line_id": "line", "line": line_text}]
    transcripts = [
        f"Thanks for watching! {line_text}",
        "This place",
        "It's nothing but a rest stop for eccentrics and recluses",
        "This place",
        "It's nothing but a rest stop for eccentrics and recluses",
    ]
    segments = [
        {
            "start_seconds": index * 6.0,
            "end_seconds": index * 6.0 + 5.0,
            "transcript": transcript,
            "asr_probability": 0.95,
        }
        for index, transcript in enumerate(transcripts)
    ]
    actions = [
        {
            "type": "assigned",
            "start_index": index,
            "count": 1,
            "line_index": 0,
            "match_score": text_similarity(line_text, transcript),
            "transcript": transcript,
            "duration_plausibility": 80.0,
            "order_hint": 0.0,
            "top_matches": [],
        }
        for index, transcript in enumerate(transcripts)
    ]

    joined = _multisentence_fragment_join_actions(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            span_search={
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 12.0,
            }
        ),
    )

    assert any(
        action["start_index"] == 1 and action["count"] == 2
        for action in joined
    )
    assert any(
        action["start_index"] == 3 and action["count"] == 2
        for action in joined
    )
    assert all(action["start_index"] > 0 for action in joined)


def test_unordered_alignment_does_not_merge_empty_boundary_segment() -> None:
    lines = [
        {
            "line_id": "line",
            "line": (
                "Lost a lot of good customers when Darius let half his "
                "workers go, and the new ships do not make up for it."
            ),
        }
    ]
    segments = [
        {
            "start_seconds": 0.0,
            "end_seconds": 4.0,
            "transcript": lines[0]["line"],
            "asr_probability": 0.95,
        },
        {
            "start_seconds": 4.3,
            "end_seconds": 6.0,
            "transcript": "",
            "asr_probability": None,
        },
    ]

    actions = order_independent_align(
        segments,
        lines,
        _alignment_settings(
            span_search={
                "max_segments": 2,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
                "minimum_score": 45.0,
                "require_text_boundaries": True,
            }
        ),
    )

    assert len(actions) == 1
    assert actions[0]["start_index"] == 0
    assert actions[0]["count"] == 1


def test_untranscribed_audio_in_merge_prevents_auto_acceptance() -> None:
    unsafe = _has_unsafe_untranscribed_merge(
        action={"start_index": 0, "count": 2},
        base_segments=[
            {
                "start_seconds": 0.0,
                "end_seconds": 1.0,
                "transcript": "Hello there",
                "metrics": {"duration_seconds": 1.0, "rms_dbfs": -20.0},
            },
            {
                "start_seconds": 1.2,
                "end_seconds": 2.2,
                "transcript": "",
                "metrics": {"duration_seconds": 1.0, "rms_dbfs": -24.0},
            },
        ],
        settings={},
    )

    assert unsafe is True
    assert _candidate_reliability(
        line={"line": "Hello there."},
        match_score=100.0,
        margin=50.0,
        settings={},
        observed="Hello there",
        unsafe_untranscribed_merge=unsafe,
    ) == (False, "MERGED_UNTRANSCRIBED_AUDIO")


def test_long_disjoint_edge_hallucination_is_an_unsafe_merge() -> None:
    unsafe = _has_unsafe_untranscribed_merge(
        action={"start_index": 0, "count": 2},
        base_segments=[
            {
                "transcript": "The bugs always go for this one's tail.",
                "metrics": {"duration_seconds": 3.0, "rms_dbfs": -18.0},
            },
            {
                "transcript": "Transcribed by subtitles service",
                "metrics": {"duration_seconds": 10.0, "rms_dbfs": -23.0},
            },
        ],
        segment={
            "transcript": "The bugs always go for this one's tail.",
            "metrics": {"duration_seconds": 13.0},
            "words": [
                {"word": "The", "start": 0.0, "end": 0.2},
                {"word": "tail", "start": 2.0, "end": 2.4},
            ],
        },
        settings={},
    )

    assert unsafe is True


def test_quiet_boundary_padding_does_not_reject_complete_merge() -> None:
    unsafe = _has_unsafe_untranscribed_merge(
        action={
            "start_index": 0,
            "count": 2,
            "transcript": "Hello there.",
        },
        base_segments=[
            {
                "transcript": "Thank you.",
                "metrics": {"duration_seconds": 1.2, "rms_dbfs": -48.0},
                "segment_asr": {"silence_rejected": True},
            },
            {
                "transcript": "Hello there.",
                "metrics": {"duration_seconds": 1.5, "rms_dbfs": -20.0},
            },
        ],
        segment={
            "start_seconds": 0.0,
            "end_seconds": 2.7,
            "metrics": {"duration_seconds": 2.7},
            "transcript": "Hello there.",
            "words": [
                {"word": "Hello", "start": 1.2, "end": 1.6},
                {"word": "there", "start": 1.6, "end": 2.2},
            ],
        },
        settings={},
    )

    assert unsafe is False


def test_exact_asr_boundary_gap_rejects_collapsed_repeated_take() -> None:
    unsafe = _has_unsafe_untranscribed_merge(
        action={
            "start_index": 0,
            "count": 2,
            "transcript": "Hello there. Hello there.",
        },
        base_segments=[
            {
                "transcript": "Hello there.",
                "metrics": {"duration_seconds": 1.2, "rms_dbfs": -20.0},
            },
            {
                "transcript": "Hello there.",
                "metrics": {"duration_seconds": 1.5, "rms_dbfs": -20.0},
            },
        ],
        segment={
            "start_seconds": 0.0,
            "end_seconds": 2.7,
            "metrics": {"duration_seconds": 2.7},
            "transcript": "Hello there.",
            "words": [
                {"word": "Hello", "start": 0.1, "end": 0.5},
                {"word": "there", "start": 0.5, "end": 1.0},
            ],
        },
        settings={},
    )

    assert unsafe is True


def test_exact_asr_omitting_boundary_pfft_rejects_noisy_merge() -> None:
    unsafe = _has_unsafe_untranscribed_merge(
        action={
            "start_index": 0,
            "count": 2,
            "transcript": "Pfft. Better get going then.",
        },
        base_segments=[
            {
                "transcript": "Pfft.",
                "metrics": {"duration_seconds": 0.75, "rms_dbfs": -47.0},
                "segment_asr": {
                    "primary": {"transcript": "Pfft."},
                },
            },
            {
                "transcript": "Better get going then.",
                "metrics": {"duration_seconds": 2.0, "rms_dbfs": -18.0},
            },
        ],
        segment={
            "start_seconds": 0.0,
            "end_seconds": 2.75,
            "metrics": {"duration_seconds": 2.75},
            "transcript": "Better get going then.",
            "words": [
                {"word": "Better", "start": 0.8, "end": 1.2},
                {"word": "then", "start": 1.8, "end": 2.2},
            ],
        },
        settings={},
    )

    assert unsafe is True


def test_short_boundary_clause_can_use_three_decode_consensus() -> None:
    lines = [
        {
            "line_id": "target",
            "line": (
                "I know! That was different. "
                "It was a long time ago..."
            ),
        }
    ]
    evaluator = TranscriptEvaluator(lines, {})

    rescued = _boundary_clause_consensus_transcript(
        line_index=0,
        exact_transcript=(
            "Hey, no! That was different. It was a long time ago."
        ),
        preliminary_transcript=(
            "I know! That was different. It was... A long time ago."
        ),
        recording_transcript=(
            "Know That Was Different It Was A Long Time Ago"
        ),
        evaluator=evaluator,
        settings={},
    )

    assert rescued == (
        "I know! That was different. It was... A long time ago."
    )


def test_boundary_clause_consensus_requires_recording_support() -> None:
    lines = [
        {
            "line_id": "target",
            "line": (
                "I know! That was different. "
                "It was a long time ago..."
            ),
        }
    ]
    evaluator = TranscriptEvaluator(lines, {})

    rescued = _boundary_clause_consensus_transcript(
        line_index=0,
        exact_transcript=(
            "Hey, no! That was different. It was a long time ago."
        ),
        preliminary_transcript=(
            "I know! That was different. It was... A long time ago."
        ),
        recording_transcript=(
            "Hey No That Was Different It Was A Long Time Ago"
        ),
        evaluator=evaluator,
        settings={},
    )

    assert rescued is None


def test_ellipsis_hesitation_can_use_constituent_asr_without_recording_support() -> None:
    line_text = (
        "I... err... I misspoke. I simply refer to the good fortune "
        "of the Duilius household, in all aspects of life."
    )
    lines = [{"line_id": "target", "line": line_text}]
    evaluator = TranscriptEvaluator(lines, {})
    preliminary = (
        "I, um... I misspoke. I simply refer to the good fortune "
        "of the Duilius household, in all aspects of life."
    )
    exact = (
        "I misspoke. I simply refer to the good fortune "
        "of the Duilius household, in all aspects of life."
    )

    fidelity = evaluator.evaluate(0, preliminary)
    assert fidelity.sentence["clause_count"] == 2
    assert fidelity.sentence["missing_clause_count"] == 0
    assert fidelity.sentence["clauses_in_order"] is True

    rescued = _boundary_clause_consensus_transcript(
        line_index=0,
        exact_transcript=exact,
        preliminary_transcript=preliminary,
        recording_transcript="",
        evaluator=evaluator,
        settings={},
    )

    assert rescued == preliminary


def test_word_timestamp_gaps_split_a_region_into_take_candidates() -> None:
    transcription = {
        "segments": [
            {
                "start": 0.0,
                "end": 4.0,
                "text": "Take this Take this",
                "words": [
                    {"start": 0.2, "end": 0.5, "word": " Take"},
                    {"start": 0.5, "end": 0.8, "word": " this"},
                    {"start": 1.5, "end": 1.8, "word": " Take"},
                    {"start": 1.8, "end": 2.1, "word": " this"},
                ],
            }
        ]
    }
    regions = [
        {
            "speech_start": 0.0,
            "speech_end": 4.0,
            "start": 0.0,
            "end": 4.0,
        }
    ]

    refined = split_regions_on_word_gaps(
        regions,
        transcription,
        duration_seconds=4.0,
        settings={
            "word_split_enabled": True,
            "word_split_gap_seconds": 0.55,
            "word_split_min_region_seconds": 1.0,
            "word_split_max_boundaries": 2,
            "minimum_segment_seconds": 0.15,
            "pre_padding_seconds": 0.1,
            "post_padding_seconds": 0.1,
        },
    )

    assert len(refined) == 2
    assert all(region["split_source"] == "word_gap" for region in refined)


def test_segment_padding_is_clamped_to_a_shared_silence_boundary() -> None:
    regions = [
        {
            "speech_start": 0.0,
            "speech_end": 1.0,
            "start": 0.0,
            "end": 1.4,
        },
        {
            "speech_start": 1.2,
            "speech_end": 2.0,
            "start": 0.9,
            "end": 2.0,
        },
    ]

    result = prevent_region_overlaps(regions)

    assert result[0]["end"] == pytest.approx(1.1)
    assert result[1]["start"] == pytest.approx(1.1)


def test_word_gap_splitting_adapts_beyond_soft_boundary_cap() -> None:
    words = []
    for index in range(8):
        start = index * 4.0 + 0.2
        words.append(
            {
                "start": start,
                "end": start + 0.5,
                "word": f" take{index}",
            }
        )
    transcription = {
        "segments": [
            {
                "start": 0.0,
                "end": 32.0,
                "text": "takes",
                "words": words,
            }
        ]
    }

    refined = split_regions_on_word_gaps(
        [
            {
                "speech_start": 0.0,
                "speech_end": 32.0,
                "start": 0.0,
                "end": 32.0,
            }
        ],
        transcription,
        duration_seconds=32.0,
        settings={
            "word_split_enabled": True,
            "word_split_gap_seconds": 0.3,
            "word_split_min_region_seconds": 1.5,
            "word_split_max_boundaries": 2,
            "word_split_max_segment_seconds": 8.0,
            "minimum_segment_seconds": 0.15,
            "pre_padding_seconds": 0.1,
            "post_padding_seconds": 0.1,
        },
    )

    assert len(refined) > 3
    assert max(
        region["speech_end"] - region["speech_start"] for region in refined
    ) <= 8.0


def test_duplicate_weak_order_assigns_separate_take_groups() -> None:
    lines = [
        {"line_id": "a", "line": "Found you!"},
        {"line_id": "b", "line": "Found you!"},
    ]
    segments = [
        {"start_seconds": 0.0, "end_seconds": 1.0},
        {"start_seconds": 20.0, "end_seconds": 21.0},
    ]
    actions = [
        {
            "start_index": 0,
            "count": 1,
            "line_index": 0,
            "match_score": 100.0,
            "top_matches": [
                {"line_index": 0, "match_score": 100.0},
                {"line_index": 1, "match_score": 100.0},
            ],
        },
        {
            "start_index": 1,
            "count": 1,
            "line_index": 0,
            "match_score": 100.0,
            "top_matches": [
                {"line_index": 0, "match_score": 100.0},
                {"line_index": 1, "match_score": 100.0},
            ],
        },
    ]

    _apply_duplicate_line_policy(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            duplicates={
                "policy": "weak_order",
                "take_group_gap_seconds": 5.0,
            }
        ),
    )

    assert [action["line_index"] for action in actions] == [0, 1]
    assert all(action["duplicate_resolved"] for action in actions)


def test_duplicate_weak_order_distributes_nearby_distinct_takes() -> None:
    lines = [
        {"line_id": "a", "line": "I'll bury you!"},
        {"line_id": "b", "line": "I'll bury you!"},
    ]
    segments = [
        {"start_seconds": 0.0, "end_seconds": 1.0},
        {"start_seconds": 1.2, "end_seconds": 2.2},
    ]
    actions = [
        {
            "start_index": index,
            "count": 1,
            "line_index": 0,
            "match_score": 100.0,
            "top_matches": [
                {"line_index": 0, "match_score": 100.0},
                {"line_index": 1, "match_score": 100.0},
            ],
        }
        for index in range(2)
    ]

    _apply_duplicate_line_policy(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            duplicates={
                "policy": "weak_order",
                "take_group_gap_seconds": 12.0,
            }
        ),
    )

    assert [action["line_index"] for action in actions] == [0, 1]
    assert all(action["duplicate_resolved"] for action in actions)


def test_duplicate_weak_order_ignores_weak_matches_when_grouping_takes() -> None:
    lines = [
        {"line_id": "a", "line": "Found you!"},
        {"line_id": "b", "line": "Found you!"},
    ]
    correct_times = {0.0, 1.0, 100.0, 101.0, 140.0, 141.0}
    start_times = [
        0.0,
        1.0,
        *[float(value) for value in range(10, 100, 10)],
        100.0,
        101.0,
        110.0,
        120.0,
        130.0,
        140.0,
        141.0,
        *[float(value) for value in range(150, 401, 10)],
    ]
    segments = [
        {
            "start_seconds": start_seconds,
            "end_seconds": start_seconds + 0.8,
        }
        for start_seconds in start_times
    ]
    actions = [
        {
            "start_index": index,
            "count": 1,
            "line_index": 0,
            "match_score": (
                100.0 if start_seconds in correct_times else 79.0
            ),
            "top_matches": [
                {
                    "line_index": 0,
                    "match_score": (
                        100.0 if start_seconds in correct_times else 79.0
                    ),
                },
                {
                    "line_index": 1,
                    "match_score": (
                        100.0 if start_seconds in correct_times else 79.0
                    ),
                },
            ],
        }
        for index, start_seconds in enumerate(start_times)
    ]

    _apply_duplicate_line_policy(
        actions,
        lines=lines,
        base_segments=segments,
        settings=_alignment_settings(
            duplicates={
                "policy": "weak_order",
                "take_group_gap_seconds": 12.0,
            },
            reliability={"short": {"minimum_score": 88.0}},
        ),
    )

    assigned_correct_times = {
        start_times[int(action["start_index"])]: int(action["line_index"])
        for action in actions
        if float(action["match_score"]) == 100.0
    }
    assert assigned_correct_times == {
        0.0: 0,
        1.0: 0,
        100.0: 0,
        101.0: 0,
        140.0: 1,
        141.0: 1,
    }


def test_duplicate_review_and_reuse_expand_only_identical_targets() -> None:
    lines = [
        {"line": "Same words."},
        {"line": "Same words!"},
        {"line": "Different words."},
    ]
    action = {
        "type": "assigned",
        "start_index": 0,
        "count": 1,
        "line_index": 0,
        "match_score": 100.0,
        "confidence_margin": 0.0,
        "transcript": "Same words",
        "duration_plausibility": 100.0,
        "order_hint": 0.0,
    }

    review_actions = _expand_alignment_actions(
        [action],
        lines=lines,
        settings=_alignment_settings(duplicates={"policy": "review"}),
    )
    reuse_actions = _expand_alignment_actions(
        [action],
        lines=lines,
        settings=_alignment_settings(duplicates={"policy": "reuse"}),
    )

    assert [item["line_index"] for item in review_actions] == [0, 1]
    assert all(item["is_primary_match"] for item in review_actions)
    assert not any(item["duplicate_resolved"] for item in review_actions)
    assert [item["line_index"] for item in reuse_actions] == [0, 1]
    assert all(item["duplicate_resolved"] for item in reuse_actions)


def test_line_review_types_nonverbal_lines_and_uses_audible_unmatched_pool(
    tmp_path: Path,
) -> None:
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 3,
        "quest": "",
        "context": "Standing near the market.",
        "line": "(cough)",
        "acting_note": "A restrained cough.",
        "emotion": "",
        "target_filename": "cough_target",
    }
    review_path = tmp_path / "line_review.json"
    review = build_line_review(
        source_lines=[line],
        candidates_by_line={},
        unmatched_segments=[
            {
                "segment_id": "audible",
                "segment_file": "audible.wav",
                "transcript": "",
                "technical_score": 80.0,
                "audible": True,
            },
            {
                "segment_id": "quiet",
                "segment_file": "quiet.wav",
                "transcript": "",
                "technical_score": 90.0,
                "audible": False,
            },
        ],
    )
    save_line_review(review_path, review)
    loaded = load_line_review(review_path)

    assert loaded["lines"][0]["type"] == "nonverbal"
    assert loaded["lines"][0]["context"] == "Standing near the market."
    assert loaded["lines"][0]["acting_note"] == "A restrained cough."
    assert loaded["lines"][0]["status"] == "REVIEW"
    assert loaded["lines"][0]["selected_segment_id"] is None
    assert loaded["lines"][0]["candidates"] == []
    assert [
        segment["segment_id"] for segment in loaded["unmatched_segments"]
    ] == ["audible"]


def test_unreliable_verbal_merge_does_not_claim_unmatched_base_segments() -> None:
    normal_line = {
        "line_id": "Sheet::R65",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 65,
        "line": "Anyone there?",
        "target_filename": "anyone_there",
    }
    nonverbal_line = {
        "line_id": "Sheet::R64",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 64,
        "line": "Huh?",
        "target_filename": "huh",
    }
    reliable_candidate = {
        "segment_id": "session__s00027",
        "segment_file": "s00027.wav",
        "session_id": "session",
        "base_indices": [26],
        "transcript": "Anyone there?",
        "match_score": 100.0,
        "selection_score": 121.0,
        "is_primary_match": True,
        "reliable": True,
        "reliability_reason": "",
    }
    unreliable_merge = {
        "segment_id": "session__m00025_00027",
        "segment_file": "m00025_00027.wav",
        "session_id": "session",
        "base_indices": [24, 25, 26],
        "transcript": "Huh? Huh? Anyone there?",
        "match_score": 91.4,
        "selection_score": 112.7,
        "is_primary_match": True,
        "reliable": False,
        "reliability_reason": "POSSIBLE_REPEATED_TAKES",
    }
    unmatched_segments = [
        {
            "segment_id": f"session__s{base_index + 1:05d}",
            "segment_file": f"s{base_index + 1:05d}.wav",
            "session_id": "session",
            "base_index": base_index,
            "transcript": "Huh?",
            "technical_score": technical_score,
            "audible": True,
        }
        for base_index, technical_score in ((24, 97.0), (25, 95.0))
    ]

    review = build_line_review(
        source_lines=[nonverbal_line, normal_line],
        candidates_by_line={
            normal_line["line_id"]: [reliable_candidate, unreliable_merge]
        },
        unmatched_segments=unmatched_segments,
    )

    reviewed_normal = next(
        line for line in review["lines"] if line["line_id"] == normal_line["line_id"]
    )
    assert [candidate["segment_id"] for candidate in reviewed_normal["candidates"]] == [
        "session__s00027",
        "session__m00025_00027",
    ]
    assert [
        segment["segment_id"] for segment in review["unmatched_segments"]
    ] == ["session__s00025", "session__s00026"]


def test_selected_line_ids_mark_shared_nonverbal_candidates() -> None:
    review_data = {
        "lines": [
            {
                "line_id": "Sheet::R3",
                "selected_segment_id": "shared-segment",
            },
            {
                "line_id": "Sheet::R4",
                "selected_segment_id": None,
            },
            {
                "line_id": "Sheet::R5",
                "selected_segment_id": "shared-segment",
            },
        ]
    }

    selected_line_ids = _selected_line_ids_by_segment(review_data)
    assert selected_line_ids == {
        "shared-segment": ["Sheet::R3", "Sheet::R5"]
    }
    selection_text, tags = _candidate_selection_display(
        line={
            "line_id": "Sheet::R4",
            "type": "nonverbal",
            "selected_segment_id": None,
        },
        segment_id="shared-segment",
        selected_line_ids=selected_line_ids,
    )
    assert selection_text == "In use (2)"
    assert tags == ("selected_elsewhere",)


def test_context_display_removes_duplicate_multiple_context_headings() -> None:
    context = (
        "Used in multiple contexts:\n"
        "At the market.\n"
        "  USED IN MULTIPLE CONTEXTS:  \n"
        "Inside the inn."
    )

    assert _context_display_text(context) == (
        "Used in multiple contexts:\nAt the market.\nInside the inn."
    )


def test_finalize_omits_unselected_lines(tmp_path: Path) -> None:
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 3,
        "line": "No take was selected.",
        "target_filename": "unselected",
    }
    review_path = tmp_path / "line_review.json"
    save_line_review(
        review_path,
        build_line_review(
            source_lines=[line],
            candidates_by_line={},
            unmatched_segments=[],
        ),
    )
    write_json(tmp_path / "segments_manifest.json", {"sessions": []})

    result = finalize_review(
        project_dir=tmp_path,
        project={"export": {}},
        review_path=review_path,
        output_dir=tmp_path / "final",
    )

    assert result["export_count"] == 0
    assert result["error_count"] == 0


def test_mark_for_retake_clears_selection_and_is_preserved(
    tmp_path: Path,
) -> None:
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 3,
        "line": "Record this again.",
        "target_filename": "retake",
    }
    candidate = {
        "segment_id": "session__s00001",
        "segment_file": "segment.wav",
        "session_id": "session",
        "base_indices": [0],
        "transcript": line["line"],
        "match_score": 100.0,
        "selection_score": 100.0,
        "reliable": True,
    }
    review_path = tmp_path / "line_review.json"
    review = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: [candidate]},
        unmatched_segments=[],
    )
    save_line_review(review_path, review)
    app = DialogueReviewApp.__new__(DialogueReviewApp)
    app.review_path = review_path
    app.review_data = review
    app.selected_line_id = line["line_id"]
    app.render_lines = lambda: None
    app.render_candidates = lambda: None

    app.mark_for_retake()

    saved = load_line_review(review_path)
    assert saved["lines"][0]["status"] == "RETAKE"
    assert saved["lines"][0]["selected_segment_id"] is None

    regenerated = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: [candidate]},
        unmatched_segments=[],
    )
    preserved = preserve_manual_selections(regenerated, saved)
    assert preserved["lines"][0]["status"] == "RETAKE"
    assert preserved["lines"][0]["selected_segment_id"] is None


def test_retake_status_rejects_a_selected_candidate(tmp_path: Path) -> None:
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 3,
        "line": "Record this again.",
        "target_filename": "retake",
    }
    review = build_line_review(
        source_lines=[line],
        candidates_by_line={},
        unmatched_segments=[],
    )
    review["lines"][0]["status"] = "RETAKE"
    review["lines"][0]["selected_segment_id"] = "not-allowed"

    with pytest.raises(ValueError, match="cannot have a selected segment"):
        save_line_review(tmp_path / "line_review.json", review)


def test_export_retake_script_keeps_rows_on_their_original_formatted_sheets(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "lines.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    second = workbook.create_sheet("Second")
    third = workbook.create_sheet("Third")
    headers = [
        "Quest",
        "Context",
        "Line to Speak",
        "Acting Note",
        "Facial Emotion",
        "Filename",
    ]
    for worksheet in (first, second, third):
        worksheet.cell(1, 2).value = "You are voicing Actor"
        for column, header in enumerate(headers, start=1):
            worksheet.cell(2, column).value = header
        worksheet.freeze_panes = "A3"
        worksheet.column_dimensions["C"].width = 42
    first.append(
        ["Quest A", "Context A", "Keep this line.", "", "", "keep_a"]
    )
    first.append(
        ["", "Context B", "Retake from first.", "Softer", "", "retake_a"]
    )
    second.append(
        ["Quest B", "Context C", "Retake from second.", "", "Sad", "retake_b"]
    )
    third.append(
        ["Quest C", "Context D", "Keep this too.", "", "", "keep_c"]
    )
    first["C4"].fill = PatternFill("solid", fgColor="FFF2CC")
    second["C3"].fill = PatternFill("solid", fgColor="DDEBF7")
    workbook.save(workbook_path)
    workbook.close()

    source_data = parse_workbook(workbook_path)
    write_json(tmp_path / "source_lines.json", source_data)
    review = build_line_review(
        source_lines=source_data["lines"],
        candidates_by_line={},
        unmatched_segments=[],
    )
    retake_ids = {"First::R4", "Second::R3"}
    for review_line in review["lines"]:
        if review_line["line_id"] in retake_ids:
            review_line["status"] = "RETAKE"
    review_path = tmp_path / "line_review.json"
    save_line_review(review_path, review)

    output_path = tmp_path / "retakes.xlsx"
    result = export_retake_script(
        project_dir=tmp_path,
        project={
            "workbook": "lines.xlsx",
            "source_lines": "source_lines.json",
        },
        review_path=review_path,
        output_path=output_path,
    )

    assert result["export_count"] == 2
    assert result["sheet_names"] == ["First", "Second"]
    exported = load_workbook(output_path)
    assert exported.sheetnames == ["First", "Second"]
    first_export = exported["First"]
    second_export = exported["Second"]
    assert first_export.freeze_panes == "A3"
    assert second_export.freeze_panes == "A3"
    assert first_export.column_dimensions["C"].width == 42
    assert second_export.column_dimensions["C"].width == 42
    assert first_export["A3"].value == "Quest A"
    assert first_export["C3"].value == "Retake from first."
    assert first_export["F3"].value == "retake_a"
    assert second_export["A3"].value == "Quest B"
    assert second_export["C3"].value == "Retake from second."
    assert second_export["F3"].value == "retake_b"
    assert first_export["C3"].fill.fgColor.rgb == "00FFF2CC"
    assert second_export["C3"].fill.fgColor.rgb == "00DDEBF7"
    exported.close()

    parsed_export = parse_workbook(output_path)
    assert parsed_export["sheet_count"] == 2
    assert [line["line"] for line in parsed_export["lines"]] == [
        "Retake from first.",
        "Retake from second.",
    ]
    with pytest.raises(ValueError, match="cannot replace the source workbook"):
        export_retake_script(
            project_dir=tmp_path,
            project={
                "workbook": "lines.xlsx",
                "source_lines": "source_lines.json",
            },
            review_path=review_path,
            output_path=workbook_path,
            overwrite=True,
        )


def test_review_candidates_keep_primary_top_score_cluster() -> None:
    def candidate(
        segment_id: str,
        score: float,
        *,
        primary: bool = True,
    ) -> dict:
        return {
            "segment_id": segment_id,
            "session_id": "session",
            "base_indices": [int(segment_id[-1])],
            "match_score": score,
            "selection_score": score,
            "is_primary_match": primary,
            "reliable": score >= 90.0,
            "reliability_reason": "",
        }

    retained = prune_line_candidates(
        [
            candidate("segment1", 99.0),
            candidate("segment2", 96.0),
            candidate("segment3", 88.0),
            candidate("segment4", 70.0),
            candidate("segment5", 98.0, primary=False),
        ]
    )

    assert [item["segment_id"] for item in retained] == [
        "segment1",
        "segment2",
        "segment3",
    ]


def test_review_candidates_keep_reliable_disjoint_take_below_score_cutoff() -> None:
    retained = prune_line_candidates(
        [
            {
                "segment_id": "session__s00126",
                "session_id": "session",
                "base_indices": [125],
                "match_score": 100.0,
                "selection_score": 121.0,
                "is_primary_match": True,
                "reliable": True,
                "reliability_reason": "",
            },
            {
                "segment_id": "session__s00123",
                "session_id": "session",
                "base_indices": [122],
                "match_score": 86.27450980392157,
                "selection_score": 107.45633353439032,
                "is_primary_match": True,
                "reliable": True,
                "reliability_reason": "",
            },
        ]
    )

    assert [candidate["segment_id"] for candidate in retained] == [
        "session__s00126",
        "session__s00123",
    ]


def test_review_candidates_drop_reliable_overlapping_variant_below_cutoff() -> None:
    retained = prune_line_candidates(
        [
            {
                "segment_id": "session__best",
                "session_id": "session",
                "base_indices": [122],
                "match_score": 100.0,
                "selection_score": 121.0,
                "is_primary_match": True,
                "reliable": True,
                "reliability_reason": "",
            },
            {
                "segment_id": "session__alternate",
                "session_id": "session",
                "base_indices": [122],
                "match_score": 86.0,
                "selection_score": 107.0,
                "is_primary_match": True,
                "reliable": True,
                "reliability_reason": "",
            },
        ]
    )

    assert [candidate["segment_id"] for candidate in retained] == [
        "session__best"
    ]


def test_structurally_incomplete_superset_does_not_hide_strict_join() -> None:
    oversized_provisional = {
        "segment_id": "session__m00065_00068",
        "session_id": "session",
        "base_indices": [64, 65, 66, 67],
        "match_score": 88.0,
        "selection_score": 109.5,
        "is_primary_match": True,
        "reliable": False,
        "reliability_reason": "UNCERTAIN_BOUNDARY_AUDIO",
        "missing_clause_count": 0,
        "fragment_join": True,
        "fragment_join_provisional": True,
    }
    contained_strict_join = {
        "segment_id": "session__m00067_00068",
        "session_id": "session",
        "base_indices": [66, 67],
        "match_score": 72.0,
        "selection_score": 90.5,
        "is_primary_match": True,
        "reliable": False,
        "reliability_reason": "MISSING_SENTENCE",
        "missing_clause_count": 2,
        "fragment_join": True,
        "fragment_join_provisional": False,
    }

    retained = prune_line_candidates(
        [oversized_provisional, contained_strict_join]
    )

    assert [candidate["segment_id"] for candidate in retained] == [
        "session__m00065_00068",
        "session__m00067_00068",
    ]


def test_unsafe_long_merge_does_not_hide_reliable_contained_take() -> None:
    unsafe_merge = {
        "segment_id": "session__m00197_00200",
        "session_id": "session",
        "base_indices": [196, 197, 198, 199],
        "match_score": 97.0,
        "selection_score": 118.2,
        "is_primary_match": True,
        "reliable": False,
        "reliability_reason": "MERGED_UNTRANSCRIBED_AUDIO",
    }
    clean_take = {
        "segment_id": "session__m00197_00198",
        "session_id": "session",
        "base_indices": [196, 197],
        "match_score": 97.0,
        "selection_score": 118.1,
        "is_primary_match": True,
        "reliable": True,
        "reliability_reason": "",
    }

    retained = prune_line_candidates([unsafe_merge, clean_take])

    assert [candidate["segment_id"] for candidate in retained] == [
        "session__m00197_00198"
    ]


def test_structurally_incomplete_disjoint_take_remains_reviewable() -> None:
    reliable_take = {
        "segment_id": "session__m00004_00005",
        "session_id": "session",
        "base_indices": [3, 4],
        "match_score": 97.6,
        "selection_score": 119.0,
        "is_primary_match": True,
        "reliable": True,
        "reliability_reason": "",
    }
    disjoint_take = {
        "segment_id": "session__m00006_00007",
        "session_id": "session",
        "base_indices": [5, 6],
        "match_score": 92.4,
        "selection_score": 114.1,
        "is_primary_match": True,
        "reliable": False,
        "reliability_reason": "MISSING_LINE_END",
        "fragment_join": True,
        "fragment_join_provisional": True,
    }

    retained = prune_line_candidates([reliable_take, disjoint_take])

    assert [candidate["segment_id"] for candidate in retained] == [
        "session__m00004_00005",
        "session__m00006_00007",
    ]


def test_review_auto_selects_best_quality_within_reliable_cluster() -> None:
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "excel_row": 3,
        "line": "Choose the clean take.",
        "target_filename": "clean_take",
    }
    candidates = [
        {
            "segment_id": "highest_text",
            "segment_file": "highest_text.wav",
            "session_id": "session",
            "base_indices": [0],
            "match_score": 100.0,
            "selection_score": 105.0,
            "technical_score": 50.0,
            "is_primary_match": True,
            "reliable": True,
            "reliability_reason": "",
        },
        {
            "segment_id": "cleanest",
            "segment_file": "cleanest.wav",
            "session_id": "session",
            "base_indices": [1],
            "match_score": 99.0,
            "selection_score": 116.0,
            "technical_score": 100.0,
            "is_primary_match": True,
            "reliable": True,
            "reliability_reason": "",
        },
    ]

    review = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: candidates},
        unmatched_segments=[],
    )

    assert review["lines"][0]["status"] == "AUTO_OK"
    assert review["lines"][0]["selected_segment_id"] == "cleanest"


def test_review_prefers_verified_boundary_trim_when_near_tied() -> None:
    line = {
        "line_id": "Opponent::R5",
        "sheet": "Opponent",
        "excel_row": 5,
        "line": "You should've picked an easier fight.",
        "target_filename": "easier_fight",
    }
    candidates = [
        {
            "segment_id": "trimmed",
            "segment_file": "trimmed.wav",
            "session_id": "session",
            "base_indices": [3],
            "match_score": 100.0,
            "selection_score": 121.565,
            "technical_score": 100.0,
            "is_primary_match": True,
            "reliable": True,
            "reliability_reason": "",
            "boundary_voice_trim": True,
        },
        {
            "segment_id": "intact",
            "segment_file": "intact.wav",
            "session_id": "session",
            "base_indices": [4],
            "match_score": 100.0,
            "selection_score": 121.549,
            "technical_score": 100.0,
            "is_primary_match": True,
            "reliable": True,
            "reliability_reason": "",
        },
    ]

    review = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: candidates},
        unmatched_segments=[],
    )

    review_line = review["lines"][0]
    assert review_line["selected_segment_id"] == "trimmed"
    assert review_line["candidates"][0]["boundary_voice_trim"] is True


def test_review_prefers_take_sized_candidate_when_near_tied() -> None:
    line = {
        "line_id": "Hjorik::R27",
        "sheet": "Hjorik",
        "excel_row": 27,
        "line": "A complete line with several words in it.",
        "target_filename": "take_sized",
    }
    common = {
        "session_id": "session",
        "base_indices": [0],
        "match_score": 100.0,
        "technical_score": 100.0,
        "is_primary_match": True,
        "reliable": True,
        "reliability_reason": "",
        "boundary_voice_trim": True,
    }
    candidates = [
        {
            **common,
            "segment_id": "merged",
            "segment_file": "merged.wav",
            "selection_score": 121.86,
            "duration_plausibility": 48.0,
            "start_seconds": 0.0,
            "end_seconds": 19.6,
            "duration_seconds": 19.6,
        },
        {
            **common,
            "segment_id": "single_take",
            "segment_file": "single_take.wav",
            "selection_score": 121.80,
            "duration_plausibility": 88.0,
            "start_seconds": 10.0,
            "end_seconds": 19.3,
            "duration_seconds": 9.3,
        },
    ]

    review = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: candidates},
        unmatched_segments=[],
    )

    assert review["lines"][0]["selected_segment_id"] == "single_take"


def test_review_candidates_keep_best_fragment_join_when_none_are_reliable() -> None:
    retained = prune_line_candidates(
        [
            {
                "segment_id": "partial",
                "session_id": "session",
                "base_indices": [1],
                "match_score": 91.0,
                "selection_score": 91.0,
                "is_primary_match": True,
                "reliable": False,
                "reliability_reason": "MISSING_SENTENCE",
                "fragment_join": False,
            },
            {
                "segment_id": "joined",
                "session_id": "session",
                "base_indices": [0, 1],
                "match_score": 73.0,
                "selection_score": 73.0,
                "is_primary_match": True,
                "reliable": False,
                "reliability_reason": "UNCERTAIN_BOUNDARY_AUDIO",
                "fragment_join": True,
            },
        ]
    )

    assert [candidate["segment_id"] for candidate in retained] == [
        "partial",
        "joined",
    ]


def test_review_regeneration_preserves_manual_selection() -> None:
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "excel_row": 3,
        "line": "Hello there.",
        "target_filename": "hello",
    }
    candidate = {
        "segment_id": "session__s00001",
        "segment_file": "segment.wav",
        "session_id": "session",
        "base_indices": [0],
        "transcript": "Hello there.",
        "match_score": 100.0,
        "selection_score": 100.0,
        "reliable": True,
    }
    previous = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: [candidate]},
        unmatched_segments=[],
    )
    previous["lines"][0]["status"] = "MANUALLY_REVIEWED"
    new = build_line_review(
        source_lines=[line],
        candidates_by_line={},
        unmatched_segments=[],
    )

    merged = preserve_manual_selections(new, previous)

    assert merged["lines"][0]["status"] == "MANUALLY_REVIEWED"
    assert merged["lines"][0]["selected_segment_id"] == "session__s00001"
    assert merged["lines"][0]["candidates"][0]["segment_id"] == (
        "session__s00001"
    )


def test_waveform_scroll_zoom_is_cursor_centered_and_keeps_markers_visible() -> None:
    zoomed = _zoomed_sample_window(
        view_start=0,
        view_end=120000,
        context_start=0,
        context_end=120000,
        selection_start=50000,
        selection_end=70000,
        anchor_sample=60000,
        zoom_factor=0.5,
        sample_rate=1000,
    )
    assert zoomed == (30000, 90000)

    closest = _zoomed_sample_window(
        view_start=zoomed[0],
        view_end=zoomed[1],
        context_start=0,
        context_end=120000,
        selection_start=50000,
        selection_end=70000,
        anchor_sample=50000,
        zoom_factor=0.01,
        sample_rate=1000,
    )
    assert closest[1] - closest[0] == 20100
    assert closest[0] <= 50000
    assert closest[1] >= 70000

    zoomed_out = _zoomed_sample_window(
        view_start=closest[0],
        view_end=closest[1],
        context_start=0,
        context_end=120000,
        selection_start=50000,
        selection_end=70000,
        anchor_sample=50000,
        zoom_factor=100.0,
        sample_rate=1000,
    )
    assert zoomed_out == (0, 120000)


def test_waveform_opens_close_and_right_drag_pans_the_timeline() -> None:
    initial = _initial_segment_window(
        context_start=0,
        context_end=120000,
        selection_start=50000,
        selection_end=70000,
        sample_rate=1000,
    )
    assert initial == (46500, 73500)
    assert 0.70 < 20000 / (initial[1] - initial[0]) < 0.80

    dragged_right = _panned_sample_window(
        view_start=initial[0],
        view_end=initial[1],
        context_start=0,
        context_end=120000,
        drag_pixels=100,
        canvas_width=1000,
    )
    assert dragged_right == (43800, 70800)

    dragged_left = _panned_sample_window(
        view_start=initial[0],
        view_end=initial[1],
        context_start=0,
        context_end=120000,
        drag_pixels=-100,
        canvas_width=1000,
    )
    assert dragged_left == (49200, 76200)

    clamped = _panned_sample_window(
        view_start=initial[0],
        view_end=initial[1],
        context_start=0,
        context_end=120000,
        drag_pixels=100000,
        canvas_width=1000,
    )
    assert clamped == (0, 27000)


def test_copy_edit_saves_quickly_then_can_transcribe_and_delete_manual_candidate(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.review as review_module

    transcription_calls = []

    def fake_transcribe_candidate_span(**kwargs: Any) -> dict[str, Any]:
        transcription_calls.append(kwargs)
        return {
            "transcript": "New exact edited transcript.",
            "word_count": 4,
            "words": [{"word": "New", "start": 0.1, "end": 0.3}],
            "asr_probability": 0.88,
        }

    monkeypatch.setattr(
        review_module,
        "transcribe_candidate_span",
        fake_transcribe_candidate_span,
    )
    sample_rate = 48000
    source_path = tmp_path / "source.wav"
    _write_tone(source_path, duration_seconds=4.0)
    source_segment_path = (
        tmp_path / "segments" / "session" / "session__s00001.wav"
    )
    source_metrics = cut_pcm_wav(
        source_path,
        source_segment_path,
        start_sample=sample_rate,
        end_sample=sample_rate * 2,
        fade_ms=0.0,
    )
    source_segment = {
        "segment_id": "session__s00001",
        "kind": "base",
        "session_id": "session",
        "source_audio": "source.wav",
        "source_sha256": "source-hash",
        "base_indices": [0],
        "start_sample": sample_rate,
        "end_sample": sample_rate * 2,
        "start_seconds": 1.0,
        "end_seconds": 2.0,
        "file": "segments/session/session__s00001.wav",
        "transcript": "Hello there.",
        "word_count": 2,
        "asr_probability": 0.97,
        "metrics": source_metrics,
    }
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "schema_version": 1,
            "settings": {"fade_ms": 0.0},
            "sessions": [
                {
                    "session_id": "session",
                    "audio": "source.wav",
                    "working_audio": "source.wav",
                    "source_sha256": "source-hash",
                    "sample_rate": sample_rate,
                    "source_frames": sample_rate * 4,
                    "segments": [source_segment],
                    "derived_segments": [],
                }
            ],
        },
    )
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "excel_row": 3,
        "line": "Hello there.",
        "target_filename": "hello",
    }
    source_candidate = {
        "segment_id": source_segment["segment_id"],
        "segment_file": source_segment["file"],
        "session_id": "session",
        "base_indices": [0],
        "transcript": "Hello there.",
        "match_score": 100.0,
        "selection_score": 100.0,
        "technical_score": 98.0,
        "is_primary_match": True,
        "reliable": True,
        "reliability_reason": "",
        "source_audio": "source.wav",
        "start_seconds": 1.0,
        "end_seconds": 2.0,
        "duration_seconds": 1.0,
    }
    review_path = tmp_path / "line_review.json"
    review = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: [source_candidate]},
        unmatched_segments=[],
    )
    save_line_review(review_path, review)

    context = segment_edit_source(
        project_dir=tmp_path,
        segment_id=source_segment["segment_id"],
    )
    assert context["audio_path"] == source_path
    assert context["sample_rate"] == sample_rate

    edited = save_edited_candidate(
        project_dir=tmp_path,
        project={"segmentation": {"fade_ms": 0.0}},
        review_path=review_path,
        review_data=review,
        line_id=line["line_id"],
        source_candidate=review["lines"][0]["candidates"][0],
        start_sample=sample_rate // 2,
        end_sample=sample_rate * 5 // 2,
    )

    assert edited["manual_edit"] is True
    assert edited["edited_from_segment_id"] == source_segment["segment_id"]
    assert edited["reliable"] is False
    assert edited["reliability_reason"] == "MANUALLY_EDITED_BOUNDARIES"
    assert edited["duration_seconds"] == pytest.approx(2.0)
    assert edited["transcript"] == ""
    assert edited["transcript_source"] == "manual_copy_edit_untranscribed"
    assert edited["asr_probability"] is None
    assert transcription_calls == []
    with wave.open(str(tmp_path / edited["segment_file"]), "rb") as reader:
        assert reader.getnframes() == sample_rate * 2

    manifest = read_json(tmp_path / "segments_manifest.json")
    derived = manifest["sessions"][0]["derived_segments"]
    assert len(derived) == 1
    assert derived[0]["kind"] == "manual_edit"
    assert derived[0]["edited_from_segment_id"] == source_segment["segment_id"]
    assert derived[0]["manual_line_ids"] == [line["line_id"]]
    assert derived[0]["transcript"] == ""
    assert derived[0]["transcript_source"] == "manual_copy_edit_untranscribed"
    assert derived[0]["asr_probability"] is None
    assert derived[0]["start_sample"] == sample_rate // 2
    assert derived[0]["end_sample"] == sample_rate * 5 // 2

    loaded = load_line_review(review_path)
    assert [candidate["segment_id"] for candidate in loaded["lines"][0]["candidates"]] == [
        source_segment["segment_id"],
        edited["segment_id"],
    ]
    assert loaded["lines"][0]["selected_segment_id"] == source_segment["segment_id"]

    save_edited_candidate(
        project_dir=tmp_path,
        project={"segmentation": {"fade_ms": 0.0}},
        review_path=review_path,
        review_data=review,
        line_id=line["line_id"],
        source_candidate=review["lines"][0]["candidates"][0],
        start_sample=sample_rate // 2,
        end_sample=sample_rate * 5 // 2,
    )
    assert len(
        read_json(tmp_path / "segments_manifest.json")["sessions"][0][
            "derived_segments"
        ]
    ) == 1
    assert len(load_line_review(review_path)["lines"][0]["candidates"]) == 2
    assert transcription_calls == []

    transcribed = transcribe_edited_candidate(
        project_dir=tmp_path,
        project={"segmentation": {"fade_ms": 0.0}},
        review_path=review_path,
        review_data=review,
        segment_id=edited["segment_id"],
    )
    assert transcribed["transcript"] == "New exact edited transcript."
    assert transcribed["transcript_source"] == "candidate_asr_manual_copy_edit"
    assert transcribed["asr_probability"] == pytest.approx(0.88)
    assert len(transcription_calls) == 1

    duplicate_save = save_edited_candidate(
        project_dir=tmp_path,
        project={"segmentation": {"fade_ms": 0.0}},
        review_path=review_path,
        review_data=review,
        line_id=line["line_id"],
        source_candidate=review["lines"][0]["candidates"][0],
        start_sample=sample_rate // 2,
        end_sample=sample_rate * 5 // 2,
    )
    assert duplicate_save["transcript"] == "New exact edited transcript."
    assert len(transcription_calls) == 1
    loaded = load_line_review(review_path)

    regenerated = build_line_review(
        source_lines=[line],
        candidates_by_line={},
        unmatched_segments=[],
    )
    preserved = preserve_manual_selections(regenerated, loaded)
    assert preserved["lines"][0]["status"] == "REVIEW"
    assert [
        candidate["segment_id"]
        for candidate in preserved["lines"][0]["candidates"]
    ] == [edited["segment_id"]]

    review["lines"][0]["selected_segment_id"] = edited["segment_id"]
    review["lines"][0]["status"] = "MANUALLY_REVIEWED"
    save_line_review(review_path, review)
    finalized = finalize_review(
        project_dir=tmp_path,
        project={"export": {}},
        review_path=review_path,
        output_dir=tmp_path / "final",
        dry_run=True,
    )
    assert finalized["export_count"] == 1

    segment_path = tmp_path / edited["segment_file"]
    assert segment_path.is_file()
    cache_path = (
        tmp_path
        / "segment_transcripts"
        / "candidates"
        / f"{edited['segment_id']}.json"
    )
    write_json(cache_path, {"transcript": "cached"})
    deleted = delete_edited_candidate(
        project_dir=tmp_path,
        review_path=review_path,
        review_data=review,
        segment_id=edited["segment_id"],
    )
    assert deleted["warnings"] == []
    assert not segment_path.exists()
    assert not cache_path.exists()
    assert read_json(tmp_path / "segments_manifest.json")["sessions"][0][
        "derived_segments"
    ] == []
    after_delete = load_line_review(review_path)["lines"][0]
    assert edited["segment_id"] not in {
        candidate["segment_id"] for candidate in after_delete["candidates"]
    }
    assert after_delete["selected_segment_id"] is None
    assert after_delete["status"] == "REVIEW"


def test_missing_line_can_select_and_preserve_unmatched_segment() -> None:
    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "excel_row": 3,
        "line": "No automatic candidate.",
        "target_filename": "missing",
    }
    unmatched = {
        "segment_id": "session__s00001",
        "segment_file": "segment.wav",
        "session_id": "session",
        "base_index": 0,
        "transcript": "Maybe this is it",
        "technical_score": 82.0,
        "audible": True,
    }
    previous = build_line_review(
        source_lines=[line],
        candidates_by_line={},
        unmatched_segments=[unmatched],
    )
    previous_line = previous["lines"][0]
    assert previous_line["status"] == "MISSING"
    assert _uses_unmatched_candidates(previous_line)
    previous_line["selected_segment_id"] = unmatched["segment_id"]
    previous_line["status"] = "MANUALLY_REVIEWED"
    assert _uses_unmatched_candidates(previous_line)
    assert _selected_segment_score(previous, previous_line) == 82.0

    regenerated = build_line_review(
        source_lines=[line],
        candidates_by_line={},
        unmatched_segments=[],
    )
    merged = preserve_manual_selections(regenerated, previous)

    assert merged["lines"][0]["status"] == "MANUALLY_REVIEWED"
    assert merged["lines"][0]["selected_segment_id"] == unmatched["segment_id"]
    assert merged["unmatched_segments"][0]["segment_id"] == unmatched["segment_id"]


def test_text_aligner_excludes_nonverbal_lines() -> None:
    actions = order_independent_align(
        [
            {
                "start_seconds": 0.0,
                "end_seconds": 1.0,
                "transcript": "cough",
                "asr_probability": 0.9,
            },
            {
                "start_seconds": 2.0,
                "end_seconds": 3.0,
                "transcript": "hello there",
                "asr_probability": 0.9,
            },
        ],
        [
            {"line_id": "nonverbal", "line": "(cough)"},
            {"line_id": "spoken", "line": "Hello there."},
        ],
        _alignment_settings(
            span_search={
                "max_segments": 1,
                "minimum_score": 35.0,
            }
        ),
    )

    assert actions
    assert {action["line_index"] for action in actions} == {1}


def test_segment_asr_transcribes_empty_base_clip_without_vad_and_caches(
    tmp_path: Path,
) -> None:
    segment_file = tmp_path / "segment.wav"
    _write_tone(segment_file)
    calls = []
    fake_word = SimpleNamespace(
        start=0.1,
        end=0.4,
        word=" Yes",
        probability=0.96,
    )
    fake_segment = SimpleNamespace(
        start=0.1,
        end=0.4,
        text="Yes",
        avg_logprob=-0.05,
        no_speech_prob=0.01,
        words=[fake_word],
    )

    class FakeModel:
        def transcribe(self, *_args, **kwargs):
            calls.append(kwargs)
            return iter([fake_segment]), SimpleNamespace(language="en")

    line = {
        "line_id": "Actor::R2",
        "sheet": "Actor",
        "sheet_index": 0,
        "excel_row": 2,
        "line": "Yes?",
    }
    write_json(tmp_path / "source_lines.json", {"lines": [line]})
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "sessions": [
                {
                    "session_id": "session",
                    "segments": [
                        {
                            "segment_id": "session__s00001",
                            "kind": "base",
                            "file": "segment.wav",
                            "source_sha256": "source-hash",
                            "start_sample": 0,
                            "end_sample": 48000,
                            "start_seconds": 0.0,
                            "end_seconds": 1.0,
                            "transcript": "",
                            "words": [],
                            "word_count": 0,
                            "asr_probability": None,
                            "metrics": {"duration_seconds": 1.0},
                        }
                    ],
                    "derived_segments": [],
                }
            ]
        },
    )
    project = {
        "source_lines": "source_lines.json",
        "language": "en",
        "transcription": {
            "model": "large-v3",
            "device": "cpu",
            "compute_type": "int8",
        },
        "segment_transcription": {
            "enabled": True,
            "prompt_fallback_enabled": False,
        },
        "sessions": [
            {
                "id": "session",
                "enabled": True,
                "sheets": ["Actor"],
                "excel_rows": [],
                "line_ids": [],
            }
        ],
    }
    runtime = {"model": FakeModel()}
    transcribe_segments_project(
        project_dir=tmp_path,
        project=project,
        runtime=runtime,
    )

    manifest = read_json(tmp_path / "segments_manifest.json")
    segment = manifest["sessions"][0]["segments"][0]
    assert segment["session_transcript"] == ""
    assert segment["transcript"] == "Yes"
    assert segment["transcript_source"] == "segment_asr"
    assert segment["segment_asr"]["primary"]["asr_probability"] == pytest.approx(
        0.96
    )
    assert calls[0]["vad_filter"] is False
    assert calls[0]["condition_on_previous_text"] is False

    transcribe_segments_project(
        project_dir=tmp_path,
        project=project,
        runtime={},
    )
    assert len(calls) == 1


def test_high_no_speech_low_rms_transcript_is_rejected_as_hallucination() -> None:
    segment = {
        "metrics": {
            "duration_seconds": 1.1,
            "rms_dbfs": -48.0,
        }
    }
    primary = {
        "transcript": "Thank you.",
        "segments": [{"no_speech_prob": 0.91}],
    }

    assert _likely_silence_hallucination(segment, primary, {}) is True
    assert (
        _likely_silence_hallucination(
            {**segment, "metrics": {"rms_dbfs": -20.0}},
            primary,
            {},
        )
        is False
    )


def test_segment_asr_batches_independent_clips_with_configured_size(
    tmp_path: Path,
) -> None:
    segment_records = []
    for index in range(3):
        segment_file = tmp_path / f"segment_{index}.wav"
        _write_tone(segment_file, duration_seconds=0.5)
        segment_records.append(
            {
                "segment_id": f"session__s{index + 1:05d}",
                "kind": "base",
                "file": segment_file.name,
                "source_sha256": "source-hash",
                "start_sample": index * 24000,
                "end_sample": (index + 1) * 24000,
                "start_seconds": index * 0.5,
                "end_seconds": (index + 1) * 0.5,
                "transcript": "",
                "words": [],
                "word_count": 0,
                "asr_probability": None,
                "metrics": {"duration_seconds": 0.5},
            }
        )
    line = {
        "line_id": "Actor::R2",
        "sheet": "Actor",
        "sheet_index": 0,
        "excel_row": 2,
        "line": "Hello.",
    }
    write_json(tmp_path / "source_lines.json", {"lines": [line]})
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "sessions": [
                {
                    "session_id": "session",
                    "segments": segment_records,
                    "derived_segments": [],
                }
            ]
        },
    )
    project = {
        "source_lines": "source_lines.json",
        "language": "en",
        "transcription": {
            "model": "large-v3",
            "device": "cpu",
            "compute_type": "int8",
        },
        "segment_transcription": {
            "enabled": True,
            "batch_size": 2,
            "prompt_fallback_enabled": False,
        },
        "sessions": [
            {
                "id": "session",
                "enabled": True,
                "sheets": ["Actor"],
                "excel_rows": [],
                "line_ids": [],
            }
        ],
    }
    batch_sizes = []

    class FakeBatchedModel:
        def transcribe(self, _audio, **kwargs):
            clips = kwargs["clip_timestamps"]
            batch_sizes.append(kwargs["batch_size"])
            decoded = []
            for index, clip in enumerate(clips):
                start = int(clip["start"] * 16000) / 16000
                decoded.append(
                    SimpleNamespace(
                        start=start + 0.05,
                        end=start + 0.35,
                        text=f"Hello {index + 1}",
                        avg_logprob=-0.1,
                        no_speech_prob=0.01,
                        words=[
                            SimpleNamespace(
                                start=start + 0.05,
                                end=start + 0.25,
                                word=" Hello",
                                probability=0.95,
                            )
                        ],
                    )
                )
            return iter(decoded), SimpleNamespace(language="en")

    runtime = {
        "model": SimpleNamespace(
            feature_extractor=SimpleNamespace(sampling_rate=16000)
        ),
        "batched_model": FakeBatchedModel(),
    }
    transcribe_segments_project(
        project_dir=tmp_path,
        project=project,
        runtime=runtime,
    )

    manifest = read_json(tmp_path / "segments_manifest.json")
    segments = manifest["sessions"][0]["segments"]
    assert batch_sizes == [2, 1]
    assert manifest["segment_transcription"]["batch_size"] == 2
    assert all(segment["transcript_source"] == "segment_asr" for segment in segments)
    assert segments[1]["segment_asr"]["primary"]["words"][0]["start"] == pytest.approx(
        0.05
    )


def test_segment_asr_batches_prompted_fallbacks_that_share_a_prompt(
    tmp_path: Path,
) -> None:
    segment_records = []
    for index in range(3):
        segment_file = tmp_path / f"prompt_segment_{index}.wav"
        _write_tone(segment_file, duration_seconds=0.5)
        segment_records.append(
            {
                "segment_id": f"session__s{index + 1:05d}",
                "kind": "base",
                "file": segment_file.name,
                "source_sha256": "source-hash",
                "start_sample": index * 24000,
                "end_sample": (index + 1) * 24000,
                "start_seconds": index * 0.5,
                "end_seconds": (index + 1) * 0.5,
                "transcript": "",
                "words": [],
                "word_count": 0,
                "asr_probability": None,
                "metrics": {"duration_seconds": 0.5},
            }
        )
    line = {
        "line_id": "Actor::R2",
        "sheet": "Actor",
        "sheet_index": 0,
        "excel_row": 2,
        "line": "Hello.",
    }
    write_json(tmp_path / "source_lines.json", {"lines": [line]})
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "sessions": [
                {
                    "session_id": "session",
                    "segments": segment_records,
                    "derived_segments": [],
                }
            ]
        },
    )
    project = {
        "source_lines": "source_lines.json",
        "language": "en",
        "transcription": {
            "model": "large-v3",
            "device": "cpu",
            "compute_type": "int8",
        },
        "segment_transcription": {
            "enabled": True,
            "batch_size": 3,
            "prompt_fallback_enabled": True,
        },
        "sessions": [
            {
                "id": "session",
                "enabled": True,
                "sheets": ["Actor"],
                "excel_rows": [],
                "line_ids": [],
            }
        ],
    }
    calls = []

    class FakeBatchedModel:
        def transcribe(self, _audio, **kwargs):
            calls.append(kwargs)
            decoded = []
            for clip in kwargs["clip_timestamps"]:
                start = int(clip["start"] * 16000) / 16000
                prompted = bool(kwargs.get("initial_prompt"))
                decoded.append(
                    SimpleNamespace(
                        start=start,
                        end=start + 0.4,
                        text="Hello" if prompted else "",
                        avg_logprob=-0.1 if prompted else -1.0,
                        no_speech_prob=0.01 if prompted else 0.8,
                        words=(
                            [
                                SimpleNamespace(
                                    start=start + 0.05,
                                    end=start + 0.25,
                                    word=" Hello",
                                    probability=0.95,
                                )
                            ]
                            if prompted
                            else []
                        ),
                    )
                )
            return iter(decoded), SimpleNamespace(language="en")

    runtime = {
        "model": SimpleNamespace(
            feature_extractor=SimpleNamespace(
                sampling_rate=16000,
                chunk_length=30,
            )
        ),
        "batched_model": FakeBatchedModel(),
    }
    transcribe_segments_project(
        project_dir=tmp_path,
        project=project,
        runtime=runtime,
    )

    manifest = read_json(tmp_path / "segments_manifest.json")
    assert [call["batch_size"] for call in calls] == [3, 3]
    assert "initial_prompt" not in calls[0]
    assert calls[1]["initial_prompt"] == "Hello."
    assert manifest["segment_transcription"]["primary_inference_batch_count"] == 1
    assert manifest["segment_transcription"]["prompted_fallback_count"] == 3
    assert manifest["segment_transcription"]["prompted_inference_batch_count"] == 1
    assert all(
        segment["transcript_source"] == "segment_asr_prompted"
        for segment in manifest["sessions"][0]["segments"]
    )


def test_long_clip_avoids_batched_pipeline_truncation(tmp_path: Path) -> None:
    audio_path = tmp_path / "long.wav"
    _write_tone(audio_path, duration_seconds=30.1)
    direct_calls = []
    batched_calls = []
    decoded_segment = SimpleNamespace(
        start=0.0,
        end=30.1,
        text="Complete long clip",
        avg_logprob=-0.1,
        no_speech_prob=0.01,
        words=[],
    )

    class FakeModel:
        feature_extractor = SimpleNamespace(
            sampling_rate=16000,
            chunk_length=30,
        )

        def transcribe(self, audio, **kwargs):
            direct_calls.append((audio, kwargs))
            return iter([decoded_segment]), SimpleNamespace(language="en")

    class FakeBatchedModel:
        def transcribe(self, audio, **kwargs):
            batched_calls.append((audio, kwargs))
            raise AssertionError("Long clips must not use batched clip timestamps")

    profile = {
        "model": "large-v3",
        "device": "cpu",
        "compute_type": "int8",
        "beam_size": 5,
        "batch_size": 8,
        "batch_size_max": 32,
    }
    result = _decode_clips_batched(
        audio_paths=[audio_path],
        project={"language": "en"},
        profile=profile,
        runtime={
            "model": FakeModel(),
            "batched_model": FakeBatchedModel(),
            "device": "cpu",
            "compute_type": "int8",
            "model_name": "large-v3",
        },
    )

    assert not batched_calls
    assert len(direct_calls) == 1
    assert result[0]["transcript"] == "Complete long clip"
    assert result[0]["segments"][0]["end"] == pytest.approx(30.1)


def test_automatic_batch_size_uses_free_gpu_memory_conservatively() -> None:
    batch_size, source, memory = _automatic_batch_size(
        device="cuda",
        model_name="large-v3",
        compute_type="float16",
        memory_info={
            "index": 0,
            "name": "Test GPU",
            "total_mib": 16384,
            "free_mib": 10000,
        },
    )
    assert batch_size == 16
    assert "10000 MiB free" in source
    assert memory and memory["name"] == "Test GPU"

    assert _automatic_batch_size(
        device="cuda",
        model_name="large-v3",
        compute_type="float16",
        maximum=8,
        memory_info={
            "index": 0,
            "name": "Test GPU",
            "total_mib": 16384,
            "free_mib": 10000,
        },
    )[0] == 8
    assert _automatic_batch_size(
        device="cpu",
        model_name="large-v3",
        compute_type="int8",
    )[0] == 1


def test_recording_asr_auto_sizes_cpu_and_batch_changes_reuse_cache(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    audio_path = tmp_path / "source.wav"
    _write_tone(audio_path)
    line = {
        "line_id": "Actor::R2",
        "sheet": "Actor",
        "sheet_index": 0,
        "excel_row": 2,
        "line": "Hello.",
    }
    write_json(tmp_path / "source_lines.json", {"lines": [line]})
    write_json(
        tmp_path / "audio_inventory.json",
        {
            "files": [
                {
                    "path": str(audio_path.resolve()),
                    "sha256": "source-hash",
                }
            ]
        },
    )
    project = {
        "source_lines": "source_lines.json",
        "audio_inventory": "audio_inventory.json",
        "language": "en",
        "transcription": {
            "model": "large-v3",
            "device": "cpu",
            "compute_type": "int8",
        },
        "sessions": [
            {
                "id": "session",
                "enabled": True,
                "audio": "source.wav",
                "sheets": ["Actor"],
                "excel_rows": [],
                "line_ids": [],
            }
        ],
    }
    calls = []
    fake_segment = SimpleNamespace(
        id=1,
        start=0.0,
        end=0.5,
        text="Hello",
        avg_logprob=-0.1,
        no_speech_prob=0.01,
        words=[],
    )

    class FakePipeline:
        def __init__(self, model):
            self.model = model

        def transcribe(self, _audio, **kwargs):
            calls.append(kwargs)
            return iter([fake_segment]), SimpleNamespace(
                language="en",
                language_probability=1.0,
                duration=1.0,
                duration_after_vad=1.0,
            )

    import dialogue_pipeline.transcription as transcription_module
    import faster_whisper

    monkeypatch.setattr(
        transcription_module,
        "_make_model",
        lambda *_args, **_kwargs: (object(), "cpu", "int8"),
    )
    monkeypatch.setattr(
        faster_whisper,
        "BatchedInferencePipeline",
        FakePipeline,
    )
    outputs = transcribe_project(
        project_dir=tmp_path,
        project=project,
    )

    assert len(outputs) == 1
    assert calls[0]["batch_size"] == 1
    assert calls[0]["without_timestamps"] is False
    payload = read_json(outputs[0])
    assert payload["batched_inference"] is True
    assert payload["batch_size_requested"] == "auto"
    assert payload["batch_size"] == 1

    stable_cache_key = payload["cache_key"]
    project["transcription"]["batch_size"] = 8
    cached_outputs = transcribe_project(
        project_dir=tmp_path,
        project=project,
    )
    assert len(calls) == 1
    assert read_json(cached_outputs[0])["cache_key"] == stable_cache_key

    # A cache written by the previous schema is also reusable when only the
    # batch size changes.
    legacy_settings = dict(project["transcription"])
    legacy_settings.pop("batch_size")
    legacy_batch_size = 16
    payload.pop("cache_identity")
    payload.pop("batch_size_requested")
    payload.pop("batch_size_source")
    payload["batch_size"] = legacy_batch_size
    payload["cache_key"] = stable_hash(
        {
            "audio_sha256": "source-hash",
            "model": "large-v3",
            "device_request": "cpu",
            "compute_type_request": "int8",
            "language": "en",
            "batched_inference": True,
            "batch_size": legacy_batch_size,
            "settings": legacy_settings,
        }
    )
    write_json(outputs[0], payload)
    project["transcription"]["batch_size"] = 4
    transcribe_project(project_dir=tmp_path, project=project)
    assert len(calls) == 1


def test_prompted_segment_asr_is_evidence_but_cannot_verify_auto_ok(
    tmp_path: Path,
) -> None:
    segment_file = tmp_path / "segment.wav"
    _write_tone(segment_file)
    outputs = [
        SimpleNamespace(
            start=0.0,
            end=1.0,
            text="",
            avg_logprob=-1.0,
            no_speech_prob=0.8,
            words=[],
        ),
        SimpleNamespace(
            start=0.0,
            end=1.0,
            text="Yes",
            avg_logprob=-0.1,
            no_speech_prob=0.01,
            words=[
                SimpleNamespace(
                    start=0.1,
                    end=0.4,
                    word=" Yes",
                    probability=0.95,
                )
            ],
        ),
    ]

    class FakeModel:
        def transcribe(self, *_args, **_kwargs):
            return iter([outputs.pop(0)]), SimpleNamespace(language="en")

    line = {
        "line_id": "Actor::R2",
        "sheet": "Actor",
        "sheet_index": 0,
        "excel_row": 2,
        "line": "Yes?",
    }
    write_json(tmp_path / "source_lines.json", {"lines": [line]})
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "sessions": [
                {
                    "session_id": "session",
                    "segments": [
                        {
                            "segment_id": "session__s00001",
                            "kind": "base",
                            "file": "segment.wav",
                            "source_sha256": "source-hash",
                            "start_sample": 0,
                            "end_sample": 48000,
                            "start_seconds": 0.0,
                            "end_seconds": 1.0,
                            "transcript": "",
                            "words": [],
                            "word_count": 0,
                            "asr_probability": None,
                            "metrics": {"duration_seconds": 1.0},
                        }
                    ],
                    "derived_segments": [],
                }
            ]
        },
    )
    project = {
        "source_lines": "source_lines.json",
        "language": "en",
        "transcription": {
            "model": "large-v3",
            "device": "cpu",
            "compute_type": "int8",
        },
        "segment_transcription": {
            "enabled": True,
            "prompt_fallback_enabled": True,
        },
        "sessions": [
            {
                "id": "session",
                "enabled": True,
                "sheets": ["Actor"],
                "excel_rows": [],
                "line_ids": [],
            }
        ],
    }
    runtime = {"model": FakeModel()}
    transcribe_segments_project(
        project_dir=tmp_path,
        project=project,
        runtime=runtime,
    )
    segment = read_json(tmp_path / "segments_manifest.json")["sessions"][0][
        "segments"
    ][0]

    assert segment["transcript"] == "Yes"
    assert segment["transcript_source"] == "segment_asr_prompted"
    assert segment["segment_asr"]["primary"]["transcript"] == ""
    assert (
        segment["segment_asr"]["prompted_fallback"]["transcript"]
        == "Yes"
    )
    exact = transcribe_candidate_span(
        project_dir=tmp_path,
        project=project,
        segment=segment,
        runtime=runtime,
    )
    assert exact["transcript"] == ""


def test_merged_candidate_asr_decodes_the_exact_span_once(
    tmp_path: Path,
) -> None:
    segment_file = tmp_path / "merged.wav"
    _write_tone(segment_file, duration_seconds=2.0)
    calls = []
    fake_segment = SimpleNamespace(
        start=0.0,
        end=1.8,
        text="Hello there hello there",
        avg_logprob=-0.1,
        no_speech_prob=0.01,
        words=[],
    )

    class FakeModel:
        def transcribe(self, *_args, **kwargs):
            calls.append(kwargs)
            return iter([fake_segment]), SimpleNamespace(language="en")

    project = {
        "language": "en",
        "transcription": {
            "model": "large-v3",
            "device": "cpu",
            "compute_type": "int8",
        },
        "segment_transcription": {"enabled": True},
        "segmentation": {"fade_ms": 5.0},
        "export": {
            "sample_rate": 48000,
            "channels": 1,
            "bits_per_sample": 16,
        },
    }
    segment = {
        "segment_id": "session__m00001_00002",
        "kind": "merged",
        "file": "merged.wav",
        "source_sha256": "source-hash",
        "start_sample": 0,
        "end_sample": 96000,
    }
    result = transcribe_candidate_span(
        project_dir=tmp_path,
        project=project,
        segment=segment,
        runtime={"model": FakeModel()},
    )
    cached = transcribe_candidate_span(
        project_dir=tmp_path,
        project=project,
        segment=segment,
        runtime={},
    )

    assert result["transcript"] == "Hello there hello there"
    assert cached["cache_key"] == result["cache_key"]
    assert len(calls) == 1
    assert calls[0]["vad_filter"] is False
    assert calls[0]["condition_on_previous_text"] is False
    assert "initial_prompt" not in calls[0]


def test_candidate_span_asr_batches_unique_uncached_spans(
    tmp_path: Path,
) -> None:
    segments = []
    for index in range(3):
        segment_file = tmp_path / f"merged_{index}.wav"
        _write_tone(segment_file, duration_seconds=0.5)
        segments.append(
            {
                "segment_id": f"session__m{index + 1:05d}_{index + 2:05d}",
                "kind": "merged",
                "file": segment_file.name,
                "source_sha256": "source-hash",
                "start_sample": index * 24000,
                "end_sample": (index + 1) * 24000,
            }
        )
    project = {
        "language": "en",
        "transcription": {
            "model": "large-v3",
            "device": "cpu",
            "compute_type": "int8",
        },
        "segment_transcription": {
            "enabled": True,
            "batch_size": 2,
        },
        "segmentation": {"fade_ms": 5.0},
        "export": {
            "sample_rate": 48000,
            "channels": 1,
            "bits_per_sample": 16,
        },
    }
    batch_sizes = []

    class FakeBatchedModel:
        def transcribe(self, _audio, **kwargs):
            clips = kwargs["clip_timestamps"]
            batch_sizes.append(kwargs["batch_size"])
            decoded = []
            for index, clip in enumerate(clips):
                start = int(clip["start"] * 16000) / 16000
                decoded.append(
                    SimpleNamespace(
                        start=start + 0.05,
                        end=start + 0.35,
                        text=f"Line {index + 1}",
                        avg_logprob=-0.1,
                        no_speech_prob=0.01,
                        words=[
                            SimpleNamespace(
                                start=start + 0.05,
                                end=start + 0.25,
                                word=" Line",
                                probability=0.95,
                            )
                        ],
                    )
                )
            return iter(decoded), SimpleNamespace(language="en")

    runtime = {
        "model": SimpleNamespace(
            feature_extractor=SimpleNamespace(sampling_rate=16000)
        ),
        "batched_model": FakeBatchedModel(),
    }
    results = transcribe_candidate_spans(
        project_dir=tmp_path,
        project=project,
        segments=[segments[0], segments[1], segments[0], segments[2]],
        runtime=runtime,
    )
    project["segment_transcription"]["batch_size"] = 8
    cached = transcribe_candidate_spans(
        project_dir=tmp_path,
        project=project,
        segments=segments,
        runtime={},
    )

    assert batch_sizes == [2, 1]
    assert set(results) == {segment["segment_id"] for segment in segments}
    assert {
        segment_id: payload["cache_key"]
        for segment_id, payload in cached.items()
    } == {
        segment_id: payload["cache_key"]
        for segment_id, payload in results.items()
    }


def test_sample_accurate_cut_and_finalize(tmp_path: Path) -> None:
    source = tmp_path / "source.wav"
    segment_file = tmp_path / "segments" / "session" / "session__s00001.wav"
    _write_tone(source)
    metrics = cut_pcm_wav(
        source,
        segment_file,
        start_sample=4800,
        end_sample=14400,
        fade_ms=5.0,
    )
    assert metrics["frame_count"] == 9600
    assert metrics["sample_rate"] == 48000

    line = {
        "line_id": "Sheet::R3",
        "sheet": "Sheet",
        "sheet_index": 0,
        "excel_row": 3,
        "quest": "Quest",
        "context": "Context",
        "line": "Hello there.",
        "acting_note": "",
        "emotion": "",
        "target_filename": "TargetFile_00000001_1",
    }
    candidate = {
        "line_id": line["line_id"],
        "session_id": "session",
        "segment_id": "session__s00001",
        "segment_file": segment_file.relative_to(tmp_path).as_posix(),
        "transcript": "Hello there.",
        "match_score": 100.0,
        "ordered_similarity": 100.0,
        "token_coverage": 1.0,
        "token_precision": 1.0,
        "extra_word_count": 0,
        "clause_count": 1,
        "minimum_clause_score": 100.0,
        "missing_clause_count": 0,
        "fragment_join": False,
        "fragment_source_count": 0,
        "confidence_margin": 50.0,
        "technical_score": 100.0,
        "selection_score": 115.0,
        "reliable": True,
        "reliability_reason": "",
        "source_audio": "source.wav",
        "start_seconds": 0.1,
        "end_seconds": 0.3,
        "duration_seconds": 0.2,
        "asr_probability": 0.99,
        "base_indices": [0],
        "rank": 1,
    }
    review_path = tmp_path / "line_review.json"
    review_data = build_line_review(
        source_lines=[line],
        candidates_by_line={line["line_id"]: [candidate]},
        unmatched_segments=[
            {
                "segment_id": candidate["segment_id"],
                "segment_file": candidate["segment_file"],
                "source_wav": candidate["source_audio"],
                "start_seconds": candidate["start_seconds"],
                "end_seconds": candidate["end_seconds"],
                "duration_seconds": candidate["duration_seconds"],
                "transcript": "Unmatched introduction",
                "asr_confidence": candidate["asr_probability"],
                "reason": "NO_RELIABLE_MATCH",
                "suggested_line_1": line["line_id"],
                "suggested_line_1_text": line["line"],
                "suggested_line_1_score": 40.0,
                "suggested_line_2": "",
                "suggested_line_2_text": "",
                "suggested_line_2_score": 0.0,
                "technical_flags": "",
                "technical_score": 80.0,
                "audible": True,
            }
        ],
    )
    save_line_review(review_path, review_data)
    loaded_review = load_line_review(review_path)
    assert loaded_review["lines"][0]["type"] == "normal"
    assert loaded_review["lines"][0]["status"] == "AUTO_OK"
    assert loaded_review["lines"][0]["suggested_segment_id"] == (
        "session__s00001"
    )
    assert loaded_review["lines"][0]["selected_segment_id"] == (
        "session__s00001"
    )
    assert loaded_review["lines"][0]["candidates"][0]["score"] == 100.0
    assert loaded_review["unmatched_segments"] == []

    manifest = {
        "schema_version": 1,
        "sessions": [
            {
                "session_id": "session",
                "segments": [
                    {
                        "segment_id": "session__s00001",
                        "file": segment_file.relative_to(tmp_path).as_posix(),
                        "source_audio": "source.wav",
                        "start_seconds": 0.1,
                        "end_seconds": 0.3,
                        "transcript": "Hello there.",
                    }
                ],
                "derived_segments": [],
            }
        ],
    }
    write_json(tmp_path / "segments_manifest.json", manifest)
    project = {
        "export": {
            "extension": ".wav",
            "sample_rate": 48000,
            "channels": 1,
            "bits_per_sample": 16,
        }
    }
    output_dir = tmp_path / "final"
    result = finalize_review(
        project_dir=tmp_path,
        project=project,
        review_path=review_path,
        output_dir=output_dir,
    )
    assert result["error_count"] == 0
    assert (output_dir / "TargetFile_00000001_1.wav").is_file()


def test_finalize_can_optionally_reuse_one_segment(tmp_path: Path) -> None:
    segment_file = tmp_path / "segment.wav"
    _write_tone(segment_file)
    lines = [
        {
            "line_id": f"Sheet::R{row}",
            "sheet": "Sheet",
            "sheet_index": 0,
            "excel_row": row,
            "quest": "",
            "context": "",
            "line": "Found you!",
            "acting_note": "",
            "emotion": "",
            "target_filename": f"target_{row}",
        }
        for row in (3, 4)
    ]
    candidates = {}
    for line in lines:
        candidates[line["line_id"]] = [
            {
                "line_id": line["line_id"],
                "session_id": "session",
                "segment_id": "session__s00001",
                "segment_file": "segment.wav",
                "transcript": "Found you",
                "match_score": 100.0,
                "confidence_margin": 0.0,
                "technical_score": 100.0,
                "selection_score": 115.0,
                "reliable": True,
                "reliability_reason": "",
                "source_audio": "source.wav",
                "start_seconds": 0.0,
                "end_seconds": 1.0,
                "duration_seconds": 1.0,
                "asr_probability": 0.99,
                "base_indices": [0],
                "rank": 1,
            }
        ]
    review_path = tmp_path / "line_review.json"
    save_line_review(
        review_path,
        build_line_review(
            source_lines=lines,
            candidates_by_line=candidates,
            unmatched_segments=[],
        ),
    )
    write_json(
        tmp_path / "segments_manifest.json",
        {
            "sessions": [
                {
                    "session_id": "session",
                    "segments": [
                        {
                            "segment_id": "session__s00001",
                            "file": "segment.wav",
                            "source_audio": "source.wav",
                            "start_seconds": 0.0,
                            "end_seconds": 1.0,
                            "transcript": "Found you",
                        }
                    ],
                    "derived_segments": [],
                }
            ]
        },
    )
    project = {
        "export": {
            "extension": ".wav",
            "sample_rate": 48000,
            "channels": 1,
            "bits_per_sample": 16,
        }
    }

    with pytest.raises(ValueError, match="Finalization stopped"):
        finalize_review(
            project_dir=tmp_path,
            project=project,
            review_path=review_path,
            output_dir=tmp_path / "blocked",
            dry_run=True,
        )
    result = finalize_review(
        project_dir=tmp_path,
        project=project,
        review_path=review_path,
        output_dir=tmp_path / "allowed",
        allow_segment_reuse=True,
        dry_run=True,
    )

    assert result["error_count"] == 0
    assert result["export_count"] == 2
    assert result["allow_segment_reuse"] is True


def test_segmentation_and_alignment_integration(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import dialogue_pipeline.segmentation as segmentation_module

    def fake_voice_regions(
        _path: Path,
        *,
        start_sample: int,
        end_sample: int,
        threshold: float,
    ) -> list[tuple[int, int]]:
        inset = 2400 if threshold < 0.7 else 3600
        return [(start_sample + inset, end_sample - inset)]

    monkeypatch.setattr(
        segmentation_module,
        "pcm_voice_regions",
        fake_voice_regions,
    )
    source = tmp_path / "source.wav"
    _write_pattern(source)
    source_hash = sha256_file(source)
    source_lines = {
        "schema_version": 1,
        "line_count": 2,
        "sheet_count": 1,
        "sheets": [
            {
                "name": "Actor",
                "index": 0,
                "voice_header": "Actor",
                "line_ids": ["Actor::R3", "Actor::R4"],
                "line_count": 2,
            }
        ],
        "lines": [
            {
                "line_id": "Actor::R3",
                "sheet": "Actor",
                "sheet_index": 0,
                "excel_row": 3,
                "quest": "",
                "context": "",
                "line": "Hello there",
                "acting_note": "",
                "emotion": "",
                "target_filename": "hello_target",
            },
            {
                "line_id": "Actor::R4",
                "sheet": "Actor",
                "sheet_index": 0,
                "excel_row": 4,
                "quest": "",
                "context": "",
                "line": "Goodbye friend",
                "acting_note": "",
                "emotion": "",
                "target_filename": "goodbye_target",
            },
        ],
    }
    write_json(tmp_path / "source_lines.json", source_lines)
    write_json(
        tmp_path / "audio_inventory.json",
        {
            "files": [
                {
                    "path": str(source.resolve()),
                    "sha256": source_hash,
                    "duration_seconds": 5.0,
                    "sample_rate": 48000,
                    "channels": 1,
                    "bits_per_sample": 16,
                }
            ]
        },
    )
    project = {
        "source_lines": "source_lines.json",
        "audio_inventory": "audio_inventory.json",
        "language": "en",
        "segmentation": {
            "silence_noise_db": -45.0,
            "silence_detection_min_seconds": 0.2,
            "split_gap_seconds": 0.4,
            "minimum_segment_seconds": 0.1,
            "pre_padding_seconds": 0.05,
            "post_padding_seconds": 0.05,
            "fade_ms": 5.0,
        },
        "alignment": {
            "span_search": {
                "max_segments": 2,
                "max_gap_seconds": 1.0,
                "max_duration_seconds": 10.0,
                "minimum_score": 45.0,
            },
            "reliability": {
                "normal": {
                    "minimum_score": 72.0,
                    "minimum_margin": 8.0,
                },
                "short": {
                    "minimum_score": 88.0,
                    "minimum_margin": 15.0,
                },
            },
            "ranking": {"noise_penalty": 2.2},
        },
        "export": {
            "extension": ".wav",
            "sample_rate": 48000,
            "channels": 1,
            "bits_per_sample": 16,
        },
        "sessions": [
            {
                "id": "session",
                "enabled": True,
                "audio": "source.wav",
                "sheets": ["Actor"],
                "excel_rows": [],
                "line_ids": [],
            }
        ],
    }
    write_json(
        tmp_path / "transcripts" / "session.json",
        {
            "schema_version": 1,
            "session_id": "session",
            "audio": "source.wav",
            "audio_sha256": source_hash,
            "cache_key": "synthetic",
            "segments": [
                {
                    "id": 0,
                    "start": 0.25,
                    "end": 0.9,
                    "text": "Hello there",
                    "words": [
                        {
                            "start": 0.25,
                            "end": 0.5,
                            "word": " Hello",
                            "probability": 0.99,
                        },
                        {
                            "start": 0.55,
                            "end": 0.9,
                            "word": " there",
                            "probability": 0.99,
                        },
                    ],
                },
                {
                    "id": 1,
                    "start": 1.4,
                    "end": 2.05,
                    "text": "Hello there",
                    "words": [
                        {
                            "start": 1.4,
                            "end": 1.65,
                            "word": " Hello",
                            "probability": 0.98,
                        },
                        {
                            "start": 1.7,
                            "end": 2.05,
                            "word": " there",
                            "probability": 0.98,
                        },
                    ],
                },
                {
                    "id": 2,
                    "start": 3.0,
                    "end": 3.75,
                    "text": "Goodbye friend",
                    "words": [
                        {
                            "start": 3.0,
                            "end": 3.35,
                            "word": " Goodbye",
                            "probability": 0.99,
                        },
                        {
                            "start": 3.4,
                            "end": 3.75,
                            "word": " friend",
                            "probability": 0.99,
                        },
                    ],
                },
            ],
        },
    )
    manifest_path = segment_project(project_dir=tmp_path, project=project)
    assert manifest_path.is_file()
    manifest = read_json(manifest_path)
    manifest_session = manifest["sessions"][0]
    assert manifest_session["voice_boundary_detection"] == {
        "enabled": True,
        "vad_threshold": 0.5,
        "breath_vad_threshold": 0.7,
    }
    assert all(
        segment["voice_bounds"]["speech"] is not None
        and segment["voice_bounds"]["strict_speech"] is not None
        for segment in manifest_session["segments"]
    )
    stale_unmatched = tmp_path / "B_unmatched_segments.tsv"
    stale_unmatched.write_text("stale", encoding="utf-8")
    outputs = align_project(
        project_dir=tmp_path,
        project=project,
        session_filter={"session"},
    )
    assert outputs["review"].is_file()
    assert outputs["review"].name == "line_review.json"
    assert "unmatched" not in outputs
    assert not stale_unmatched.exists()
    review = load_line_review(outputs["review"])
    assert review["lines"][0]["selected_segment_id"]
    assert review["lines"][1]["selected_segment_id"]
    assert review["lines"][0]["candidates"]
    assert "unmatched_segments" in review
